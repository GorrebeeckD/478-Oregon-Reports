import os
import pyodbc  
import pandas as pd
from datetime import datetime, timedelta
from dateutil.relativedelta import relativedelta
from openpyxl import Workbook
from openpyxl.styles import Border, Side, Alignment, Font
from openpyxl.styles import NamedStyle
import win32com.client as win32

# Database connection details (update these)
DB_DRIVER = "{IBM DB2 ODBC DRIVER}"  
DB_SERVER = "10.30.207.182"
DB_DATABASE = "ARRPROD"
DB_PORT = "50001"  # Default DB2 port, update if different

# Generate current date in different formats
current_date_mmddyyyy = datetime.now().strftime("%m%d%Y")
current_date_yyyymmdd = datetime.now().strftime("%Y%m%d")

# Get the user's downloads folder path
shared_folder = r"\\hmsfs\general\General\Client Folders\ACCOUNT DELIVERY SOLUTIONS - JUDE\Reporting\478-Oregon\Monthly Billing Report"

# Define file paths in the downloads folder
ffs_excel_filename = os.path.join(shared_folder, f"FFS Monthly Billing Report_{current_date_mmddyyyy}.xlsx")
cco_excel_filename = os.path.join(shared_folder, f"CCO Monthly Billing Report_{current_date_mmddyyyy}.xlsx")
ffs_txt_filename = os.path.join(shared_folder, f"ORMTHLY.MTHLYBILLFFS.{current_date_yyyymmdd}.txt")
cco_txt_filename = os.path.join(shared_folder, f"ORMTHLY.MTHLYBILLCCO.{current_date_yyyymmdd}.txt")
ffs_denied_excel_filename = os.path.join(shared_folder, f"FFS Denied_{current_date_yyyymmdd}.xlsx")
cco_denied_excel_filename = os.path.join(shared_folder, f"CCO Denied_{current_date_yyyymmdd}.xlsx")
ffs_denied_txt_filename = os.path.join(shared_folder, f"ORMTHLY.ORDENIALSOUTFFS.{current_date_yyyymmdd}.txt")
cco_denied_txt_filename = os.path.join(shared_folder, f"ORMTHLY.ORDENIALSOUTCCO.{current_date_yyyymmdd}.txt")

# Calculate the first and last day of the previous month
first_day_prev_month = (datetime.now().replace(day=1) - relativedelta(months=1)).strftime("%Y-%m-%d")
last_day_prev_month = (datetime.now().replace(day=1) - timedelta(days=1)).strftime("%Y-%m-%d")

# Connection string using Windows Authentication
conn_str = (
    f"DRIVER={DB_DRIVER};"
    f"DATABASE={DB_DATABASE};"
    f"HOSTNAME={DB_SERVER};"
    f"PORT={DB_PORT};"
    f"PROTOCOL=TCPIP;"
    f"Trusted_Connection=yes;"
)

# Function to add borders to all cells in a worksheet
def add_borders(ws):
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    for row in ws.iter_rows():
        for cell in row:
            cell.border = thin_border

# Function to format columns as dates or currency
# Function to format columns as dates or currency and apply font to all cells
def format_columns_and_apply_font(ws):
    date_columns = ['G', 'H', 'L', 'P', 'Q']
    currency_columns = ['I', 'J', 'K', 'M']
    
    date_style = NamedStyle(name="date_style", number_format="MM/DD/YYYY")
    currency_style = NamedStyle(name="currency_style", number_format="$#,##0.00")
    
    for col in date_columns:
        for cell in ws[col]:
            cell.style = date_style
            cell.font = Font(name="Aptos Narrow")
            cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    
    for col in currency_columns:
        for cell in ws[col]:
            cell.style = currency_style
            cell.font = Font(name="Aptos Narrow")
            cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    
    # Apply font to all cells
    for row in ws.iter_rows():
        for cell in row:
            cell.font = Font(name="Aptos Narrow")
            cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))         

# Establish toad connection
try:
    connection = pyodbc.connect(conn_str)
    cursor = connection.cursor()
    
    # OPEN FFS CLAIMS
    query1 = f"""
    --FFS SQL
    ---Updated 20241103 AZN - CUS0176411 - Remove Centene exchange data from all GW reporting
    SELECT 
        "TMP_TBL"."ARSEQ" AS "ARSEQ",
        "TMP_TBL"."CLIENT LAST NAME" AS "CLIENT LAST NAME",
        "TMP_TBL"."CLIENT FIRST NAME" AS "CLIENT FIRST NAME",
        "TMP_TBL"."RECIPIENT ID" AS "RECIPIENT ID",
        "TMP_TBL"."ICN" AS "ICN",
        "TMP_TBL"."ICN_DETAIL_LINE" AS "ICN_DETAIL_LINE",
        "TMP_TBL"."FROM DOS" AS "FROM DOS",
        "TMP_TBL"."THRU DOS" AS "THRU DOS",
        "TMP_TBL"."BILLED AMOUNT" AS "BILLED AMOUNT",
        "TMP_TBL"."LINE MEDICAID PAID" AS "LINE MEDICAID PAID",
        "TMP_TBL"."MEDICAID PAID" AS "MEDICAID PAID",
        "TMP_TBL"."MEDICAID PAID DATE" AS "MEDICAID PAID DATE",
        "TMP_TBL"."AMOUNT RECOVERED" AS "AMOUNT RECOVERED",
        "TMP_TBL"."HMS CARRIER CODE" AS "HMS CARRIER CODE",
        "TMP_TBL"."HMS CARRIER NAME" AS "HMS CARRIER NAME",
        "TMP_TBL"."ORIGINAL BILL DATE" AS "ORIGINAL BILL DATE",
        "TMP_TBL"."REBILL DATE" AS "REBILL DATE",
        "TMP_TBL"."STATUS" AS "STATUS",
        "TMP_TBL"."SOURCE CODE" AS "SOURCE CODE"
    FROM  (     SELECT 
                AR_SEQ_NUM AS "ARSEQ",
                    LAST_NM AS "CLIENT LAST NAME",
                    FIRST_NM AS "CLIENT FIRST NAME",
                    MA_NUM AS "RECIPIENT ID",
                    ICN AS "ICN",
                    ICN_DETAIL_LINE AS  "ICN_DETAIL_LINE",
                    FROM_DOS_DT AS "FROM DOS",
                    THRU_DOS_DT AS "THRU DOS",
                    BILL_AMT AS "BILLED AMOUNT",
                    MA_PAID_AMT AS "LINE MEDICAID PAID",
            MA_PAID_HDR_AMT AS "MEDICAID PAID",
                    MA_PAID_DT AS "MEDICAID PAID DATE",
            SUM(REMIT_AMT) AS "AMOUNT RECOVERED",
            CARRIER_CD AS "HMS CARRIER CODE",
            CARRIER_NM AS "HMS CARRIER NAME",
            BILL_DT AS "ORIGINAL BILL DATE",
            REBILL_DT AS "REBILL DATE",
            CLAIM_STATUS AS  "STATUS",
                    SOURCE_CODE AS "SOURCE CODE"

              FROM (
              SELECT 
                CLAIMS.AR_SEQ_NUM,
                    CLAIMS.LAST_NM,
                    CLAIMS.FIRST_NM,
                    CLAIMS.MA_NUM,
                    SUBSTR(CLAIMS.ICN_NUM,1,LENGTH(CLAIMS.ICN_NUM)-2) AS ICN,
                    RIGHT(CLAIMS.ICN_NUM,2) AS  ICN_DETAIL_LINE,
                    CLAIMS.FROM_DOS_DT,
                    CLAIMS.THRU_DOS_DT,
                    CLAIMS.BILL_AMT,
                    CLAIMS.MA_PAID_AMT, 
            CLAIMCONT.MA_PAID_HDR_AMT, 
                    CLAIMCONT.MA_PAID_DT, 
            CASE WHEN POSTING_REPORTS.REMIT_AMT IS NULL THEN 0.00 ELSE POSTING_REPORTS.REMIT_AMT END AS REMIT_AMT,
            CLAIMS.CARRIER_CD,
            ARTCARM.CARRIER_NM, 
            CLAIMS.BILL_DT,
            CASE WHEN CLAIMS.REBILL_DT IS NOT NULL THEN CLAIMS.REBILL_DT ELSE CLAIMS.BILL_DT END AS REBILL_DT,
            CASE 
            WHEN CLAIMS.TRNST_RF IN ('OPEN', 'DENIED', 'REVERSED') THEN
                CLAIMS.TRNST_RF
            WHEN CLAIMS.TRNST_RF IN ('PAID-PD', 'PAID-PP', 'PAID-OP', 'PAID-EP') THEN
                'PAID'
            ELSE
                NULL
            END  AS CLAIM_STATUS,
                    CASE 
                      WHEN CLAIMS.ORIG_SRCE_ELIG_CD = 'RS' THEN 
              'RSC'                 
                      ELSE 
              'TPL'                 
                    END AS SOURCE_CODE
                FROM AR.CLAIMS CLAIMS 
             JOIN AR.CLAIM_CONTROLLER CLAIMCONT ON CLAIMS.CLM_CONTROLLER_ID = CLAIMCONT.CLM_CONTROLLER_ID 
              LEFT JOIN AR.POSTING_REPORTS  POSTING_REPORTS ON CLAIMS.CLAIM_ID = POSTING_REPORTS.CLAIM_ID  AND POSTING_REPORTS.DELETE_IND = 'N'
                      LEFT OUTER JOIN CAR.ARTCARM_BASE ARTCARM ON ARTCARM.CARRIER_CD = CLAIMS.CARRIER_CD AND ARTCARM.DATA_SOURCE_CD = 'GCS' AND ARTCARM.CARRIER_OFFICE_CD='0000'
                WHERE CLAIMS.DELETE_IND='N' 
          AND CLAIMS.CONTRACT_NUM = '478' 
          AND CLAIMS.PROJECT_CD >= '30' 
          AND (SUBSTR(CLAIMS.ICN_NUM,1,2) NOT IN('60','70')) 
	      AND CLAIMS.BILL_TYPE_CD <> 'SP'
          AND CLAIMS.BILL_DT BETWEEN '{first_day_prev_month}' AND '{last_day_prev_month}'
          AND CLAIMS.REBILL_DT IS NULL
          AND TRIM(claims.carrier_cd) <> 'CNEXC' ---Updated 20241103 AZN - CUS0176411 - Remove Centene exchange data from all GW reporting
            )

              GROUP BY
                  AR_SEQ_NUM,
                  LAST_NM,
                  FIRST_NM,
                  MA_NUM,
                  ICN,
                  ICN_DETAIL_LINE,
          FROM_DOS_DT,
                  THRU_DOS_DT,
                  BILL_AMT,
          MA_PAID_AMT,
          MA_PAID_HDR_AMT,
          MA_PAID_DT,
          CARRIER_CD,
          CARRIER_NM,
          BILL_DT,
          REBILL_DT,
          CLAIM_STATUS,
          SOURCE_CODE


    ) "TMP_TBL"
    
    Where TMP_TBL.STATUS = 'OPEN'
    WITH UR;

    """
    cursor.execute(query1)
    columns1 = [column[0] for column in cursor.description]
    results1 = cursor.fetchall()
    results1 = [list(row) for row in results1] 
    print("Results from FFS OPEN query:")
    for row in results1:
        print(row)

    # Export FFS Open query results to Excel and TXT
    df1 = pd.DataFrame(results1, columns=columns1)

    # Save to Excel with borders and without bold column names
    with pd.ExcelWriter(ffs_excel_filename, engine='openpyxl') as writer:
        df1.to_excel(writer, index=False)
        workbook  = writer.book
        worksheet = writer.sheets['Sheet1']
        add_borders(worksheet)
        
        # Align column names to the left and remove bold formatting
        for cell in worksheet["1:1"]:
            cell.alignment = Alignment(horizontal="left")
            cell.font = Font(bold=False, name="Aptos Narrow")
        
        # Format columns as dates or currency and apply font to all cells
        format_columns_and_apply_font(worksheet)

    # Save to TXT with commas and quotes around each row
    with open(ffs_txt_filename, 'w') as f:
        f.write('"' + '","'.join(columns1) + '"\n')
        for row in results1:
            f.write('"' + '","'.join(map(str, row)) + '"\n')

    print("FFS Open Files Successfully Exported")
    
    # OPEN CCO CLAIMS
    query2 = f"""
    --CCO SQL
    ---Updated 20241103 AZN - CUS0176411 - Remove Centene exchange data from all GW reporting
    SELECT 
        "TMP_TBL"."ARSEQ" AS "ARSEQ",
        "TMP_TBL"."CLIENT LAST NAME" AS "CLIENT LAST NAME",
        "TMP_TBL"."CLIENT FIRST NAME" AS "CLIENT FIRST NAME",
        "TMP_TBL"."RECiPIENT ID" AS "RECiPIENT ID",
        "TMP_TBL"."ICN" AS "ICN",
        "TMP_TBL"."ICN_DETAIL_LINE" AS "ICN_DETAIL_LINE",
        "TMP_TBL"."FROM DOS" AS "FROM DOS",
        "TMP_TBL"."THRU DOS" AS "THRU DOS",
        "TMP_TBL"."BILLED AMOUNT" AS "BILLED AMOUNT",
        "TMP_TBL"."LINE MEDICAID PAID" AS "LINE MEDICAID PAID",
        "TMP_TBL"."MEDICAID PAID" AS "MEDICAID PAID",
        "TMP_TBL"."MEDICAID PAID DATE" AS "MEDICAID PAID DATE",
        "TMP_TBL"."AMOUNT RECOVERED" AS "AMOUNT RECOVERED",
        "TMP_TBL"."HMS CARRIER CODE" AS "HMS CARRIER CODE",
        "TMP_TBL"."HMS CARRIER NAME" AS "HMS CARRIER NAME",
        "TMP_TBL"."ORIGINAL BILL DATE" AS "ORIGINAL BILL DATE",
        "TMP_TBL"."REBILL DATE" AS "REBILL DATE",
        "TMP_TBL"."STATUS" AS "STATUS",
        "TMP_TBL"."SOURCE CODE" AS "SOURCE CODE"
    FROM  (               SELECT 
                AR_SEQ_NUM AS "ARSEQ",
                    LAST_NM AS "CLIENT LAST NAME",
                    FIRST_NM AS "CLIENT FIRST NAME",
                    MA_NUM AS "RECiPIENT ID",
                    ICN AS "ICN",
                    ICN_DETAIL_LINE AS  "ICN_DETAIL_LINE",
                    FROM_DOS_DT AS "FROM DOS",
                    THRU_DOS_DT AS "THRU DOS",
                    BILL_AMT AS "BILLED AMOUNT",
                    MA_PAID_AMT AS "LINE MEDICAID PAID",
            MA_PAID_HDR_AMT AS "MEDICAID PAID",
                    MA_PAID_DT AS "MEDICAID PAID DATE",
            SUM(REMIT_AMT) AS "AMOUNT RECOVERED",
            CARRIER_CD AS "HMS CARRIER CODE",
            CARRIER_NM AS "HMS CARRIER NAME",
            BILL_DT AS "ORIGINAL BILL DATE",
            REBILL_DT AS "REBILL DATE",
            CLAIM_STATUS AS  "STATUS",
                    SOURCE_CODE AS "SOURCE CODE"

              FROM (
              SELECT 
                CLAIMS.AR_SEQ_NUM,
                    CLAIMS.LAST_NM,
                    CLAIMS.FIRST_NM,
                    CLAIMS.MA_NUM,
                    SUBSTR(CLAIMS.ICN_NUM,1,LENGTH(CLAIMS.ICN_NUM)-2) AS ICN,
                    RIGHT(CLAIMS.ICN_NUM,2) AS  ICN_DETAIL_LINE,
                    CLAIMS.FROM_DOS_DT,
                    CLAIMS.THRU_DOS_DT,
                    CLAIMS.BILL_AMT,
                    CLAIMS.MA_PAID_AMT, 
            CLAIMCONT.MA_PAID_HDR_AMT, 
                    CLAIMCONT.MA_PAID_DT, 
            CASE WHEN POSTING_REPORTS.REMIT_AMT IS NULL THEN 0.00 ELSE POSTING_REPORTS.REMIT_AMT END AS REMIT_AMT,
            CLAIMS.CARRIER_CD,
            ARTCARM.CARRIER_NM, 
            CLAIMS.BILL_DT,
            CASE WHEN CLAIMS.REBILL_DT IS NOT NULL THEN CLAIMS.REBILL_DT ELSE CLAIMS.BILL_DT END AS REBILL_DT,
            CASE 
            WHEN CLAIMS.TRNST_RF IN ('OPEN', 'DENIED', 'REVERSED') THEN
                CLAIMS.TRNST_RF
            WHEN CLAIMS.TRNST_RF IN ('PAID-PD', 'PAID-PP', 'PAID-OP', 'PAID-EP') THEN
                'PAID'
            ELSE
                NULL
            END  AS CLAIM_STATUS,
                    CASE 
                      WHEN CLAIMS.ORIG_SRCE_ELIG_CD = 'RS' THEN 
              'RSC'                 
                      ELSE 
              'TPL'                 
                    END AS SOURCE_CODE
                FROM AR.CLAIMS CLAIMS 
             JOIN AR.CLAIM_CONTROLLER CLAIMCONT ON CLAIMS.CLM_CONTROLLER_ID = CLAIMCONT.CLM_CONTROLLER_ID 
              LEFT JOIN AR.POSTING_REPORTS  POSTING_REPORTS ON CLAIMS.CLAIM_ID = POSTING_REPORTS.CLAIM_ID  AND POSTING_REPORTS.DELETE_IND = 'N'
                      LEFT OUTER JOIN CAR.ARTCARM_BASE ARTCARM ON ARTCARM.CARRIER_CD = CLAIMS.CARRIER_CD AND ARTCARM.DATA_SOURCE_CD = 'GCS' AND ARTCARM.CARRIER_OFFICE_CD='0000'
                WHERE CLAIMS.DELETE_IND='N' 
          AND CLAIMS.CONTRACT_NUM = '478' 
          AND CLAIMS.PROJECT_CD >= '30' 
          AND (SUBSTR(CLAIMS.ICN_NUM,1,2) IN('60','70')) 
	      AND CLAIMS.BILL_TYPE_CD <> 'SP'
          AND CLAIMS.BILL_DT BETWEEN '{first_day_prev_month}' AND '{last_day_prev_month}'
          AND CLAIMS.REBILL_DT IS NULL 
          AND TRIM(claims.carrier_cd) <> 'CNEXC' ---Updated 20241103 AZN - CUS0176411 - Remove Centene exchange data from all GW reporting
          )

              GROUP BY
                  AR_SEQ_NUM,
                  LAST_NM,
                  FIRST_NM,
                  MA_NUM,
                  ICN,
                  ICN_DETAIL_LINE,
          FROM_DOS_DT,
                  THRU_DOS_DT,
                  BILL_AMT,
          MA_PAID_AMT,
          MA_PAID_HDR_AMT,
          MA_PAID_DT,
          CARRIER_CD,
          CARRIER_NM,
          BILL_DT,
          REBILL_DT,
          CLAIM_STATUS,
          SOURCE_CODE


    ) "TMP_TBL"
    Where TMP_TBL.STATUS = 'OPEN'
    WITH UR;

    """
    cursor.execute(query2)
    columns2 = [column[0] for column in cursor.description]
    results2 = cursor.fetchall()  
    results2 = [list(row) for row in results2]  
    print("Results from CCO Open query:")
    for row in results2:
        print(row)

    # Export second query results to Excel and TXT
    df2 = pd.DataFrame(results2, columns=columns2)

    # Save to Excel with borders and without bold column names
    with pd.ExcelWriter(cco_excel_filename, engine='openpyxl') as writer:
        df2.to_excel(writer, index=False)
        workbook  = writer.book
        worksheet = writer.sheets['Sheet1']
        add_borders(worksheet)
        
        # Align column names to the left and remove bold formatting
        for cell in worksheet["1:1"]:
            cell.alignment = Alignment(horizontal="left")
            cell.font = Font(bold=False, name="Aptos Narrow")
        
        # Format columns as dates or currency and apply font to all cells
        format_columns_and_apply_font(worksheet)
    
    # Save to TXT with commas and quotes around each row
    with open(cco_txt_filename, 'w') as f:
        f.write('"' + '","'.join(columns2) + '"\n')
        for row in results2:
            f.write('"' + '","'.join(map(str, row)) + '"\n')

    print("CCO OPEN Files Successfully Exported")
    
except pyodbc.DatabaseError as e:
    print("Database connection error:", e)
finally:
    if cursor:
        cursor.close()
    if connection:
        connection.close()
    print("Connection closed.")

try:
    connection = pyodbc.connect(DSN='EDWPROD')
    cursor = connection.cursor()
    
    # FFS DENIED CLAIMS
    teradata_query1 = f"""
    SELECT DISTINCT
            A.AR_SEQ_NUM AS "ARSEQ",
            A.LAST_NM AS "CLIENT LAST NAME",
            A.FIRST_NM AS "CLIENT FIRST NAME",
            A.MA_NUM AS "RECIPIENT ID", 
            SUBSTR(A.ICN_NUM,1,LENGTH(ICN_NUM)-2) AS ICN,
            RIGHT(A.ICN_NUM,2) AS "ICN DETAIL LINE",  
            TO_CHAR(A.FROM_DOS_DT, 'MM/DD/YY') AS "FROM DOS",
            TO_CHAR(A.THRU_DOS_DT, 'MM/DD/YY') AS "THRU DOS",
            A.BILL_AMT AS "BILLED AMOUNT",
            A.MA_PAID_AMT AS "LINE MEDICAID PAID",
            A.MED_HDRPD_AMT AS "MEDICAID PAID",
            TO_CHAR(A.MA_PAID_DT, 'MM/DD/YY') AS "MEDICAID PAID DATE",
            SUM(A.REMIT_AMT) AS "ACTUAL AMOUNT RECOVERED",
            C.ACTCD_RF1 AS "HMS ACTION CD",
            CASE WHEN A.ACTION_CD = 'ADDR' THEN 'Submit Claim to a Different Address'
            WHEN A.ACTION_CD = 'ALCARE' THEN 'Alternate Services were Available and Should Have Been Utilized'
            WHEN A.ACTION_CD = 'BCLCL   ' THEN 'File Claim with Local BCBS Plan'
            WHEN A.ACTION_CD = 'BPNPI' THEN 'Billing provider NPI'
            WHEN A.ACTION_CD = 'BPTAX' THEN 'Invalid billing provider Tax ID'
            WHEN A.ACTION_CD = 'BPTXN' THEN 'Missing/invalid billing provider taxonomy ID'
            WHEN A.ACTION_CD = 'CCLM' THEN 'Corrected Claim Needed'
            WHEN A.ACTION_CD = 'CLAMAJ' THEN 'Care may not be Covered by Another Payer per Coordination of Benefits'
            WHEN A.ACTION_CD = 'CLMFWD' THEN 'Claim Forwarded to Payer by NEIC'
            WHEN A.ACTION_CD = 'COBPD' THEN 'Paid in Accordance with COB'
            WHEN A.ACTION_CD = 'CPT' THEN 'CPT Code Missing'
            WHEN A.ACTION_CD = 'DUP' THEN 'Duplicate Claim Submission'
            WHEN A.ACTION_CD = 'DUPHMS' THEN 'Duplicate to Medicaid Claim'
            WHEN A.ACTION_CD = 'ELEC' THEN 'Submit Claims Electronically'
            WHEN A.ACTION_CD = 'ERRDES' THEN 'No description/definition of the denial code found on the EOB in DocDNA'
            WHEN A.ACTION_CD = 'ERRDNA' THEN 'Either the claim or the denial code could not be found in DocDNA'
            WHEN A.ACTION_CD = 'ERREOB' THEN 'There is no denial code present on the EOB in DocDNA'
            WHEN A.ACTION_CD = 'EXPER' THEN 'Procedure Considered Experimental'
            WHEN A.ACTION_CD = 'FAULT' THEN 'Applied to No Fault Benefit'
            WHEN A.ACTION_CD = 'HCFA' THEN 'Claim Needs to be Submitted on CMS/HCFA 1500 Form'
            WHEN A.ACTION_CD = 'HMSI' THEN 'Additional Info Needed (crosswalk to INFO)'
            WHEN A.ACTION_CD = 'IACC' THEN 'Accident Report Required'
            WHEN A.ACTION_CD = 'IANPI' THEN 'Missing/Invalid Attending NPI Number'
            WHEN A.ACTION_CD = 'IBNPI' THEN 'Missing/Invalid Billing NPI Number'
            WHEN A.ACTION_CD = 'ICD10' THEN 'Invalid ICD-10 code.'
            WHEN A.ACTION_CD = 'ICLF' THEN 'Claim Form Requested'
            WHEN A.ACTION_CD = 'ICOB' THEN 'COB Information Needed'
            WHEN A.ACTION_CD = 'ICPD' THEN 'CPT4 Code Description Needed'
            WHEN A.ACTION_CD = 'ICPT' THEN 'CPT4 Code Required'
            WHEN A.ACTION_CD = 'IDAYS' THEN 'Missing/Invalid Days Supply'
            WHEN A.ACTION_CD = 'IDEA' THEN 'Missing/Invalid DEA Number'
            WHEN A.ACTION_CD = 'IDIG' THEN 'DX Code Required'
            WHEN A.ACTION_CD = 'IDOB' THEN 'Invalid/Missing DOB'
            WHEN A.ACTION_CD = 'IDOS' THEN 'Date(s) of Service Needed'
            WHEN A.ACTION_CD = 'IEOB' THEN 'Primary EOB Needed'
            WHEN A.ACTION_CD = 'ILAC' THEN 'Nature of Illness or Accident Required'
            WHEN A.ACTION_CD = 'ILLEG' THEN 'Illegible Code on EOB'
            WHEN A.ACTION_CD = 'IMCV' THEN 'Medicare EOB Required'
            WHEN A.ACTION_CD = 'INAME' THEN 'Invalid Character in Name or Illegible Name'
            WHEN A.ACTION_CD = 'INFO' THEN 'Additional Information Needed'
            WHEN A.ACTION_CD = 'INOPID' THEN 'Missing/Invalid Operating Physician ID '
            WHEN A.ACTION_CD = 'INPI' THEN 'Missing/Invalid NPI Number'
            WHEN A.ACTION_CD = 'INPRC' THEN 'Invalid Procedure Code'
            WHEN A.ACTION_CD = 'INVADD' THEN 'Invalid/Incomplete Address'
            WHEN A.ACTION_CD = 'INVAM' THEN 'Invalid Amount'
            WHEN A.ACTION_CD = 'INVBN' THEN 'Missing/Invalid BIN Number'
            WHEN A.ACTION_CD = 'INVBT' THEN 'Invalid Bill Type'
            WHEN A.ACTION_CD = 'INVCD' THEN 'Invalid Code'
            WHEN A.ACTION_CD = 'INVDOS' THEN 'Invalid Date of Service'
            WHEN A.ACTION_CD = 'INVDT' THEN 'Invalid Admit Date'
            WHEN A.ACTION_CD = 'INVDX' THEN 'DX Not Valid for Procedure'
            WHEN A.ACTION_CD = 'INVES' THEN 'Claim Being Investigated - Includes Possible Mismatch Patients'
            WHEN A.ACTION_CD = 'INVGR' THEN 'Invalid Group Number (use INVGRP)'
            WHEN A.ACTION_CD = 'INVGRP  ' THEN 'Invalid Group Number'
            WHEN A.ACTION_CD = 'INVID   ' THEN 'Invalid ID Number'
            WHEN A.ACTION_CD = 'INVMOD' THEN 'Missing/Invalid Modifier'
            WHEN A.ACTION_CD = 'INVNAM' THEN 'Invalid Name'
            WHEN A.ACTION_CD = 'INVND' THEN 'Non-Matched Service Provider Id'
            WHEN A.ACTION_CD = 'INVPCN' THEN 'Missing/Invalid PCN Number'
            WHEN A.ACTION_CD = 'INVPOS' THEN 'Incorrect Place of Service'
            WHEN A.ACTION_CD = 'IPOA' THEN 'Missing/Invalid POA Indicator'
            WHEN A.ACTION_CD = 'IPRO' THEN 'Provider Information Required'
            WHEN A.ACTION_CD = 'IREC    ' THEN 'Medical Records Needed'
            WHEN A.ACTION_CD = 'IREL' THEN 'Patient Relation to Insured Required'
            WHEN A.ACTION_CD = 'IRNPI' THEN 'Missing/Invalid Referring NPI'
            WHEN A.ACTION_CD = 'ITAX' THEN 'Invalid billing provider Tax ID'
            WHEN A.ACTION_CD = 'ITMZ' THEN 'Itemized Bill Required'
            WHEN A.ACTION_CD = 'KEYED' THEN 'Denial Code Keyed Incorrectly'
            WHEN A.ACTION_CD = 'LEGAL   ' THEN 'HMS Legal Working on Claims Population'
            WHEN A.ACTION_CD = 'LIABL' THEN 'Injury/Illness is Covered by the Liability Carrier'
            WHEN A.ACTION_CD = 'MCV' THEN 'Medicare Voucher Requested (Use IMCV)'
            WHEN A.ACTION_CD = 'MDCLM' THEN 'Carrier Acknowledges as a Medicaid Reclamation Claim'
            WHEN A.ACTION_CD = 'NCRTP' THEN 'No Code Carrier Refused to Process Claims'
            WHEN A.ACTION_CD = 'NEWMOM' THEN 'Newborn Claims Must be Submitted Under the Mothers Policy'
            WHEN A.ACTION_CD = 'NOAUTH' THEN 'No Authorization Obtained'
            WHEN A.ACTION_CD = 'NOCOD   ' THEN 'No Denial Code on Remittance'
            WHEN A.ACTION_CD = 'NOINT' THEN 'Interim Bills Not Processed, Submit for Entire Admission'
            WHEN A.ACTION_CD = 'NOTES' THEN 'Physicians Orders or Nursing Notes Needed'
            WHEN A.ACTION_CD = 'NOXWK' THEN 'No-Crosswalk'
            WHEN A.ACTION_CD = 'NSXWK' THEN 'Non Standard Crosswalk'
            WHEN A.ACTION_CD = 'NWAIT   ' THEN 'Waiting Period for Benefit Not Satisfied'
            WHEN A.ACTION_CD = 'OCCUR' THEN 'NUBC Occurrence Code(s)'
            WHEN A.ACTION_CD = 'OPN' THEN 'Open Claim Not Yet Adjudicated by Carrier'
            WHEN A.ACTION_CD = 'OUTNW' THEN 'Out of Network'
            WHEN A.ACTION_CD = 'PAPER   ' THEN 'Resubmit on Paper'
            WHEN A.ACTION_CD = 'PBM' THEN 'Send Claim to PBM'
            WHEN A.ACTION_CD = 'PCPRE' THEN 'Not Authorized by PCP (Use NOAUTH)'
            WHEN A.ACTION_CD = 'PDIA' THEN 'Invalid Principal/Admit Diagnosis'
            WHEN A.ACTION_CD = 'PEND    ' THEN 'Claim Pended'
            WHEN A.ACTION_CD = 'PREAUT' THEN 'Pre Auth Needed (use NOAUTH)'
            WHEN A.ACTION_CD = 'PTINFO' THEN 'Need Information From Patient'
            WHEN A.ACTION_CD = 'QUAL' THEN 'Qualifying Procedure Not Received'
            WHEN A.ACTION_CD = 'RENPRO' THEN 'Pending Information from Rendering Provider'
            WHEN A.ACTION_CD = 'RESENT  ' THEN 'Resent Claims'
            WHEN A.ACTION_CD = 'RESUB   ' THEN 'Submit Claim to Another Entity'
            WHEN A.ACTION_CD = 'RNID' THEN 'Missing/invalid rendering provider ID number'
            WHEN A.ACTION_CD = 'SPLIT' THEN 'Split into Multiple Claims at Carrier'
            WHEN A.ACTION_CD = 'SPPBM' THEN 'Must Fill through Specialty Pharmacy'
            WHEN A.ACTION_CD = 'STDNT' THEN 'Coverage for Student Terminated due to Reaching Maximum Age'
            WHEN A.ACTION_CD = 'SUBMH' THEN 'Submit to Mental Health Carrier'
            WHEN A.ACTION_CD = 'TIMEL' THEN 'Claim Past Timely Filing Limit'
            WHEN A.ACTION_CD = 'TIMELY' THEN 'Time Limit for filing has expired'
            WHEN A.ACTION_CD = 'TIMRX' THEN 'Prescription Too Old'
            WHEN A.ACTION_CD = 'TOOTH' THEN 'Missing/Invalid Tooth Number'
            WHEN A.ACTION_CD = 'TPA' THEN 'Send Claim to TPA'
            WHEN A.ACTION_CD = 'TPLNC' THEN 'Group does not Allow Third Party Claims'
            WHEN A.ACTION_CD = 'UB92' THEN 'Resubmit Claim on a UB Form'
            WHEN A.ACTION_CD = 'UMGRP' THEN 'Group Number does not Match Carriers System (use INVGRP)'
            WHEN A.ACTION_CD = 'UMID    ' THEN 'ID Number does not Match Carriers System (use INVID)'
            WHEN A.ACTION_CD = 'UNITS' THEN 'Units Field Invalid for Number of Days'
            WHEN A.ACTION_CD = 'UNPRO' THEN 'Unprocessed Claim'
            WHEN A.ACTION_CD = 'ADJUD' THEN 'Adjudicated per Plan Contract/Allowable'
            WHEN A.ACTION_CD = 'AMBORG' THEN 'Facility point of origin and destination - ambulance.'
            WHEN A.ACTION_CD = 'APDTF   ' THEN 'Timely Filing Appeal Denied'
            WHEN A.ACTION_CD = 'ARDUP   ' THEN 'Claim is a Duplicate of Previously Billed Claim'
            WHEN A.ACTION_CD = 'AUTHDN' THEN 'Pre-Authorization Denied Prior to Service'
            WHEN A.ACTION_CD = 'BANK' THEN 'Carrier Has Filed for Bankruptcy'
            WHEN A.ACTION_CD = 'BILLER  ' THEN 'Billing Error'
            WHEN A.ACTION_CD = 'CANCEL' THEN 'Policy Canceled'
            WHEN A.ACTION_CD = 'CAPIT' THEN 'Capitated Service'
            WHEN A.ACTION_CD = 'CLMAJ   ' THEN 'Claim Adjusted'
            WHEN A.ACTION_CD = 'CLMFRQ' THEN 'Frequency of service.'
            WHEN A.ACTION_CD = 'CMPMI' THEN 'Missing/Invalid Compound Code'
            WHEN A.ACTION_CD = 'CMPNC' THEN 'Compounds Not Covered'
            WHEN A.ACTION_CD = 'COINS' THEN 'MA Paid Less Than Co-Insurance'
            WHEN A.ACTION_CD = 'CONTR   ' THEN 'Contraception Not Covered'
            WHEN A.ACTION_CD = 'COPAY' THEN 'MA Paid Less Than Copay'
            WHEN A.ACTION_CD = 'DAYSUP  ' THEN 'Days Supply Exceeds Plan Limits'
            WHEN A.ACTION_CD = 'DEDUC' THEN 'Payment Applied to Patient Deductible'
            WHEN A.ACTION_CD = 'DEFER' THEN 'Portion of Payment Deferred'
            WHEN A.ACTION_CD = 'DEFUNCT' THEN 'Carrier No longer Exist'
            WHEN A.ACTION_CD = 'DEPNE   ' THEN 'Dependent Not Eligible'
            WHEN A.ACTION_CD = 'DEPPG' THEN 'Dependent Pregnancy Not Covered'
            WHEN A.ACTION_CD = 'DNYPD' THEN 'Uncollectable Claim that was Paid to Provider, Patient or State'
            WHEN A.ACTION_CD = 'DOSNE' THEN 'Coverage not in effect at time of service '
            WHEN A.ACTION_CD = 'DPWMI' THEN 'Date Prescription Written Missing/Invalid'
            WHEN A.ACTION_CD = 'DRGNC' THEN 'Drug Not Covered'
            WHEN A.ACTION_CD = 'DSPMI' THEN 'Missing/Invalid Dispense as Written Code'
            WHEN A.ACTION_CD = 'DUPPRO' THEN 'Duplicate to Provider Claim'
            WHEN A.ACTION_CD = 'DURCF' THEN 'Insert Fail DUR-Conflict'
            WHEN A.ACTION_CD = 'DURNC' THEN 'Durable Medical Equipment Not Covered'
            WHEN A.ACTION_CD = 'DXNC' THEN 'Diagnosis Not Covered'
            WHEN A.ACTION_CD = 'EMPNE' THEN 'Employee not Eligible'
            WHEN A.ACTION_CD = 'EXFREQ' THEN 'Service Exceeds Approved Frequency'
            WHEN A.ACTION_CD = 'EXRCR' THEN 'Exceeds Reasonable and Customary Rate'
            WHEN A.ACTION_CD = 'FLEX    ' THEN 'Flexible Spending Account Payments'
            WHEN A.ACTION_CD = 'FRMLRY' THEN 'Product Not On Formulary'
            WHEN A.ACTION_CD = 'FRONT   ' THEN 'Front End Edit Reject'
            WHEN A.ACTION_CD = 'GENSB' THEN 'Generic Substitution Required'
            WHEN A.ACTION_CD = 'GRPNE' THEN 'Group Not Eligible'
            WHEN A.ACTION_CD = 'HOSPIC' THEN 'Patient not covered for Hospice care'
            WHEN A.ACTION_CD = 'ICNPD   ' THEN 'DO NOT USE - Duplicate Claim Closed Out - ICN billed to and paid by different carrier (Auto Deny Job)'
            WHEN A.ACTION_CD = 'IDOL' THEN 'Mapaid > Billed'
            WHEN A.ACTION_CD = 'INCL' THEN 'Carrier will not Pay Separately for this Service'
            WHEN A.ACTION_CD = 'INDC' THEN 'Missing or Invalid NDC Number'
            WHEN A.ACTION_CD = 'INGMX' THEN 'Ingredient Cost Reduced to Maximum'
            WHEN A.ACTION_CD = 'INVGN' THEN 'Invalid Gender'
            WHEN A.ACTION_CD = 'INVPL' THEN 'Invalid or Incomplete Protocol Requirements'
            WHEN A.ACTION_CD = 'INVQT' THEN 'Invalid Quantity Entered for Medication Package '
            WHEN A.ACTION_CD = 'IRX     ' THEN 'Missing/Invalid RX Number'
            WHEN A.ACTION_CD = 'LMAX' THEN 'Lifetime Benefit Maximum Met'
            WHEN A.ACTION_CD = 'LOSC' THEN 'Invalid Location of Service'
            WHEN A.ACTION_CD = 'MAXBEN' THEN 'Benefit Maximum has been reached'
            WHEN A.ACTION_CD = 'MAXBN' THEN 'Maximum Benefits Reached (Use MAXBEN)'
            WHEN A.ACTION_CD = 'MEDAJ' THEN 'Medicaid Claim Adjudicated'
            WHEN A.ACTION_CD = 'MEDNCO' THEN 'Not a Medicare Covered Service'
            WHEN A.ACTION_CD = 'MEDNEC' THEN 'Claim Not Medically Necessary'
            WHEN A.ACTION_CD = 'MHNC    ' THEN 'Mental Health Not Covered'
            WHEN A.ACTION_CD = 'NABP    ' THEN 'Carrier Needs NABP Number'
            WHEN A.ACTION_CD = 'NBPT' THEN 'Non-Billable Provider Type'
            WHEN A.ACTION_CD = 'NDCNC' THEN 'NDC code Not Covered'
            WHEN A.ACTION_CD = 'NEICPD' THEN 'Payment Per Negotiated Rate'
            WHEN A.ACTION_CD = 'NENTPT' THEN 'Entity Not Found: Patient'
            WHEN A.ACTION_CD = 'NOBIL   ' THEN 'Claim Should Not Have Been Billed per Client - Do Not Work/Rebill Claims'
            WHEN A.ACTION_CD = 'NOCOV' THEN 'No Coverage'
            WHEN A.ACTION_CD = 'NODENT' THEN 'No Dental Coverage'
            WHEN A.ACTION_CD = 'NOMAT' THEN 'Maternity Charges Not Covered'
            WHEN A.ACTION_CD = 'NOMED   ' THEN 'Medical Services Not Covered'
            WHEN A.ACTION_CD = 'NONEW' THEN 'Insured has No Coverage for Newborns'
            WHEN A.ACTION_CD = 'NONURS  ' THEN 'Home Nursing Services Not Covered'
            WHEN A.ACTION_CD = 'NOOBES' THEN 'Obesity Services Not Covered'
            WHEN A.ACTION_CD = 'NOPAY' THEN 'No Payment to be Issued'
            WHEN A.ACTION_CD = 'NOVIS' THEN 'Vision Services Not Covered'
            WHEN A.ACTION_CD = 'NPHARM' THEN 'Non-Matched Pharmacy Number'
            WHEN A.ACTION_CD = 'OTCNC   ' THEN 'Over the Counter Drugs Not Covered'
            WHEN A.ACTION_CD = 'PARTB' THEN 'Claim not Processed - Medicare Part B policy'
            WHEN A.ACTION_CD = 'PARTD' THEN 'Claim not Processed - Medicare Part D policy'
            WHEN A.ACTION_CD = 'PDCR    ' THEN 'Paid at Customary and Reasonable Rate'
            WHEN A.ACTION_CD = 'PDPVPB' THEN 'Paid to Provider Prior to HMS Billing'
            WHEN A.ACTION_CD = 'PDTOPT  ' THEN 'Claim Paid to Patient'
            WHEN A.ACTION_CD = 'PDTOPV' THEN 'Claim Paid to Provider of Service'
            WHEN A.ACTION_CD = 'PDTOST  ' THEN 'Paid to State Medicaid Agency'
            WHEN A.ACTION_CD = 'PDTOWC' THEN 'Paid to Incorrect HMS Client'
            WHEN A.ACTION_CD = 'PHARNE  ' THEN 'Pharmacy Not Eligible'
            WHEN A.ACTION_CD = 'PHYNC' THEN 'Physician Not Covered '
            WHEN A.ACTION_CD = 'PRCNC' THEN 'Procedure Code Not Covered'
            WHEN A.ACTION_CD = 'PREEX' THEN 'Pre-Existing Condition'
            WHEN A.ACTION_CD = 'PROBL' THEN 'Servicing Provider Must Bill Carrier Directly'
            WHEN A.ACTION_CD = 'PROVN' THEN 'Provider Not Covered (use PRVNC)'
            WHEN A.ACTION_CD = 'PRSC' THEN 'Expected a Prescriber Field'
            WHEN A.ACTION_CD = 'PRSNC' THEN 'Drug Not Covered for Prescriber'
            WHEN A.ACTION_CD = 'PRVNC' THEN 'Provider Not Covered'
            WHEN A.ACTION_CD = 'PTAGE' THEN 'Patient Outside Age Limit for this Type of Benefit'
            WHEN A.ACTION_CD = 'PTRES' THEN 'Patient Responsibility'
            WHEN A.ACTION_CD = 'QUANT   ' THEN 'Quantity Not Covered'
            WHEN A.ACTION_CD = 'REBIL   ' THEN 'Claim Closed to Pass through New Cycle'
            WHEN A.ACTION_CD = 'REFIL' THEN 'Refill Too Soon'
            WHEN A.ACTION_CD = 'REFMI' THEN 'Refill Number Missing/Invalid'
            WHEN A.ACTION_CD = 'RIDER' THEN 'Considered Under Rider Coverage'
            WHEN A.ACTION_CD = 'ROUTN' THEN 'Routine Services Not Covered'
            WHEN A.ACTION_CD = 'RXAGE' THEN 'Participants Age Restricts Medication Coverage'
            WHEN A.ACTION_CD = 'RXDSG' THEN 'Fail Rx Dosage Rule Table'
            WHEN A.ACTION_CD = 'RXGEN' THEN 'Participants Gender Restricts Medication Coverage'
            WHEN A.ACTION_CD = 'RXLMT   ' THEN 'Medication Exceeds Plan Limits'
            WHEN A.ACTION_CD = 'RXNC' THEN 'Prescription Drugs Not Covered'
            WHEN A.ACTION_CD = 'TERMGRP' THEN 'The group termed with this carrier and the run out period has ended'
            WHEN A.ACTION_CD = 'TOSNE' THEN 'Type of service not covered'
            WHEN A.ACTION_CD = 'TRVLNC' THEN 'Travel/Transportation Not Covered'
            WHEN A.ACTION_CD = 'UNID' THEN 'Unable to Identify Member'
            WHEN A.ACTION_CD = 'WASTE' THEN 'Procedure code that are not covered by the carrier and should never go back to the same carrier'
            WHEN A.ACTION_CD = 'WCOMP' THEN 'Claim should be Processed by Workers Comp Carrier'
            WHEN A.ACTION_CD = 'WWEST'THEN 'Outsourced to Washington & West (Massachusetts Medicaid)'
            WHEN A.ACTION_CD = 'YMAXBN'THEN 'Annual Maximum Benefits Reached (use MAXBEN)'
            WHEN A.ACTION_CD = 'ZPDPR'THEN 'Zero Pay Carrier - Billed from a PAR/Imputed Carrier Feed'
            WHEN A.ACTION_CD = 'ZPDRS'THEN 'Zero Pay Carrier - Billed from the State Resource Feed'
            WHEN A.ACTION_CD = 'PO1'THEN 'Full payment received'
            WHEN A.ACTION_CD = 'PO2'THEN 'Partial Payment received'
            WHEN A.ACTION_CD = 'P03'THEN 'Voluntary/Excess payment received'
            WHEN A.ACTION_CD = 'PEO1'THEN 'Provider grant extension.'
            WHEN A.ACTION_CD = 'PEO2'THEN 'Credit for Claim due to HMS (do not recoup)'
            WHEN A.ACTION_CD = 'RO1'THEN 'Provider agrees with recoupment'
            WHEN A.ACTION_CD = 'RO2'THEN 'Recoupment scheduled for future dates'
            WHEN A.ACTION_CD = 'RO3'THEN 'Recoupment for future date (final balance)'
            WHEN A.ACTION_CD = 'RECUP10'THEN 'Recovery has been submitted for invoicing'
            WHEN A.ACTION_CD = 'V01'THEN 'Not Eligible for Medicare on DOS'
            WHEN A.ACTION_CD = 'V02'THEN 'Medicare Benefits Exhausted'
            WHEN A.ACTION_CD = 'V03'THEN 'Invalid MC HIC # '
            WHEN A.ACTION_CD = 'V05'THEN 'Non-covered Medicare Service(Provider Relations denial)'
            WHEN A.ACTION_CD = 'V06'THEN 'Prior MA Recoupment'
            WHEN A.ACTION_CD = 'V10'THEN 'Untimely'
            WHEN A.ACTION_CD = 'V11'THEN 'Administrative days or PRO denial (bed hold days)'
            WHEN A.ACTION_CD = 'V12'THEN 'Correctly Billed by Medicare and Medicaid'
            WHEN A.ACTION_CD = 'V13'THEN 'Not a Medicare Assigned Provider'
            WHEN A.ACTION_CD = 'V14'THEN 'Mail returned'
            WHEN A.ACTION_CD = 'V15'THEN 'HH:Patient not home bound'
            WHEN A.ACTION_CD = 'V16'THEN 'HH:PT/OT/ST not under treatment plan'
            WHEN A.ACTION_CD = 'V17'THEN 'HH:Nursing care not under treatment plan'
            WHEN A.ACTION_CD = 'V18'THEN 'HH: Nursing care not skilled.'
            WHEN A.ACTION_CD = 'V19'THEN 'HH: Nursing no intermittent'
            WHEN A.ACTION_CD = 'V20'THEN 'HH: Patient condition chronic'
            WHEN A.ACTION_CD = 'V21 'THEN 'Disallow amount is for DED/Coins'
            WHEN A.ACTION_CD = 'V29'THEN 'Patient not affiliated with provider. (Patient not found by provider.)'
            WHEN A.ACTION_CD = 'VO4'THEN 'Bankrupt provider'
            WHEN A.ACTION_CD = 'VO7'THEN 'Prior providers refund'
            WHEN A.ACTION_CD = 'VO8'THEN 'SNF/Non-skilled level of care'
            WHEN A.ACTION_CD = 'VO9'THEN 'Legal providers (MAT)'
            WHEN A.ACTION_CD = 'V30'THEN 'Case Disallowed '
            WHEN A.ACTION_CD = 'V31'THEN 'Direct Payment'
            WHEN A.ACTION_CD = 'V32'THEN 'Duplicate - MRM'
            WHEN A.ACTION_CD = 'V33'THEN 'Duplicate - vendor'
            WHEN A.ACTION_CD = 'V34'THEN 'Failure to pursue'
            WHEN A.ACTION_CD = 'V35'THEN 'Member Paid'
            WHEN A.ACTION_CD = 'V36'THEN 'Negotiated Rate'
            WHEN A.ACTION_CD = 'V37'THEN 'No Auth for Service'
            WHEN A.ACTION_CD = 'V38'THEN 'No Causal Relationship'
            WHEN A.ACTION_CD = 'V39'THEN 'Provider Takeback'
            WHEN A.ACTION_CD = 'V40'THEN 'Refund MRM Fee'
            WHEN A.ACTION_CD = 'V41'THEN 'Settlement Approved'
            WHEN A.ACTION_CD = 'V42'THEN 'Aetna Unrecoverable Amount'
            WHEN A.ACTION_CD = 'V43'THEN 'Provider Denied'
            WHEN A.ACTION_CD = 'OPN01 'THEN 'Open Claim This Claim Will Be Recouped'
            WHEN A.ACTION_CD = 'OPN02 'THEN 'No Response Extension'
            WHEN A.ACTION_CD = 'OPN03 'THEN 'Claim Recovered/Missing Documentation'
            WHEN A.ACTION_CD = 'OPN04 'THEN 'No Medical Records Received'
            WHEN A.ACTION_CD = 'OPN05 'THEN 'No Response During 2nd Extension'
            WHEN A.ACTION_CD = 'OPN06 'THEN 'Final Notice Of Recovery Sent'
            WHEN A.ACTION_CD = 'OPN07 'THEN 'No Response During 3rd Extension'
            WHEN A.ACTION_CD = 'P01 'THEN 'Full Payment Received'
            WHEN A.ACTION_CD = 'P02 'THEN 'Partial Payment Received'
            WHEN A.ACTION_CD = 'P03 'THEN 'Partial Payment P03'
            WHEN A.ACTION_CD = 'P04 'THEN 'MD Checks Received Via Provider'
            WHEN A.ACTION_CD = 'P05 'THEN 'Extension Payment Received'
            WHEN A.ACTION_CD = 'P06 'THEN 'Payment From EMOMED Forms'
            WHEN A.ACTION_CD = 'P50 'THEN 'Paid By Check - Private Health Insurance Paid'
            WHEN A.ACTION_CD = 'P51 'THEN 'Paid By Check - Medicaid Paid Twice (Duplicate)'
            WHEN A.ACTION_CD = 'P52 'THEN 'Paid By Check - Medicare Paid'
            WHEN A.ACTION_CD = 'P53 'THEN 'Paid By Check - Patient Paid.'
            WHEN A.ACTION_CD = 'P54 'THEN 'Paid By Check - Medicare/Medicaid Cross-Over Pymt'
            WHEN A.ACTION_CD = 'P55 'THEN 'Paid By Check - Other Noted By Provider'
            WHEN A.ACTION_CD = 'P56 'THEN 'Paid By Check - Casualty/Estate Recovery'
            WHEN A.ACTION_CD = 'P57 'THEN 'Paid By Check - Billing Error'
            WHEN A.ACTION_CD = 'P58 'THEN 'Paid By Check - Medicare SNF Stay'
            WHEN A.ACTION_CD = 'PAMDP 'THEN 'Pending Additional Medical Records To Permedion'
            WHEN A.ACTION_CD = 'PE01 'THEN 'Provider Granted Extension'
            WHEN A.ACTION_CD = 'PE02 'THEN 'Provider Granted 2nd 60day Extension (CATPL Only)'
            WHEN A.ACTION_CD = 'PE03 'THEN 'Provider Granted 3rd 60day Extension (CATPL Only)'
            WHEN A.ACTION_CD = 'PE04 'THEN 'Provider Is Sending Check Into Lockbox'
            WHEN A.ACTION_CD = 'PE05 'THEN 'Claims Are Pending/ In Process Of Being Worked'
            WHEN A.ACTION_CD = 'PE06 'THEN 'Medical Records Rec`D In Process Of Being Worked'
            WHEN A.ACTION_CD = 'PE10 'THEN 'CATPL ONLY Extension Per CA PD 120 Days'
            WHEN A.ACTION_CD = 'PE100 'THEN 'Provider Requesting Extension'
            WHEN A.ACTION_CD = 'PE30 'THEN '1st 30 Day Extension Request'
            WHEN A.ACTION_CD = 'PE500 'THEN '(PIRA)Providers Wants To Appeal, Pending Records'
            WHEN A.ACTION_CD = 'PE505 'THEN 'Rec`D Appeal Docs In Process Of Working Docs'
            WHEN A.ACTION_CD = 'PE510 'THEN '(PIRA) Pending Appeal To State'
            WHEN A.ACTION_CD = 'PE60 'THEN '2nd 30 Day Extension Request'
            WHEN A.ACTION_CD = 'PE600 'THEN 'CA-Provider Has Billed Or Attempted To Bill'
            WHEN A.ACTION_CD = 'PE90 'THEN '3rd 30 Day Extension Request'
            WHEN A.ACTION_CD = 'R01 'THEN 'Provider Agrees With Recoupment'
            WHEN A.ACTION_CD = 'R02 'THEN 'Recoupment Scheduled For Future Date'
            WHEN A.ACTION_CD = 'R03 'THEN 'Result Of Appeal Process Thru Permedion'
            WHEN A.ACTION_CD = 'R04 'THEN 'Provider Verbally Agreed To Recoupment'
            WHEN A.ACTION_CD = 'R05 'THEN 'Patient Liability Not Applicable Partial Recoup'
            WHEN A.ACTION_CD = 'R06 'THEN 'Recoup During Extension'
            WHEN A.ACTION_CD = 'R07 'THEN 'Recoupment Confirmed'
            WHEN A.ACTION_CD = 'R08 'THEN 'Adjustment Form Received-Not Processed By Client'
            WHEN A.ACTION_CD = 'R09 'THEN 'Provider Verbally Agree To Recoup During Extension'
            WHEN A.ACTION_CD = 'R10 'THEN 'Recoup/Untimely Denied/Appeal'
            WHEN A.ACTION_CD = 'R11 'THEN 'Provider Agrees To Recoupment - NC Recoup File'
            WHEN A.ACTION_CD = 'R12 'THEN 'Recoup During 2nd - 30 Day Extension'
            WHEN A.ACTION_CD = 'R13 'THEN 'Recoup During 3rd- 30 Day Extension'
            WHEN A.ACTION_CD = 'R14 'THEN 'Provider Verbally Agrees To Recoup - NC File'
            WHEN A.ACTION_CD = 'R15 'THEN 'Partial Recoupment'
            WHEN A.ACTION_CD = 'R16 'THEN 'Recoupment File Processed'
            WHEN A.ACTION_CD = 'R17 'THEN 'Follow-Up Completed.'
            WHEN A.ACTION_CD = 'R18 'THEN 'Possible Double Recoupment'
            WHEN A.ACTION_CD = 'R19 'THEN 'Unacceptable Documentation Received'
            WHEN A.ACTION_CD = 'R20 'THEN 'Claim Recouped Mid Cycle'
            WHEN A.ACTION_CD = 'R21 'THEN 'Incomp Documentation Rec`D 1st Ext (Days 90-120)'
            WHEN A.ACTION_CD = 'R22 'THEN 'Incomp Documentation Rec`D 2nd Ext (Days 120-150)'
            WHEN A.ACTION_CD = 'R23 'THEN 'CX Provider Agrees To Refund'
            WHEN A.ACTION_CD = 'R24 'THEN 'Responded But No Documentation Provided'
            WHEN A.ACTION_CD = 'R25 'THEN 'VA-Provider Initated Adj/Refund Due To HMS Audit'
            WHEN A.ACTION_CD = 'R26 'THEN 'Extension Request Denied By Account Team'
            WHEN A.ACTION_CD = 'R27 'THEN 'CX Cleveland Clinic No Response From Provider'
            WHEN A.ACTION_CD = 'R30 'THEN 'Recoup At 30 Days'
            WHEN A.ACTION_CD = 'R50 'THEN 'Recoupment - Private Health Insurance Paid'
            WHEN A.ACTION_CD = 'R51 'THEN 'Recoupment - Medicaid Paid Twice. (Duplicate)'
            WHEN A.ACTION_CD = 'R52 'THEN 'Recoupment - Medicare Paid'
            WHEN A.ACTION_CD = 'R53 'THEN 'Recoupment - Patient Paid'
            WHEN A.ACTION_CD = 'R54 'THEN 'Recoupment - Medicare/Medcaid Crossover Payment'
            WHEN A.ACTION_CD = 'R55 'THEN 'Recoupment - Other Noted By Provider'
            WHEN A.ACTION_CD = 'R56 'THEN 'Recoupment - Casualty/Estate Recovery'
            WHEN A.ACTION_CD = 'R57 'THEN 'Recoupment - Billing Error'
            WHEN A.ACTION_CD = 'R58 'THEN 'Recoupment - Medicare SNF Stay'
            WHEN A.ACTION_CD = 'R59 'THEN 'Claim Voided Via Mmis By Provider'
            WHEN A.ACTION_CD = 'R60 'THEN 'Recoup At 60 Days'
            WHEN A.ACTION_CD = 'R70 'THEN 'Reconsideration- Upheld'
            WHEN A.ACTION_CD = 'R72 'THEN 'Docs Does Not Support Scenario'
            WHEN A.ACTION_CD = 'R80 'THEN 'Unacceptable Docs Received During Extension'
            WHEN A.ACTION_CD = 'R81 'THEN 'Claim Denied For Untimely Filing During Extension'
            WHEN A.ACTION_CD = 'R82 'THEN 'No Documentation Received'
            WHEN A.ACTION_CD = 'R90 'THEN 'Recoup At 90 Days'
            WHEN A.ACTION_CD = 'R91 'THEN 'Preliminary Findings'
            WHEN A.ACTION_CD = 'R92 'THEN 'Upheld/ Result Of Appeal'
            WHEN A.ACTION_CD = 'R93 'THEN 'Upheld/ Appeal To State'
            WHEN A.ACTION_CD = 'R94 'THEN 'Technical Denial'
            WHEN A.ACTION_CD = 'R95 'THEN 'Final Recovery Letter Sent'
            WHEN A.ACTION_CD = 'RECON 'THEN 'Reconsideration Request Received'
            WHEN A.ACTION_CD = 'RITA 'THEN 'Untimely Rebuttal Received - Uphold'
            WHEN A.ACTION_CD = 'RLTNO 'THEN 'Tentative Notice Of Overpayment Sent'
            WHEN A.ACTION_CD = 'RPCOM 'THEN 'Complex Review Payment Received'
            WHEN A.ACTION_CD = 'RUP 'THEN 'Upheld Letter Generated'
            WHEN A.ACTION_CD = 'SDR 'THEN 'Supporting Documentation Rec`Vd-In Review Status'
            WHEN A.ACTION_CD = 'TNOOS 'THEN 'Tentative Notic Of Overpayment Sent'
            WHEN A.ACTION_CD = 'U01 'THEN 'Rac Identified Underpayments'
            WHEN A.ACTION_CD = 'U02 'THEN 'RAC Identified Dollar Value Of Underpayments'
            WHEN A.ACTION_CD = 'V01 'THEN 'No Coverage For DOS (Dates Of Service)'
            WHEN A.ACTION_CD = 'V02 'THEN 'Medicare Benefit Days Exhausted'
            WHEN A.ACTION_CD = 'V03 'THEN 'Medicaid Recipient Mismatched'
            WHEN A.ACTION_CD = 'V04 'THEN 'Bankrupt Provider'
            WHEN A.ACTION_CD = 'V05 'THEN 'Non-Covered Service'
            WHEN A.ACTION_CD = 'V06 'THEN 'Prior Medicaid Recoupment'
            WHEN A.ACTION_CD = 'V07 'THEN 'Prior Provider Refund'
            WHEN A.ACTION_CD = 'V08 'THEN 'Medicaid Recouped Before Cycle'
            WHEN A.ACTION_CD = 'V09 'THEN 'MATPL ONLY Legal Providers'
            WHEN A.ACTION_CD = 'V10 'THEN 'Time Limit For Filing Has Expired'
            WHEN A.ACTION_CD = 'V100 'THEN 'Provider Does Not Agree With Recoupment'
            WHEN A.ACTION_CD = 'V11 'THEN 'Death Certificate Rec'
            WHEN A.ACTION_CD = 'V12 'THEN 'Correctly Billed By Blue Cross, Medicare, Medicaid'
            WHEN A.ACTION_CD = 'V13 'THEN 'Non-Participating Provider'
            WHEN A.ACTION_CD = 'V14 'THEN 'Provider Electronic Adjustment'
            WHEN A.ACTION_CD = 'V15 'THEN 'Missing Modifier-No Impact On Reimbursement'
            WHEN A.ACTION_CD = 'V16 'THEN 'HH: PT/OT/ST Not Under Treatment Plan'
            WHEN A.ACTION_CD = 'V17 'THEN 'HH:Nursing Care Not Under Treatment Plan'
            WHEN A.ACTION_CD = 'V18 'THEN 'HH: Nursing Care Not Skilled'
            WHEN A.ACTION_CD = 'V19 'THEN 'HH: Nursing No Intermittent'
            WHEN A.ACTION_CD = 'V20 'THEN 'HH: Patient Condition Chronic'
            WHEN A.ACTION_CD = 'V21 'THEN 'Disallow Amount Is For Ded/Co (MA Liability)'
            WHEN A.ACTION_CD = 'V22 'THEN 'Dupe Surgeons (Billed Correctly)'
            WHEN A.ACTION_CD = 'V23 'THEN 'Non-Dupes- Twin Birth'
            WHEN A.ACTION_CD = 'V24 'THEN 'Billed Correctly(Same Proc Left/Right Forced Clm)'
            WHEN A.ACTION_CD = 'V25 'THEN 'Wrong Claim Selected/Supporting Claim'
            WHEN A.ACTION_CD = 'V26 'THEN 'Bill Correctly(Diffrent Level Of Care)'
            WHEN A.ACTION_CD = 'V27 'THEN 'Paid Correctly - Crossover Claim'
            WHEN A.ACTION_CD = 'V28 'THEN 'Provider Refuses To Bill MC (MATPL)'
            WHEN A.ACTION_CD = 'V29 'THEN 'Patient Not Affiliated With Provider'
            WHEN A.ACTION_CD = 'V30 'THEN 'Void Per Client/PD'
            WHEN A.ACTION_CD = 'V31 'THEN 'SNF - Nursing Care Not Skilled'
            WHEN A.ACTION_CD = 'V32 'THEN 'SNF Beneficiary Not In A Medicare-Certified Bed'
            WHEN A.ACTION_CD = 'V33 'THEN 'SNF Benefit Days Exhausted (Please Verify W/CWF)'
            WHEN A.ACTION_CD = 'V34 'THEN 'SNF Hosp Stay Unrelated To Subseq SNF Skilled Care'
            WHEN A.ACTION_CD = 'V35 'THEN 'Nv - Soi, Noridian Admin. Svc.'
            WHEN A.ACTION_CD = 'V36 'THEN 'NV - SOI, Mutual Of Omaha'
            WHEN A.ACTION_CD = 'V37 'THEN 'MO - Prev. Adj, Credit Or Void'
            WHEN A.ACTION_CD = 'V38 'THEN 'Pre-Cert/Pre-Auth Denied'
            WHEN A.ACTION_CD = 'V39 'THEN 'Void Per Client/PD MA Paid Amount To Be Recouped'
            WHEN A.ACTION_CD = 'V40 'THEN 'Entered As Self Audit When Actually MCA'
            WHEN A.ACTION_CD = 'V41 'THEN 'Change Of Ownership- Provider Not Held Responsible'
            WHEN A.ACTION_CD = 'V42 'THEN 'BC Benefit Days Exhausted'
            WHEN A.ACTION_CD = 'V43 'THEN 'Date Of Death Incorrect'
            WHEN A.ACTION_CD = 'V44 'THEN 'Provider Closed'
            WHEN A.ACTION_CD = 'V45 'THEN 'Under Carrier Review'
            WHEN A.ACTION_CD = 'V46 'THEN 'Recoupment Confirmed (Non Credit)'
            WHEN A.ACTION_CD = 'V47 'THEN 'Verbally Stated Time Limit For Filing Has Expired'
            WHEN A.ACTION_CD = 'V48 'THEN 'Return Mail'
            WHEN A.ACTION_CD = 'V49 'THEN 'Other Insurance Is Primary'
            WHEN A.ACTION_CD = 'V50 'THEN 'Provider Refunded Another Entity'
            WHEN A.ACTION_CD = 'V51 'THEN 'Claims For Timely Filing Appealed Correctly'
            WHEN A.ACTION_CD = 'V52 'THEN 'Provider Initiated Online Adj Due To HMS Audit'
            WHEN A.ACTION_CD = 'V53 'THEN 'No Findings'
            WHEN A.ACTION_CD = 'V54 'THEN 'Reconsideration-Overturned'
            WHEN A.ACTION_CD = 'V55 'THEN 'Overturn/Result Of Appeal'
            WHEN A.ACTION_CD = 'V56 'THEN 'Overturn/Appeal To State'
            WHEN A.ACTION_CD = 'V57 'THEN 'Suspected Fraud'
            WHEN A.ACTION_CD = 'V58 'THEN 'Incorrect Policy Number'
            WHEN A.ACTION_CD = 'V59 'THEN 'Policy Not In The States MMIS'
            WHEN A.ACTION_CD = 'V60 'THEN 'Hardship/Disaster Relief'
            WHEN A.ACTION_CD = 'V61 'THEN 'Appeal Overturned'
            WHEN A.ACTION_CD = 'V62 'THEN 'No Records Received, Closing Per Client'
            WHEN A.ACTION_CD = 'V63 'THEN 'Nj Mmis Void'
            WHEN A.ACTION_CD = 'V64 'THEN 'Void Claim Prior To Mailing'
            WHEN A.ACTION_CD = 'V65 'THEN 'Void Audit Overlap Claim'
            WHEN A.ACTION_CD = 'V66 'THEN 'Claim Closed, Recovery Initiated'
            WHEN A.ACTION_CD = 'V67 'THEN 'Corrected Claim Submitted By Provider'
            WHEN A.ACTION_CD = 'V68 'THEN 'Benefits Exhausted -Ci Cycle La'
            WHEN A.ACTION_CD = 'V72'THEN 'Re-Disallow Claims'
            WHEN A.ACTION_CD = 'V99 'THEN 'Void Duplicate Claim'
            WHEN A.ACTION_CD = 'v99 'THEN 'Duplicate Claims Disallowed'
            WHEN A.ACTION_CD = 'BDRFPV'THEN 'Bad referring provider info'
            WHEN A.ACTION_CD = 'BLANKPIN'THEN 'the rendering provider number is missing'
            WHEN A.ACTION_CD = 'DODB4C'THEN 'Date of death prior to DOS'
            WHEN A.ACTION_CD = 'DOSNE'THEN 'Coverage not in effect at time of service '
            WHEN A.ACTION_CD = 'DUPE'THEN 'Duplicate Claim (use DUP)'
            WHEN A.ACTION_CD = 'DXINC'THEN 'DX code incomplete'
            WHEN A.ACTION_CD = 'DXNOTCOV'THEN 'DX is not covered'
            WHEN A.ACTION_CD = 'FSSINFO'THEN 'missing facility name'
            WHEN A.ACTION_CD = 'HOSPIC'THEN 'Patient not covered for Hospice care'
            WHEN A.ACTION_CD = 'INAPPSER'THEN 'Treatment rendered inappropriate'
            WHEN A.ACTION_CD = 'INVLDDX/INVLDX'THEN 'Claim denied based on diagnosis'
            WHEN A.ACTION_CD = 'INVLDPIN'THEN 'invalid rendering provider number'
            WHEN A.ACTION_CD = 'INVRNFN'THEN 'Invalid rendering provider first name'
            WHEN A.ACTION_CD = 'INVRNLN'THEN 'Invalid rendering provider last name'
            WHEN A.ACTION_CD = 'MAXBEN'THEN 'Benefit Maximum has been reached'
            WHEN A.ACTION_CD = 'MAXCHG'THEN 'Maximum charges exceeded'
            WHEN A.ACTION_CD = 'MISSINFO/MSINFO'THEN 'Claim lacks info needed for adjudication'
            WHEN A.ACTION_CD = 'NOHICNUM'THEN 'Missing HIC number'
            WHEN A.ACTION_CD = 'NOTPDSEP'THEN 'Procedure is not paid separately'
            WHEN A.ACTION_CD = 'NOUPIN'THEN 'the UPIN is missing on the claim'
            WHEN A.ACTION_CD = 'PRNOCERT/PRNOCE'THEN 'Provider was not certified for this procedure on this date of service'
            WHEN A.ACTION_CD = 'PRVNOB'THEN 'Provider not allowed to bill for service'
            WHEN A.ACTION_CD = 'Routine'THEN 'Non covered charges, routine exam'
            WHEN A.ACTION_CD = 'SEPINP'THEN 'Not paid separately when patient in Hosp.'
            WHEN A.ACTION_CD = 'Timely'THEN 'Time Limit for filing has expired (USE Timel)'
            WHEN A.ACTION_CD = 'TOSNE'THEN 'Type of service not covered'
            WHEN A.ACTION_CD = 'V01'THEN 'Coverage termination prior to DOS'
            WHEN A.ACTION_CD = 'V02'THEN 'DOS prior to coverage effective date'
            WHEN A.ACTION_CD = 'V03'THEN 'Dependent not eligible for coverage'
            WHEN A.ACTION_CD = 'V04'THEN 'Patient not covered under this policy'
            WHEN A.ACTION_CD = 'V05'THEN 'Carrier issued duplicate payment'
            WHEN A.ACTION_CD = 'V06'THEN 'Type of service not eligible for coverage'
            WHEN A.ACTION_CD = 'V07'THEN 'Carrier benefits incorrectly calculated'
            WHEN A.ACTION_CD = 'V08'THEN 'Patients maximum has been exceeded'
            WHEN A.ACTION_CD = 'V09'THEN 'Carrier filing limit exceeded'
            WHEN A.ACTION_CD = 'V10'THEN 'Pre-existing conditions not covered under this policy'
            WHEN A.ACTION_CD = 'V11'THEN 'Services not covered without prior authorization'
            WHEN A.ACTION_CD = 'V12'THEN 'Coverage was not coordinated correctly'
            WHEN A.ACTION_CD = 'V13'THEN 'Services rendered by this provider not covered'
            WHEN A.ACTION_CD = 'V14'THEN 'Services not rendered by provider'
            WHEN A.ACTION_CD = 'V15'THEN 'Carrier paid in excess of Medicaid paid amount'
            WHEN A.ACTION_CD = 'V16'THEN 'Member was retro-terminated'
    ELSE NULL
    END AS "HMS DESCRIPTION",

    C.CARRIER_ACTION_CD1 AS "CARRIER ACTION CD1",
    CASE 
    WHEN A.CARRIER_ACTION_CD = '277ACC' THEN '277 ACCEPTANCE/ACKNOWLEDGMENT'
    WHEN A.CARRIER_ACTION_CD = '-999' THEN 'Reason Code Unavailable'
    WHEN A.CARRIER_ACTION_CD = 'ACCDT' THEN 'Reason Code Unavailable'
    WHEN A.CARRIER_ACTION_CD = 'ADJUD' THEN 'Adjudicated per Plan Contract/Allowable'
    WHEN A.CARRIER_ACTION_CD = 'APID' THEN 'Reason Code Unavailable'
    WHEN A.CARRIER_ACTION_CD = 'ARDUP' THEN 'Claim is a Duplicate of Previously Billed Claim'
    WHEN A.CARRIER_ACTION_CD = 'B2BREJ' THEN 'Reason Code Unavailable'
    WHEN A.CARRIER_ACTION_CD = 'BADCLMS' THEN 'Reason Code Unavailable'
    WHEN A.CARRIER_ACTION_CD = 'BADEL' THEN 'Reason Code Unavailable'
    WHEN A.CARRIER_ACTION_CD = 'BILLER' THEN 'Billing Error'
    WHEN A.CARRIER_ACTION_CD = 'BPNAP' THEN 'Reason Code Unavailable'
    WHEN A.CARRIER_ACTION_CD = 'BPNPI' THEN 'Reason Code Unavailable'
    WHEN A.CARRIER_ACTION_CD = 'BPTAX' THEN 'Reason Code Unavailable'
    WHEN A.CARRIER_ACTION_CD = 'BPTXN' THEN 'Reason Code Unavailable'
    WHEN A.CARRIER_ACTION_CD = 'CAID' THEN 'Reason Code Unavailable'
    WHEN A.CARRIER_ACTION_CD = 'CANCEL' THEN 'Reason Code Unavailable'
    WHEN A.CARRIER_ACTION_CD = 'CAPIT' THEN 'Capitated Service'
    WHEN A.CARRIER_ACTION_CD = 'CCLM' THEN 'Corrected Claim Needed'
    WHEN A.CARRIER_ACTION_CD = 'CLAMAJ' THEN 'Care may not be Covered by Another Payer per Coordination of Benefits'
    WHEN A.CARRIER_ACTION_CD = 'CLMAJ' THEN 'Claim Adjusted'
    WHEN A.CARRIER_ACTION_CD = 'CLMFRQ' THEN 'Reason Code Unavailable'
    WHEN A.CARRIER_ACTION_CD = 'CLMFWD' THEN 'Claim Forwarded to Payer by NEIC'
    WHEN A.CARRIER_ACTION_CD = 'CMPMI' THEN 'Missing/Invalid Compound Code'
    WHEN A.CARRIER_ACTION_CD = 'CNBC' THEN 'Reason Code Unavailable'
    WHEN A.CARRIER_ACTION_CD = 'CO45' THEN 'Charge exceeds fee schedule/maximum allowable or contracted/legislated fee arrangement. Usage: This adjustment amount cannot equal the total service or claim charge amount; and must not duplicate provider adjustment amounts (payments and contractual reductions) that have resulted from prior payer(s) adjudication. (Use only with Group Codes PR or CO depending upon liability)'
    WHEN A.CARRIER_ACTION_CD = 'COBPD' THEN 'Paid in Accordance with COB'
    WHEN A.CARRIER_ACTION_CD = 'COINS' THEN 'MA Paid Less Than Co-Insurance'
    WHEN A.CARRIER_ACTION_CD = 'COPAY' THEN 'MA Paid Less Than Copay'
    WHEN A.CARRIER_ACTION_CD = 'DAVITA' THEN 'Reason Code Unavailable'
    WHEN A.CARRIER_ACTION_CD = 'DAYSUP' THEN 'Days Supply Exceeds Plan Limits'
    WHEN A.CARRIER_ACTION_CD = 'DEDUC' THEN 'Payment Applied to Patient Deductible'
    WHEN A.CARRIER_ACTION_CD = 'DEFER' THEN 'Portion of Payment Deferred'
    WHEN A.CARRIER_ACTION_CD = 'DENTL' THEN 'Group Has Dental Coverage Only'
    WHEN A.CARRIER_ACTION_CD = 'DEPNE' THEN 'Dependent Not Eligible'
    WHEN A.CARRIER_ACTION_CD = 'DEPPG' THEN 'Dependent Pregnancy Not Covered'
    WHEN A.CARRIER_ACTION_CD = 'DISDT' THEN 'Reason Code Unavailable'
    WHEN A.CARRIER_ACTION_CD = 'DISSTT' THEN 'Reason Code Unavailable'
    WHEN A.CARRIER_ACTION_CD = 'DOSNE' THEN 'Patient Not Eligible on Date of Service'
    WHEN A.CARRIER_ACTION_CD = 'DPWMI' THEN 'Date Prescription Written Missing/Invalid'
    WHEN A.CARRIER_ACTION_CD = 'DRGNC' THEN 'Drug Not Covered'
    WHEN A.CARRIER_ACTION_CD = 'DSPMI' THEN 'Missing/Invalid Dispense as Written Code'
    WHEN A.CARRIER_ACTION_CD = 'DUP' THEN 'Duplicate Claim Submission'
    WHEN A.CARRIER_ACTION_CD = 'DURCF' THEN 'Insert Fail DUR-Conflict'
    WHEN A.CARRIER_ACTION_CD = 'DXNC' THEN 'Diagnosis Not Covered'
    WHEN A.CARRIER_ACTION_CD = 'EDITGOV' THEN 'Reason Code Unavailable'
    WHEN A.CARRIER_ACTION_CD = 'EMPNE' THEN 'Employee not Eligible'
    WHEN A.CARRIER_ACTION_CD = 'ENDUP' THEN 'Reason Code Unavailable'
    WHEN A.CARRIER_ACTION_CD = 'ERRDNA' THEN 'Reason Code Unavailable'
    WHEN A.CARRIER_ACTION_CD = 'ERREOB' THEN 'Reason Code Unavailable'
    WHEN A.CARRIER_ACTION_CD = 'EXFREQ' THEN 'Service Exceeds Approved Frequency'
    WHEN A.CARRIER_ACTION_CD = 'EXPER' THEN 'Procedure Considered Experimental'
    WHEN A.CARRIER_ACTION_CD = 'EXRCR' THEN 'Exceeds Reasonable and Customary Rate'
    WHEN A.CARRIER_ACTION_CD = 'FRMLRY' THEN 'Product Not On Formulary'
    WHEN A.CARRIER_ACTION_CD = 'GENREJ' THEN 'Reason Code Unavailable'
    WHEN A.CARRIER_ACTION_CD = 'GRPNE' THEN 'Group Not Eligible'
    WHEN A.CARRIER_ACTION_CD = 'GRPSTL' THEN 'Reason Code Unavailable'
    WHEN A.CARRIER_ACTION_CD = 'HCPCS' THEN 'Reason Code Unavailable'
    WHEN A.CARRIER_ACTION_CD = 'HOSPIC' THEN 'Hospice Care Not Covered'
    WHEN A.CARRIER_ACTION_CD = 'ICD10' THEN 'Reason Code Unavailable'
    WHEN A.CARRIER_ACTION_CD = 'ICNPD' THEN 'DO NOT USE - Duplicate Claim Closed Out - ICN billed to and paid by different carrier (Auto Deny Job'
    WHEN A.CARRIER_ACTION_CD = 'ICOB' THEN 'COB Information Needed'
    WHEN A.CARRIER_ACTION_CD = 'IDAYS' THEN 'Missing/Invalid Days Supply'
    WHEN A.CARRIER_ACTION_CD = 'IDEA' THEN 'Missing/Invalid DEA Number'
    WHEN A.CARRIER_ACTION_CD = 'IDIG' THEN 'DX Code Required'
    WHEN A.CARRIER_ACTION_CD = 'IDOB' THEN 'Invalid/Missing DOB'
    WHEN A.CARRIER_ACTION_CD = 'IDOS' THEN 'Date(s) of Service Needed'
    WHEN A.CARRIER_ACTION_CD = 'IEOB' THEN 'Primary EOB Needed'
    WHEN A.CARRIER_ACTION_CD = 'ILAC' THEN 'Nature of Illness or Accident Required'
    WHEN A.CARRIER_ACTION_CD = 'IMCV' THEN 'Medicare EOB Required'
    WHEN A.CARRIER_ACTION_CD = 'INAME' THEN 'Invalid Character in Name or Illegible Name'
    WHEN A.CARRIER_ACTION_CD = 'INCL' THEN 'Carrier will not Pay Separately for this Service'
    WHEN A.CARRIER_ACTION_CD = 'INDC' THEN 'Missing or Invalid NDC Number'
    WHEN A.CARRIER_ACTION_CD = 'INFO' THEN 'Additional Information Needed'
    WHEN A.CARRIER_ACTION_CD = 'INGEX' THEN 'Reject - Ingredient Cost Exceeds Plan Maximum'
    WHEN A.CARRIER_ACTION_CD = 'INGMX' THEN 'Ingredient Cost Reduced to Maximum'
    WHEN A.CARRIER_ACTION_CD = 'INOPID' THEN 'Reason Code Unavailable'
    WHEN A.CARRIER_ACTION_CD = 'INPI' THEN 'Missing/Invalid NPI Number'
    WHEN A.CARRIER_ACTION_CD = 'INPRC' THEN 'Invalid Procedure Code'
    WHEN A.CARRIER_ACTION_CD = 'INSADD' THEN 'Reason Code Unavailable'
    WHEN A.CARRIER_ACTION_CD = 'INVADD' THEN 'Invalid/Incomplete Address'
    WHEN A.CARRIER_ACTION_CD = 'INVBN' THEN 'Missing/Invalid BIN Number'
    WHEN A.CARRIER_ACTION_CD = 'INVBT' THEN 'Invalid Bill Type'
    WHEN A.CARRIER_ACTION_CD = 'INVCD' THEN 'Invalid Code'
    WHEN A.CARRIER_ACTION_CD = 'INVDOS' THEN 'Invalid Date of Service'
    WHEN A.CARRIER_ACTION_CD = 'INVDT' THEN 'Invalid Admit Date'
    WHEN A.CARRIER_ACTION_CD = 'INVDX' THEN 'DX Not Valid for Procedure'
    WHEN A.CARRIER_ACTION_CD = 'INVGN' THEN 'Invalid Gender'
    WHEN A.CARRIER_ACTION_CD = 'INVGR' THEN 'UNKNOWN'
    WHEN A.CARRIER_ACTION_CD = 'INVGRP' THEN 'Invalid Group Number'
    WHEN A.CARRIER_ACTION_CD = 'INVID' THEN 'Invalid ID Number'
    WHEN A.CARRIER_ACTION_CD = 'INVMOD' THEN 'Missing/Invalid Modifier'
    WHEN A.CARRIER_ACTION_CD = 'INVNAM' THEN 'Invalid Name'
    WHEN A.CARRIER_ACTION_CD = 'INVND' THEN 'Non-Matched Service Provider Id'
    WHEN A.CARRIER_ACTION_CD = 'INVPCN' THEN 'Missing/Invalid PCN Number'
    WHEN A.CARRIER_ACTION_CD = 'INVPL' THEN 'Invalid or Incomplete Protocol Requirements'
    WHEN A.CARRIER_ACTION_CD = 'INVPOS' THEN 'Incorrect Place of Service'
    WHEN A.CARRIER_ACTION_CD = 'INVPY' THEN 'Invalid Payee Code for Medicaid Agency'
    WHEN A.CARRIER_ACTION_CD = 'INVQT' THEN 'Invalid Quantity Entered for Medication Package '
    WHEN A.CARRIER_ACTION_CD = 'INVUC' THEN 'Invalid Usual and Customary Amount'
    WHEN A.CARRIER_ACTION_CD = 'IPRO' THEN 'Provider Information Required'
    WHEN A.CARRIER_ACTION_CD = 'IREC' THEN 'Medical Records Needed'
    WHEN A.CARRIER_ACTION_CD = 'IREL' THEN 'Patient Relation to Insured Required'
    WHEN A.CARRIER_ACTION_CD = 'IRNPI' THEN 'Missing/Invalid Referring NPI'
    WHEN A.CARRIER_ACTION_CD = 'ITAX' THEN 'Need Tax ID Information'
    WHEN A.CARRIER_ACTION_CD = 'ITMZ' THEN 'Itemized Bill Required'
    WHEN A.CARRIER_ACTION_CD = 'KEYED' THEN 'Denial Code Keyed Incorrectly'
    WHEN A.CARRIER_ACTION_CD = 'LIABL' THEN 'Injury/Illness is Covered by the Liability Carrier'
    WHEN A.CARRIER_ACTION_CD = 'LINE' THEN 'Reason Code Unavailable'
    WHEN A.CARRIER_ACTION_CD = 'LOWDOL' THEN 'Reason Code Unavailable'
    WHEN A.CARRIER_ACTION_CD = 'M86' THEN 'Service denied because payment already made for same/similar procedure within set time frame.'
    WHEN A.CARRIER_ACTION_CD = 'MA04' THEN 'Secondary payment cannot be considered without the identity of or payment information from the primary payer. The information was either not reported or was illegible.'
    WHEN A.CARRIER_ACTION_CD = 'MA130' THEN 'Your claim contains incomplete and/or invalid information, and no appeal rights are afforded because the claim is unprocessable. Please submit a new claim with the complete/correct information.'
    WHEN A.CARRIER_ACTION_CD = 'MA66' THEN 'Missing/incomplete/invalid principal procedure code.'
    WHEN A.CARRIER_ACTION_CD = 'MAXBEN' THEN 'Maximum Benefits Reached'
    WHEN A.CARRIER_ACTION_CD = 'MAXBN' THEN 'UNKNOWN'
    WHEN A.CARRIER_ACTION_CD = 'MCAID' THEN 'Reason Code Unavailable'
    WHEN A.CARRIER_ACTION_CD = 'MEDAJ' THEN 'Medicaid Claim Adjudicated'
    WHEN A.CARRIER_ACTION_CD = 'MEDNCO' THEN 'Not a Medicare Covered Service'
    WHEN A.CARRIER_ACTION_CD = 'MEDNCOV' THEN 'UNKNOWN'
    WHEN A.CARRIER_ACTION_CD = 'MEDNEC' THEN 'Claim Not Medically Necessary'
    WHEN A.CARRIER_ACTION_CD = 'N122' THEN 'Add-on code cannot be billed by itself.'
    WHEN A.CARRIER_ACTION_CD = 'N130' THEN 'Consult plan benefit documents/guidelines for information about restrictions for this service.'
    WHEN A.CARRIER_ACTION_CD = 'N179' THEN 'Additional information has been requested from the member. The charges will be reconsidered upon receipt of that information.'
    WHEN A.CARRIER_ACTION_CD = 'N30' THEN 'Patient ineligible for this service.'
    WHEN A.CARRIER_ACTION_CD = 'N465' THEN 'Missing Physical Therapy Notes/Report.'
    WHEN A.CARRIER_ACTION_CD = 'N522' THEN 'Duplicate of a claim processed, or to be processed, as a crossover claim.'
    WHEN A.CARRIER_ACTION_CD = 'N525' THEN 'These services are not covered when performed within the global period of another service.'
    WHEN A.CARRIER_ACTION_CD = 'N578' THEN 'Coverages do not apply to this loss.'
    WHEN A.CARRIER_ACTION_CD = 'N674' THEN 'Not covered unless a pre-requisite procedure/service has been provided.'
    WHEN A.CARRIER_ACTION_CD = 'NABP' THEN 'Carrier Needs NABP Number'
    WHEN A.CARRIER_ACTION_CD = 'NBPT' THEN 'Non-Billable Provider Type'
    WHEN A.CARRIER_ACTION_CD = 'NDCNC' THEN 'NDC code Not Covered'
    WHEN A.CARRIER_ACTION_CD = 'NEICPD' THEN 'Payment Per Negotiated Rate'
    WHEN A.CARRIER_ACTION_CD = 'NENTBP' THEN 'Reason Code Unavailable'
    WHEN A.CARRIER_ACTION_CD = 'NENTIN' THEN 'Reason Code Unavailable'
    WHEN A.CARRIER_ACTION_CD = 'NENTPT' THEN 'Reason Code Unavailable'
    WHEN A.CARRIER_ACTION_CD = 'NENTRF' THEN 'Reason Code Unavailable'
    WHEN A.CARRIER_ACTION_CD = 'NOAUTH' THEN 'No Authorization Obtained'
    WHEN A.CARRIER_ACTION_CD = 'NOBIL' THEN 'Claim Should Not Have Been Billed per Client - Do Not Work/Rebill Claims'
    WHEN A.CARRIER_ACTION_CD = 'NOCOD' THEN 'No Denial Code on Remittance'
    WHEN A.CARRIER_ACTION_CD = 'NOCOV' THEN 'No Coverage'
    WHEN A.CARRIER_ACTION_CD = 'NOELG' THEN 'Reason Code Unavailable'
    WHEN A.CARRIER_ACTION_CD = 'NONURS' THEN 'Home Nursing Services Not Covered'
    WHEN A.CARRIER_ACTION_CD = 'NOPAY' THEN 'No Payment to be Issued'
    WHEN A.CARRIER_ACTION_CD = 'NOTES' THEN 'Physicians Orders or Nursing Notes Needed'
    WHEN A.CARRIER_ACTION_CD = 'NOXWK' THEN 'No-Crosswalk'
    WHEN A.CARRIER_ACTION_CD = 'NPHARM' THEN 'Non-Matched Pharmacy Number'
    WHEN A.CARRIER_ACTION_CD = 'NUBCCD' THEN 'Reason Code Unavailable'
    WHEN A.CARRIER_ACTION_CD = 'NUBCVL' THEN 'Reason Code Unavailable'
    WHEN A.CARRIER_ACTION_CD = 'OA133' THEN 'The disposition of this service line is pending further review. (Use only with Group Code OA). Usage: Use of this code requires a reversal and correction when the service line is finalized (use only in Loop 2110 CAS segment of the 835 or Loop 2430 of the 837).'
    WHEN A.CARRIER_ACTION_CD = 'OA18' THEN 'Exact duplicate claim/service (Use only with Group Code OA except where state workers compensation regulations requires CO)'
    WHEN A.CARRIER_ACTION_CD = 'OA23' THEN 'The impact of prior payer(s) adjudication including payments and/or adjustments. (Use only with Group Code OA)'
    WHEN A.CARRIER_ACTION_CD = 'OAB13' THEN 'Previously paid. Payment for this claim/service may have been provided in a previous payment.'
    WHEN A.CARRIER_ACTION_CD = 'OUTNW' THEN 'Out of Network'
    WHEN A.CARRIER_ACTION_CD = 'P' THEN 'Paid'
    WHEN A.CARRIER_ACTION_CD = 'PARTB' THEN 'Claim not Processed - Medicare Part B policy'
    WHEN A.CARRIER_ACTION_CD = 'PCFDNY' THEN 'Reason Code Unavailable'
    WHEN A.CARRIER_ACTION_CD = 'PCFVOID' THEN 'Reason Code Unavailable'
    WHEN A.CARRIER_ACTION_CD = 'PCPRE' THEN 'UNKNOWN'
    WHEN A.CARRIER_ACTION_CD = 'PDIA' THEN 'Invalid Principal/Admit Diagnosis'
    WHEN A.CARRIER_ACTION_CD = 'PDPVPB' THEN 'Paid to Provider Prior to HMS Billing'
    WHEN A.CARRIER_ACTION_CD = 'PDTOPV' THEN 'Claim Paid to Provider of Service'
    WHEN A.CARRIER_ACTION_CD = 'PEND' THEN 'Claim Pended'
    WHEN A.CARRIER_ACTION_CD = 'PHARNE' THEN 'Pharmacy Not Eligible'
    WHEN A.CARRIER_ACTION_CD = 'PHYNC' THEN 'Physician Not Covered '
    WHEN A.CARRIER_ACTION_CD = 'PR1' THEN 'Deductible Amount'
    WHEN A.CARRIER_ACTION_CD = 'PR20' THEN 'This injury/illness is covered by the liability carrier.'
    WHEN A.CARRIER_ACTION_CD = 'PR204' THEN 'This service/equipment/drug is not covered under the patients current benefit plan'
    WHEN A.CARRIER_ACTION_CD = 'PR27' THEN 'Expenses incurred after coverage terminated.'
    WHEN A.CARRIER_ACTION_CD = 'PR33' THEN 'Insured has no dependent coverage.'
    WHEN A.CARRIER_ACTION_CD = 'PR49' THEN 'This is a non-covered service because it is a routine/preventive exam or a diagnostic/screening procedure done in conjunction with a routine/preventive exam. Usage: Refer to the 835 Healthcare Policy Identification Segment (loop 2110 Service Payment Information REF), if present.'
    WHEN A.CARRIER_ACTION_CD = 'PRB7' THEN 'This provider was not certified/eligible to be paid for this procedure/service on this date of service. Usage: Refer to the 835 Healthcare Policy Identification Segment (loop 2110 Service Payment Information REF), if present.'
    WHEN A.CARRIER_ACTION_CD = 'PRCNC' THEN 'Procedure Code Not Covered'
    WHEN A.CARRIER_ACTION_CD = 'PRPRC' THEN 'Reason Code Unavailable'
    WHEN A.CARRIER_ACTION_CD = 'PRPRD' THEN 'Reason Code Unavailable'
    WHEN A.CARRIER_ACTION_CD = 'PRSC' THEN 'Expected a Prescriber Field'
    WHEN A.CARRIER_ACTION_CD = 'PRVNC' THEN 'Provider Not Covered'
    WHEN A.CARRIER_ACTION_CD = 'PTAGE' THEN 'Patient Outside Age Limit for this Type of Benefit'
    WHEN A.CARRIER_ACTION_CD = 'PTINFO' THEN 'Need Information From Patient'
    WHEN A.CARRIER_ACTION_CD = 'PTRES' THEN 'Patient Responsibility'
    WHEN A.CARRIER_ACTION_CD = 'QUAL' THEN 'Qualifying Procedure Not Received'
    WHEN A.CARRIER_ACTION_CD = 'QUANT' THEN 'Quantity Not Covered'
    WHEN A.CARRIER_ACTION_CD = 'REBIL' THEN 'Claim Closed to Pass through New Cycle'
    WHEN A.CARRIER_ACTION_CD = 'RECAL' THEN 'Reason Code Unavailable'
    WHEN A.CARRIER_ACTION_CD = 'RECALL' THEN 'Claims Closed - Completed Client Run Out Phase'
    WHEN A.CARRIER_ACTION_CD = 'REFIL' THEN 'Refill Too Soon'
    WHEN A.CARRIER_ACTION_CD = 'REFMI' THEN 'Refill Number Missing/Invalid'
    WHEN A.CARRIER_ACTION_CD = 'RENPRO' THEN 'Pending Information from Rendering Provider'
    WHEN A.CARRIER_ACTION_CD = 'RESENT' THEN 'Resent Claims'
    WHEN A.CARRIER_ACTION_CD = 'RESUB' THEN 'Submit Claim to Another Entity'
    WHEN A.CARRIER_ACTION_CD = 'REVCOD' THEN 'Reason Code Unavailable'
    WHEN A.CARRIER_ACTION_CD = 'RNID' THEN 'Reason Code Unavailable'
    WHEN A.CARRIER_ACTION_CD = 'RNNAP' THEN 'Reason Code Unavailable'
    WHEN A.CARRIER_ACTION_CD = 'ROMRAT' THEN 'Reason Code Unavailable'
    WHEN A.CARRIER_ACTION_CD = 'ROUTN' THEN 'Routine Services Not Covered'
    WHEN A.CARRIER_ACTION_CD = 'RXAGE' THEN 'Participant Age Restricts Medication Coverage'
    WHEN A.CARRIER_ACTION_CD = 'RXLMT' THEN 'Medication Exceeds Plan Limits'
    WHEN A.CARRIER_ACTION_CD = 'RXNC' THEN 'Prescription Drugs Not Covered'
    WHEN A.CARRIER_ACTION_CD = 'SETTL' THEN 'Claim Covered in Carrier Settlement'
    WHEN A.CARRIER_ACTION_CD = 'SPPBM' THEN 'Must Fill through Specialty Pharmacy'
    WHEN A.CARRIER_ACTION_CD = 'SUBID' THEN 'Reason Code Unavailable'
    WHEN A.CARRIER_ACTION_CD = 'TERMGRP' THEN 'Reason Code Unavailable'
    WHEN A.CARRIER_ACTION_CD = 'TIMELY' THEN 'Claim Past Timely Filing Limit'
    WHEN A.CARRIER_ACTION_CD = 'TOSNE' THEN 'Type of Service Not Eligible'
    WHEN A.CARRIER_ACTION_CD = 'TPA' THEN 'Send Claim to TPA'
    WHEN A.CARRIER_ACTION_CD = 'TPLNC' THEN 'Group Does Not Allow Third Party Claims'
    WHEN A.CARRIER_ACTION_CD = 'UMID' THEN 'UNKNOWN'
    WHEN A.CARRIER_ACTION_CD = 'UNID' THEN 'Unable to Identify Member'
    WHEN A.CARRIER_ACTION_CD = 'UNITS' THEN 'Units Field Invalid for Number of Days'
    WHEN A.CARRIER_ACTION_CD = 'UNPRO' THEN 'Unprocessed Claim'
    WHEN A.CARRIER_ACTION_CD = 'WASTE' THEN 'Reason Code Unavailable'
    WHEN A.CARRIER_ACTION_CD = '' THEN 'No Data Provided'
    ELSE 'NOCOD'
            END  AS   "CARRIER DESCRIPTION1",
            
    C.CARRIER_ACTION_CD2 AS "CARRIER ACTION CD2",
    CASE 
    WHEN C.CARRIER_ACTION_CD2 = '277ACC' THEN '277 ACCEPTANCE/ACKNOWLEDGMENT'
    WHEN C.CARRIER_ACTION_CD2 = '-999' THEN 'Reason Code Unavailable'
    WHEN C.CARRIER_ACTION_CD2 = 'ACCDT' THEN 'Reason Code Unavailable'
    WHEN C.CARRIER_ACTION_CD2 = 'ADJUD' THEN 'Adjudicated per Plan Contract/Allowable'
    WHEN C.CARRIER_ACTION_CD2 = 'APID' THEN 'Reason Code Unavailable'
    WHEN C.CARRIER_ACTION_CD2 = 'ARDUP' THEN 'Claim is a Duplicate of Previously Billed Claim'
    WHEN C.CARRIER_ACTION_CD2 = 'B2BREJ' THEN 'Reason Code Unavailable'
    WHEN C.CARRIER_ACTION_CD2 = 'BADCLMS' THEN 'Reason Code Unavailable'
    WHEN C.CARRIER_ACTION_CD2 = 'BADEL' THEN 'Reason Code Unavailable'
    WHEN C.CARRIER_ACTION_CD2 = 'BILLER' THEN 'Billing Error'
    WHEN C.CARRIER_ACTION_CD2 = 'BPNAP' THEN 'Reason Code Unavailable'
    WHEN C.CARRIER_ACTION_CD2 = 'BPNPI' THEN 'Reason Code Unavailable'
    WHEN C.CARRIER_ACTION_CD2 = 'BPTAX' THEN 'Reason Code Unavailable'
    WHEN C.CARRIER_ACTION_CD2 = 'BPTXN' THEN 'Reason Code Unavailable'
    WHEN C.CARRIER_ACTION_CD2 = 'CAID' THEN 'Reason Code Unavailable'
    WHEN C.CARRIER_ACTION_CD2 = 'CANCEL' THEN 'Reason Code Unavailable'
    WHEN C.CARRIER_ACTION_CD2 = 'CAPIT' THEN 'Capitated Service'
    WHEN C.CARRIER_ACTION_CD2 = 'CCLM' THEN 'Corrected Claim Needed'
    WHEN C.CARRIER_ACTION_CD2 = 'CLAMAJ' THEN 'Care may not be Covered by Another Payer per Coordination of Benefits'
    WHEN C.CARRIER_ACTION_CD2 = 'CLMAJ' THEN 'Claim Adjusted'
    WHEN C.CARRIER_ACTION_CD2 = 'CLMFRQ' THEN 'Reason Code Unavailable'
    WHEN C.CARRIER_ACTION_CD2 = 'CLMFWD' THEN 'Claim Forwarded to Payer by NEIC'
    WHEN C.CARRIER_ACTION_CD2 = 'CMPMI' THEN 'Missing/Invalid Compound Code'
    WHEN C.CARRIER_ACTION_CD2 = 'CNBC' THEN 'Reason Code Unavailable'
    WHEN C.CARRIER_ACTION_CD2 = 'CO45' THEN 'Charge exceeds fee schedule/maximum allowable or contracted/legislated fee arrangement. Usage: This adjustment amount cannot equal the total service or claim charge amount; and must not duplicate provider adjustment amounts (payments and contractual reductions) that have resulted from prior payer(s) adjudication. (Use only with Group Codes PR or CO depending upon liability)'
    WHEN C.CARRIER_ACTION_CD2 = 'COBPD' THEN 'Paid in Accordance with COB'
    WHEN C.CARRIER_ACTION_CD2 = 'COINS' THEN 'MA Paid Less Than Co-Insurance'
    WHEN C.CARRIER_ACTION_CD2 = 'COPAY' THEN 'MA Paid Less Than Copay'
    WHEN C.CARRIER_ACTION_CD2 = 'DAVITA' THEN 'Reason Code Unavailable'
    WHEN C.CARRIER_ACTION_CD2 = 'DAYSUP' THEN 'Days Supply Exceeds Plan Limits'
    WHEN C.CARRIER_ACTION_CD2 = 'DEDUC' THEN 'Payment Applied to Patient Deductible'
    WHEN C.CARRIER_ACTION_CD2 = 'DEFER' THEN 'Portion of Payment Deferred'
    WHEN C.CARRIER_ACTION_CD2 = 'DENTL' THEN 'Group Has Dental Coverage Only'
    WHEN C.CARRIER_ACTION_CD2 = 'DEPNE' THEN 'Dependent Not Eligible'
    WHEN C.CARRIER_ACTION_CD2 = 'DEPPG' THEN 'Dependent Pregnancy Not Covered'
    WHEN C.CARRIER_ACTION_CD2 = 'DISDT' THEN 'Reason Code Unavailable'
    WHEN C.CARRIER_ACTION_CD2 = 'DISSTT' THEN 'Reason Code Unavailable'
    WHEN C.CARRIER_ACTION_CD2 = 'DOSNE' THEN 'Patient Not Eligible on Date of Service'
    WHEN C.CARRIER_ACTION_CD2 = 'DPWMI' THEN 'Date Prescription Written Missing/Invalid'
    WHEN C.CARRIER_ACTION_CD2 = 'DRGNC' THEN 'Drug Not Covered'
    WHEN C.CARRIER_ACTION_CD2 = 'DSPMI' THEN 'Missing/Invalid Dispense as Written Code'
    WHEN C.CARRIER_ACTION_CD2 = 'DUP' THEN 'Duplicate Claim Submission'
    WHEN C.CARRIER_ACTION_CD2 = 'DURCF' THEN 'Insert Fail DUR-Conflict'
    WHEN C.CARRIER_ACTION_CD2 = 'DXNC' THEN 'Diagnosis Not Covered'
    WHEN C.CARRIER_ACTION_CD2 = 'EDITGOV' THEN 'Reason Code Unavailable'
    WHEN C.CARRIER_ACTION_CD2 = 'EMPNE' THEN 'Employee not Eligible'
    WHEN C.CARRIER_ACTION_CD2 = 'ENDUP' THEN 'Reason Code Unavailable'
    WHEN C.CARRIER_ACTION_CD2 = 'ERRDNA' THEN 'Reason Code Unavailable'
    WHEN C.CARRIER_ACTION_CD2 = 'ERREOB' THEN 'Reason Code Unavailable'
    WHEN C.CARRIER_ACTION_CD2 = 'EXFREQ' THEN 'Service Exceeds Approved Frequency'
    WHEN C.CARRIER_ACTION_CD2 = 'EXPER' THEN 'Procedure Considered Experimental'
    WHEN C.CARRIER_ACTION_CD2 = 'EXRCR' THEN 'Exceeds Reasonable and Customary Rate'
    WHEN C.CARRIER_ACTION_CD2 = 'FRMLRY' THEN 'Product Not On Formulary'
    WHEN C.CARRIER_ACTION_CD2 = 'GENREJ' THEN 'Reason Code Unavailable'
    WHEN C.CARRIER_ACTION_CD2 = 'GRPNE' THEN 'Group Not Eligible'
    WHEN C.CARRIER_ACTION_CD2 = 'GRPSTL' THEN 'Reason Code Unavailable'
    WHEN C.CARRIER_ACTION_CD2 = 'HCPCS' THEN 'Reason Code Unavailable'
    WHEN C.CARRIER_ACTION_CD2 = 'HOSPIC' THEN 'Hospice Care Not Covered'
    WHEN C.CARRIER_ACTION_CD2 = 'ICD10' THEN 'Reason Code Unavailable'
    WHEN C.CARRIER_ACTION_CD2 = 'ICNPD' THEN 'DO NOT USE - Duplicate Claim Closed Out - ICN billed to and paid by different carrier (Auto Deny Job'
    WHEN C.CARRIER_ACTION_CD2 = 'ICOB' THEN 'COB Information Needed'
    WHEN C.CARRIER_ACTION_CD2 = 'IDAYS' THEN 'Missing/Invalid Days Supply'
    WHEN C.CARRIER_ACTION_CD2 = 'IDEA' THEN 'Missing/Invalid DEA Number'
    WHEN C.CARRIER_ACTION_CD2 = 'IDIG' THEN 'DX Code Required'
    WHEN C.CARRIER_ACTION_CD2 = 'IDOB' THEN 'Invalid/Missing DOB'
    WHEN C.CARRIER_ACTION_CD2 = 'IDOS' THEN 'Date(s) of Service Needed'
    WHEN C.CARRIER_ACTION_CD2 = 'IEOB' THEN 'Primary EOB Needed'
    WHEN C.CARRIER_ACTION_CD2 = 'ILAC' THEN 'Nature of Illness or Accident Required'
    WHEN C.CARRIER_ACTION_CD2 = 'IMCV' THEN 'Medicare EOB Required'
    WHEN C.CARRIER_ACTION_CD2 = 'INAME' THEN 'Invalid Character in Name or Illegible Name'
    WHEN C.CARRIER_ACTION_CD2 = 'INCL' THEN 'Carrier will not Pay Separately for this Service'
    WHEN C.CARRIER_ACTION_CD2 = 'INDC' THEN 'Missing or Invalid NDC Number'
    WHEN C.CARRIER_ACTION_CD2 = 'INFO' THEN 'Additional Information Needed'
    WHEN C.CARRIER_ACTION_CD2 = 'INGEX' THEN 'Reject - Ingredient Cost Exceeds Plan Maximum'
    WHEN C.CARRIER_ACTION_CD2 = 'INGMX' THEN 'Ingredient Cost Reduced to Maximum'
    WHEN C.CARRIER_ACTION_CD2 = 'INOPID' THEN 'Reason Code Unavailable'
    WHEN C.CARRIER_ACTION_CD2 = 'INPI' THEN 'Missing/Invalid NPI Number'
    WHEN C.CARRIER_ACTION_CD2 = 'INPRC' THEN 'Invalid Procedure Code'
    WHEN C.CARRIER_ACTION_CD2 = 'INSADD' THEN 'Reason Code Unavailable'
    WHEN C.CARRIER_ACTION_CD2 = 'INVADD' THEN 'Invalid/Incomplete Address'
    WHEN C.CARRIER_ACTION_CD2 = 'INVBN' THEN 'Missing/Invalid BIN Number'
    WHEN C.CARRIER_ACTION_CD2 = 'INVBT' THEN 'Invalid Bill Type'
    WHEN C.CARRIER_ACTION_CD2 = 'INVCD' THEN 'Invalid Code'
    WHEN C.CARRIER_ACTION_CD2 = 'INVDOS' THEN 'Invalid Date of Service'
    WHEN C.CARRIER_ACTION_CD2 = 'INVDT' THEN 'Invalid Admit Date'
    WHEN C.CARRIER_ACTION_CD2 = 'INVDX' THEN 'DX Not Valid for Procedure'
    WHEN C.CARRIER_ACTION_CD2 = 'INVGN' THEN 'Invalid Gender'
    WHEN C.CARRIER_ACTION_CD2 = 'INVGR' THEN 'UNKNOWN'
    WHEN C.CARRIER_ACTION_CD2 = 'INVGRP' THEN 'Invalid Group Number'
    WHEN C.CARRIER_ACTION_CD2 = 'INVID' THEN 'Invalid ID Number'
    WHEN C.CARRIER_ACTION_CD2 = 'INVMOD' THEN 'Missing/Invalid Modifier'
    WHEN C.CARRIER_ACTION_CD2 = 'INVNAM' THEN 'Invalid Name'
    WHEN C.CARRIER_ACTION_CD2 = 'INVND' THEN 'Non-Matched Service Provider Id'
    WHEN C.CARRIER_ACTION_CD2 = 'INVPCN' THEN 'Missing/Invalid PCN Number'
    WHEN C.CARRIER_ACTION_CD2 = 'INVPL' THEN 'Invalid or Incomplete Protocol Requirements'
    WHEN C.CARRIER_ACTION_CD2 = 'INVPOS' THEN 'Incorrect Place of Service'
    WHEN C.CARRIER_ACTION_CD2 = 'INVPY' THEN 'Invalid Payee Code for Medicaid Agency'
    WHEN C.CARRIER_ACTION_CD2 = 'INVQT' THEN 'Invalid Quantity Entered for Medication Package '
    WHEN C.CARRIER_ACTION_CD2 = 'INVUC' THEN 'Invalid Usual and Customary Amount'
    WHEN C.CARRIER_ACTION_CD2 = 'IPRO' THEN 'Provider Information Required'
    WHEN C.CARRIER_ACTION_CD2 = 'IREC' THEN 'Medical Records Needed'
    WHEN C.CARRIER_ACTION_CD2 = 'IREL' THEN 'Patient Relation to Insured Required'
    WHEN C.CARRIER_ACTION_CD2 = 'IRNPI' THEN 'Missing/Invalid Referring NPI'
    WHEN C.CARRIER_ACTION_CD2 = 'ITAX' THEN 'Need Tax ID Information'
    WHEN C.CARRIER_ACTION_CD2 = 'ITMZ' THEN 'Itemized Bill Required'
    WHEN C.CARRIER_ACTION_CD2 = 'KEYED' THEN 'Denial Code Keyed Incorrectly'
    WHEN C.CARRIER_ACTION_CD2 = 'LIABL' THEN 'Injury/Illness is Covered by the Liability Carrier'
    WHEN C.CARRIER_ACTION_CD2 = 'LINE' THEN 'Reason Code Unavailable'
    WHEN C.CARRIER_ACTION_CD2 = 'LOWDOL' THEN 'Reason Code Unavailable'
    WHEN C.CARRIER_ACTION_CD2 = 'M86' THEN 'Service denied because payment already made for same/similar procedure within set time frame.'
    WHEN C.CARRIER_ACTION_CD2 = 'MA04' THEN 'Secondary payment cannot be considered without the identity of or payment information from the primary payer. The information was either not reported or was illegible.'
    WHEN C.CARRIER_ACTION_CD2 = 'MA130' THEN 'Your claim contains incomplete and/or invalid information, and no appeal rights are afforded because the claim is unprocessable. Please submit a new claim with the complete/correct information.'
    WHEN C.CARRIER_ACTION_CD2 = 'MA66' THEN 'Missing/incomplete/invalid principal procedure code.'
    WHEN C.CARRIER_ACTION_CD2 = 'MAXBEN' THEN 'Maximum Benefits Reached'
    WHEN C.CARRIER_ACTION_CD2 = 'MAXBN' THEN 'UNKNOWN'
    WHEN C.CARRIER_ACTION_CD2 = 'MCAID' THEN 'Reason Code Unavailable'
    WHEN C.CARRIER_ACTION_CD2 = 'MEDAJ' THEN 'Medicaid Claim Adjudicated'
    WHEN C.CARRIER_ACTION_CD2 = 'MEDNCO' THEN 'Not a Medicare Covered Service'
    WHEN C.CARRIER_ACTION_CD2 = 'MEDNCOV' THEN 'UNKNOWN'
    WHEN C.CARRIER_ACTION_CD2 = 'MEDNEC' THEN 'Claim Not Medically Necessary'
    WHEN C.CARRIER_ACTION_CD2 = 'N122' THEN 'Add-on code cannot be billed by itself.'
    WHEN C.CARRIER_ACTION_CD2 = 'N130' THEN 'Consult plan benefit documents/guidelines for information about restrictions for this service.'
    WHEN C.CARRIER_ACTION_CD2 = 'N179' THEN 'Additional information has been requested from the member. The charges will be reconsidered upon receipt of that information.'
    WHEN C.CARRIER_ACTION_CD2 = 'N30' THEN 'Patient ineligible for this service.'
    WHEN C.CARRIER_ACTION_CD2 = 'N465' THEN 'Missing Physical Therapy Notes/Report.'
    WHEN C.CARRIER_ACTION_CD2 = 'N522' THEN 'Duplicate of a claim processed, or to be processed, as a crossover claim.'
    WHEN C.CARRIER_ACTION_CD2 = 'N525' THEN 'These services are not covered when performed within the global period of another service.'
    WHEN C.CARRIER_ACTION_CD2 = 'N578' THEN 'Coverages do not apply to this loss.'
    WHEN C.CARRIER_ACTION_CD2 = 'N674' THEN 'Not covered unless a pre-requisite procedure/service has been provided.'
    WHEN C.CARRIER_ACTION_CD2 = 'NABP' THEN 'Carrier Needs NABP Number'
    WHEN C.CARRIER_ACTION_CD2 = 'NBPT' THEN 'Non-Billable Provider Type'
    WHEN C.CARRIER_ACTION_CD2 = 'NDCNC' THEN 'NDC code Not Covered'
    WHEN C.CARRIER_ACTION_CD2 = 'NEICPD' THEN 'Payment Per Negotiated Rate'
    WHEN C.CARRIER_ACTION_CD2 = 'NENTBP' THEN 'Reason Code Unavailable'
    WHEN C.CARRIER_ACTION_CD2 = 'NENTIN' THEN 'Reason Code Unavailable'
    WHEN C.CARRIER_ACTION_CD2 = 'NENTPT' THEN 'Reason Code Unavailable'
    WHEN C.CARRIER_ACTION_CD2 = 'NENTRF' THEN 'Reason Code Unavailable'
    WHEN C.CARRIER_ACTION_CD2 = 'NOAUTH' THEN 'No Authorization Obtained'
    WHEN C.CARRIER_ACTION_CD2 = 'NOBIL' THEN 'Claim Should Not Have Been Billed per Client - Do Not Work/Rebill Claims'
    WHEN C.CARRIER_ACTION_CD2 = 'NOCOD' THEN 'No Denial Code on Remittance'
    WHEN C.CARRIER_ACTION_CD2 = 'NOCOV' THEN 'No Coverage'
    WHEN C.CARRIER_ACTION_CD2 = 'NOELG' THEN 'Reason Code Unavailable'
    WHEN C.CARRIER_ACTION_CD2 = 'NONURS' THEN 'Home Nursing Services Not Covered'
    WHEN C.CARRIER_ACTION_CD2 = 'NOPAY' THEN 'No Payment to be Issued'
    WHEN C.CARRIER_ACTION_CD2 = 'NOTES' THEN 'Physicians Orders or Nursing Notes Needed'
    WHEN C.CARRIER_ACTION_CD2 = 'NOXWK' THEN 'No-Crosswalk'
    WHEN C.CARRIER_ACTION_CD2 = 'NPHARM' THEN 'Non-Matched Pharmacy Number'
    WHEN C.CARRIER_ACTION_CD2 = 'NUBCCD' THEN 'Reason Code Unavailable'
    WHEN C.CARRIER_ACTION_CD2 = 'NUBCVL' THEN 'Reason Code Unavailable'
    WHEN C.CARRIER_ACTION_CD2 = 'OA133' THEN 'The disposition of this service line is pending further review. (Use only with Group Code OA). Usage: Use of this code requires a reversal and correction when the service line is finalized (use only in Loop 2110 CAS segment of the 835 or Loop 2430 of the 837).'
    WHEN C.CARRIER_ACTION_CD2 = 'OA18' THEN 'Exact duplicate claim/service (Use only with Group Code OA except where state workers compensation regulations requires CO)'
    WHEN C.CARRIER_ACTION_CD2 = 'OA23' THEN 'The impact of prior payer(s) adjudication including payments and/or adjustments. (Use only with Group Code OA)'
    WHEN C.CARRIER_ACTION_CD2 = 'OAB13' THEN 'Previously paid. Payment for this claim/service may have been provided in a previous payment.'
    WHEN C.CARRIER_ACTION_CD2 = 'OUTNW' THEN 'Out of Network'
    WHEN C.CARRIER_ACTION_CD2 = 'P' THEN 'Paid'
    WHEN C.CARRIER_ACTION_CD2 = 'PARTB' THEN 'Claim not Processed - Medicare Part B policy'
    WHEN C.CARRIER_ACTION_CD2 = 'PCFDNY' THEN 'Reason Code Unavailable'
    WHEN C.CARRIER_ACTION_CD2 = 'PCFVOID' THEN 'Reason Code Unavailable'
    WHEN C.CARRIER_ACTION_CD2 = 'PCPRE' THEN 'UNKNOWN'
    WHEN C.CARRIER_ACTION_CD2 = 'PDIA' THEN 'Invalid Principal/Admit Diagnosis'
    WHEN C.CARRIER_ACTION_CD2 = 'PDPVPB' THEN 'Paid to Provider Prior to HMS Billing'
    WHEN C.CARRIER_ACTION_CD2 = 'PDTOPV' THEN 'Claim Paid to Provider of Service'
    WHEN C.CARRIER_ACTION_CD2 = 'PEND' THEN 'Claim Pended'
    WHEN C.CARRIER_ACTION_CD2 = 'PHARNE' THEN 'Pharmacy Not Eligible'
    WHEN C.CARRIER_ACTION_CD2 = 'PHYNC' THEN 'Physician Not Covered '
    WHEN C.CARRIER_ACTION_CD2 = 'PR1' THEN 'Deductible Amount'
    WHEN C.CARRIER_ACTION_CD2 = 'PR20' THEN 'This injury/illness is covered by the liability carrier.'
    WHEN C.CARRIER_ACTION_CD2 = 'PR204' THEN 'This service/equipment/drug is not covered under the patients current benefit plan'
    WHEN C.CARRIER_ACTION_CD2 = 'PR27' THEN 'Expenses incurred after coverage terminated.'
    WHEN C.CARRIER_ACTION_CD2 = 'PR33' THEN 'Insured has no dependent coverage.'
    WHEN C.CARRIER_ACTION_CD2 = 'PR49' THEN 'This is a non-covered service because it is a routine/preventive exam or a diagnostic/screening procedure done in conjunction with a routine/preventive exam. Usage: Refer to the 835 Healthcare Policy Identification Segment (loop 2110 Service Payment Information REF), if present.'
    WHEN C.CARRIER_ACTION_CD2 = 'PRB7' THEN 'This provider was not certified/eligible to be paid for this procedure/service on this date of service. Usage: Refer to the 835 Healthcare Policy Identification Segment (loop 2110 Service Payment Information REF), if present.'
    WHEN C.CARRIER_ACTION_CD2 = 'PRCNC' THEN 'Procedure Code Not Covered'
    WHEN C.CARRIER_ACTION_CD2 = 'PRPRC' THEN 'Reason Code Unavailable'
    WHEN C.CARRIER_ACTION_CD2 = 'PRPRD' THEN 'Reason Code Unavailable'
    WHEN C.CARRIER_ACTION_CD2 = 'PRSC' THEN 'Expected a Prescriber Field'
    WHEN C.CARRIER_ACTION_CD2 = 'PRVNC' THEN 'Provider Not Covered'
    WHEN C.CARRIER_ACTION_CD2 = 'PTAGE' THEN 'Patient Outside Age Limit for this Type of Benefit'
    WHEN C.CARRIER_ACTION_CD2 = 'PTINFO' THEN 'Need Information From Patient'
    WHEN C.CARRIER_ACTION_CD2 = 'PTRES' THEN 'Patient Responsibility'
    WHEN C.CARRIER_ACTION_CD2 = 'QUAL' THEN 'Qualifying Procedure Not Received'
    WHEN C.CARRIER_ACTION_CD2 = 'QUANT' THEN 'Quantity Not Covered'
    WHEN C.CARRIER_ACTION_CD2 = 'REBIL' THEN 'Claim Closed to Pass through New Cycle'
    WHEN C.CARRIER_ACTION_CD2 = 'RECAL' THEN 'Reason Code Unavailable'
    WHEN C.CARRIER_ACTION_CD2 = 'RECALL' THEN 'Claims Closed - Completed Client Run Out Phase'
    WHEN C.CARRIER_ACTION_CD2 = 'REFIL' THEN 'Refill Too Soon'
    WHEN C.CARRIER_ACTION_CD2 = 'REFMI' THEN 'Refill Number Missing/Invalid'
    WHEN C.CARRIER_ACTION_CD2 = 'RENPRO' THEN 'Pending Information from Rendering Provider'
    WHEN C.CARRIER_ACTION_CD2 = 'RESENT' THEN 'Resent Claims'
    WHEN C.CARRIER_ACTION_CD2 = 'RESUB' THEN 'Submit Claim to Another Entity'
    WHEN C.CARRIER_ACTION_CD2 = 'REVCOD' THEN 'Reason Code Unavailable'
    WHEN C.CARRIER_ACTION_CD2 = 'RNID' THEN 'Reason Code Unavailable'
    WHEN C.CARRIER_ACTION_CD2 = 'RNNAP' THEN 'Reason Code Unavailable'
    WHEN C.CARRIER_ACTION_CD2 = 'ROMRAT' THEN 'Reason Code Unavailable'
    WHEN C.CARRIER_ACTION_CD2 = 'ROUTN' THEN 'Routine Services Not Covered'
    WHEN C.CARRIER_ACTION_CD2 = 'RXAGE' THEN 'Participant Age Restricts Medication Coverage'
    WHEN C.CARRIER_ACTION_CD2 = 'RXLMT' THEN 'Medication Exceeds Plan Limits'
    WHEN C.CARRIER_ACTION_CD2 = 'RXNC' THEN 'Prescription Drugs Not Covered'
    WHEN C.CARRIER_ACTION_CD2 = 'SETTL' THEN 'Claim Covered in Carrier Settlement'
    WHEN C.CARRIER_ACTION_CD2 = 'SPPBM' THEN 'Must Fill through Specialty Pharmacy'
    WHEN C.CARRIER_ACTION_CD2 = 'SUBID' THEN 'Reason Code Unavailable'
    WHEN C.CARRIER_ACTION_CD2 = 'TERMGRP' THEN 'Reason Code Unavailable'
    WHEN C.CARRIER_ACTION_CD2 = 'TIMELY' THEN 'Claim Past Timely Filing Limit'
    WHEN C.CARRIER_ACTION_CD2 = 'TOSNE' THEN 'Type of Service Not Eligible'
    WHEN C.CARRIER_ACTION_CD2 = 'TPA' THEN 'Send Claim to TPA'
    WHEN C.CARRIER_ACTION_CD2 = 'TPLNC' THEN 'Group Does Not Allow Third Party Claims'
    WHEN C.CARRIER_ACTION_CD2 = 'UMID' THEN 'UNKNOWN'
    WHEN C.CARRIER_ACTION_CD2 = 'UNID' THEN 'Unable to Identify Member'
    WHEN C.CARRIER_ACTION_CD2 = 'UNITS' THEN 'Units Field Invalid for Number of Days'
    WHEN C.CARRIER_ACTION_CD2 = 'UNPRO' THEN 'Unprocessed Claim'
    WHEN C.CARRIER_ACTION_CD2 = 'WASTE' THEN 'Reason Code Unavailable'
    WHEN C.CARRIER_ACTION_CD2 = '' THEN 'Reason Code Unavailable'
    ELSE 'NOCOD'
            END  AS   "CARRIER DESCRIPTION2",
        
        CASE WHEN A.ACTION_CD IN ('1100', '1109', '147C', '2100', '210B', '24OLD', '277ACC', '277REJ', '7U00', 'ABATE', 'ACCDX', 'ADDR', 'ADMIN', 'AETNA', 'AGED', 'AGED1', 'ALCARE', 'ARSRC', 'B2BREJ', 'BABFL', 'BABY', 'BCLCL', 'BILLE', 'BNFAD', 'BNFAS', 'BNFPB', 'CCLM', 'CERT', 'CGCIG', 'CLAMAJ', 'CLEAN', 'CLMFM', 'CLMFW', 'CLMFWD', 'CNTID', 'COBPD', 'CPT', 'D', 'DRSTM', 'DUA', 'DUP', 'DUPHMS', 'EFT', 'ELEC', 'EMPCER', 'EOMB', 'ERRDES', 'ERRDNA', 'ERREOB', 'EXPER', 'FAULT', 'FUTFL', 'GENREJ', 'HCFA', 'HMSA', 'HMSCLF', 'HMSCPT', 'HMSDIG', 'HMSDOS', 'HMSEOB', 'HMSI', 'HMSIC', 'HMSMCV', 'HMSN', 'HMSNC', 'HMSNPI', 'HMSO', 'HMSP', 'HMSPE', 'HMSPND', 'HMSPRO', 'HMSPTP', 'HMSPV', 'HMSR', 'HMSREC', 'HMSRJ', 'HMSSTU', 'HMSTMZ', 'HPSC', 'HRA', 'HSACC', 'IACC', 'IANPI', 'IBNPI', 'ICER', 'ICERT', 'ICLF', 'ICOB', 'ICOU', 'ICPD', 'ICPT', 'IDAYS', 'IDEA', 'IDIG', 'IDOB', 'IDOS', 'IEOB', 'IGRP', 'ILAC', 'ILLEG', 'IMAP', 'IMCV', 'INADD', 'INAME', 'INFO', 'INOPID', 'INPI', 'INPRC', 'INSADD', 'INVAD', 'INVADD', 'INVAM', 'INVBN', 'INVBT', 'INVCD', 'INVDOB', 'INVDOS', 'INVDS', 'INVDT', 'INVDX', 'INVES', 'INVGR', 'INVGRP', 'INVID', 'INVLDADD', 'INVLDCAR', 'INVLDFMT', 'INVLDOF', 'INVMOD', 'INVNAM', 'INVND', 'INVOF', 'INVPCN', 'INVPIN', 'INVPOS', 'IPAD', 'IPCS', 'IPOA', 'IPRO', 'IPRS', 'IPTA', 'IREC', 'IREL', 'IRXST', 'ISNPI', 'ISTU', 'ITAX', 'ITMZ', 'ITPL', 'KEYED', 'LEGAL', 'LIABL', 'LOST', 'LWR21', 'MAGEL', 'MCAID', 'MCPRID', 'MCV', 'MDCLM', 'MEDREC', 'MEDSUP', 'MODIF', 'NAS', 'NCRPT', 'NCRTP', 'NENTAP', 'NENTBP', 'NENTRF', 'NENTRP', 'NENTSF', 'NOAUT', 'NOAUTH', 'NOCOB', 'NOCOD', 'NOINFO', 'NOINT', 'NOPAP', 'NOTES', 'NOXWK', 'NSXWK', 'NWAIT', 'OPN', 'OTNWK', 'OUTNW', 'PAPER', 'PAYCL', 'PBM', 'PBMPD', 'PCPRE', 'PDIA', 'PEND', 'PND', 'PRCINF', 'PREAU', 'PREAUT', 'PRIMID', 'PTADD', 'PTINFO', 'PURCHASED', 'QUAL', 'REDWT', 'RENNPI', 'RENPRO', 'REPRI', 'RESENT', 'RESUB', 'REVCD', 'RFINF', 'RSCINFO', 'SBAPS', 'SPLIT', 'SPPBM', 'STDNT', 'SUBMH', 'SUBNAP', 'SURDT', 'TARS', 'TFLCL', 'TIMEL', 'TIMELY', 'TIMRX', 'TOOTH', 'TPA', 'TPLNC', 'TPLNE', 'UB92', 'UMGRP', 'UMID', 'UNITS', 'UNPRO', 'VALOP', 'ZPDPR', 'ZPDRS' )
        --C.CARRIER_ACTION_CD2 IN ('277ACC', '999', 'ACCDT', 'APID', 'B2BREJ', 'BADCLMS', 'BADEL', 'BPNAP', 'BPNPI', 'BPTAX', 'BPTXN', 'CAID', 'CANCEL', 'CCLM', 'CLAMAJ', 'CLMFRQ', 'CLMFWD', 'CNBC', 'COBPD', 'DAVITA', 'DISDT', 'DISSTT', 'DUP', 'EDITGOV', 'ENDUP', 'ERRDNA', 'ERREOB', 'EXPER', 'GENREJ', 'GRPSTL', 'HCPCS', 'ICD10', 'ICOB', 'IDAYS', 'IDEA', 'IDIG', 'IDOB', 'IDOS', 'IEOB', 'ILAC', 'IMCV', 'INAME', 'INFO', 'INOPID', 'INPI', 'INPRC', 'INSADD', 'INVADD', 'INVBN', 'INVBT', 'INVCD', 'INVDOS', 'INVDT', 'INVDX', 'INVGR', 'INVGRP', 'INVID', 'INVMOD', 'INVNAM', 'INVND', 'INVPCN', 'INVPOS', 'IPRO', 'IREC', 'IREL', 'ITAX', 'ITMZ', 'KEYED', 'LIABL', 'LINE', 'LOWDOL', 'MAXBN', 'MCAID', 'MEDNCOV', 'NENTBP', 'NENTIN', 'NENTPT', 'NENTRF', 'NOAUTH', 'NOCOD', 'NOELG', 'NOTES', 'NOXWK', 'NUBCCD', 'NUBCVL', 'OUTNW', 'PCFDNY', 'PCFVOID', 'PCPRE', 'PDIA', 'PEND', 'PRPRC', 'PRPRD', 'PTINFO', 'QUAL', 'RECAL', 'RENPRO', 'RESENT', 'RESUB', 'REVCOD', 'RNID', 'RNNAP', 'ROMRAT', 'SPPBM', 'SUBID', 'TERMGRP', 'TIMELY', 'TPA', 'TPLNC', 'UMID', 'UNITS', 'UNPRO', 'WASTE' )
        --OR C.CARRIER_ACTION_CD2 IN ('277ACC', '999', 'ACCDT', 'APID', 'B2BREJ', 'BADCLMS', 'BADEL', 'BPNAP', 'BPNPI', 'BPTAX', 'BPTXN', 'CAID', 'CANCEL', 'CCLM', 'CLAMAJ', 'CLMFRQ', 'CLMFWD', 'CNBC', 'COBPD', 'DAVITA', 'DISDT', 'DISSTT', 'DUP', 'EDITGOV', 'ENDUP', 'ERRDNA', 'ERREOB', 'EXPER', 'GENREJ', 'GRPSTL', 'HCPCS', 'ICD10', 'ICOB', 'IDAYS', 'IDEA', 'IDIG', 'IDOB', 'IDOS', 'IEOB', 'ILAC', 'IMCV', 'INAME', 'INFO', 'INOPID', 'INPI', 'INPRC', 'INSADD', 'INVADD', 'INVBN', 'INVBT', 'INVCD', 'INVDOS', 'INVDT', 'INVDX', 'INVGR', 'INVGRP', 'INVID', 'INVMOD', 'INVNAM', 'INVND', 'INVPCN', 'INVPOS', 'IPRO', 'IREC', 'IREL', 'ITAX', 'ITMZ', 'KEYED', 'LIABL', 'LINE', 'LOWDOL', 'MAXBN', 'MCAID', 'MEDNCOV', 'NENTBP', 'NENTIN', 'NENTPT', 'NENTRF', 'NOAUTH', 'NOCOD', 'NOELG', 'NOTES', 'NOXWK', 'NUBCCD', 'NUBCVL', 'OUTNW', 'PCFDNY', 'PCFVOID', 'PCPRE', 'PDIA', 'PEND', 'PRPRC', 'PRPRD', 'PTINFO', 'QUAL', 'RECAL', 'RENPRO', 'RESENT', 'RESUB', 'REVCOD', 'RNID', 'RNNAP', 'ROMRAT', 'SPPBM', 'SUBID', 'TERMGRP', 'TIMELY', 'TPA', 'TPLNC', 'UMID', 'UNITS', 'UNPRO', 'WASTE' )
        THEN 'ACTIONABLE'
        WHEN A.ACTION_CD IN('1', '2', '3', '4', '5', 'ADJ', 'ADJUD', 'AMBORG', 'AMBUL', 'APDNA', 'APDTF', 'APMRB', 'APPDN', 'ARDUP', 'ASSIGN', 'AUTHDN', 'AUTO', 'BADCLMS', 'BANK', 'BENEFITCHG', 'BILLER', 'BILPR', 'BKRPT', 'CANCEL', 'CAPAJ', 'CAPIT', 'CARD', 'CARE', 'CCSOL', 'CCU', 'CDHNE', 'CFLDP', 'CLMAJ', 'CLMFRQ', 'CLMVD', 'CMPMI', 'CMPNC', 'CMPPR', 'CMPPT', 'CNBC', 'CNPHA', 'COINS', 'CONTR', 'COPAY', 'COSMT', 'CREDIT', 'CSFAL', 'CSNID', 'CTOHA', 'CUSTD', 'DAVITA', 'DAWPT', 'DAYLI', 'DAYSU', 'DAYSUP', 'DB2', 'DEDUC', 'DEFER', 'DEFUNCT', 'DENTL', 'DEPNE', 'DEPPG', 'DISAL', 'DISPOL', 'DNYPD', 'DODIN', 'DOSNE', 'DPWMI', 'DRGEX', 'DRGNA', 'DRGNC', 'DRGPG', 'DRGTB', 'DSPMI', 'DSPMX', 'DUPPRO', 'DURCF', 'DURNC', 'DXNC', 'ELGIS', 'EMPN', 'EMPNA', 'EMPNE', 'ENDUP', 'EXDRG', 'EXFREQ', 'EXRCR', 'FACNC', 'FAIL', 'FLEX', 'FRMLRY', 'FRONT', 'FWPXX', 'GENSB', 'GRAND', 'GROSS', 'GRPNE', 'HDNA', 'HFAMILY', 'HMSADJ', 'HMSC', 'HMSD', 'HMSE', 'HMSG', 'HMSH', 'HMSM', 'HMSS', 'HMST', 'HMSU', 'HMSX', 'HMSZ', 'HOSPIC', 'HOSPOL', 'ICNPD', 'ID12M', 'ID24M', 'IDEXP', 'IDOL', 'IFILL', 'INCL', 'INDC', 'INDEM', 'INDN', 'INGEX', 'INGMI', 'INGMX', 'INPNC', 'INVDO', 'INVGN', 'INVNDC', 'INVPL', 'INVPY', 'INVQT', 'INVUC', 'IPA', 'IRGC', 'IRX', 'LIFEPO', 'LMAX', 'LOSC', 'LOWDOL', 'LRDEX', 'LTCNC', 'LTD', 'MAB', 'MAIL', 'MAMCO', 'MAPAD', 'MAREF', 'MATCH', 'MAXBE', 'MAXBEN', 'MAXBN', 'MCCOB', 'MEDAJ', 'MEDNCO', 'MEDNEC', 'MFDNY', 'MGDCR', 'MHNC', 'MMNS', 'MXBEN', 'NABP', 'NBPAT', 'NBPT', 'NDCGP', 'NDCNC', 'NEIC', 'NEICDX', 'NEICPD', 'NENTIN', 'NENTPT', 'NENTSB', 'NLP', 'NOBIL', 'NOBILL', 'NOBRK', 'NOCOL', 'NOCOV', 'NOCVGE', 'NODENT', 'NOELG', 'NOMAT', 'NOMC', 'NOMCA', 'NOMCB', 'NOMED', 'NOMHC', 'NONEW', 'NONURS', 'NOOBES', 'NOPAR', 'NOPAY', 'NOPAYADJ', 'NOPRO', 'NOSUB', 'NOVIS', 'NPHARM', 'NPICN', 'OFFADJ', 'OOP', 'OOPET', 'OTCNC', 'OTHPRO', 'OVDRG', 'OVERP', 'P', 'PARPY', 'PARTB', 'PARTD', 'PAYEE', 'PAYME', 'PAYPR', 'PCFDNY', 'PCFVOID', 'PCUHC', 'PDCR', 'PDPVPB', 'PDSTCLM', 'PDTOP', 'PDTOPT', 'PDTOPV', 'PDTOST', 'PDTOWC', 'PENL', 'PHARN', 'PHARNE', 'PHYNC', 'PIPDY', 'POSTIN', 'POUHC', 'PRCNC', 'PREEX', 'PREVOD', 'PREVOI', 'PREVPD', 'PRICE', 'PRIME', 'PROBL', 'PROD', 'PROVN', 'PROVNC', 'PRSC', 'PRSCAD', 'PRSNC', 'PRVADD', 'PRVNC', 'PTAGE', 'PTINF', 'PTRES', 'PTSTAT', 'PUHC', 'QUANT', 'RDCBN', 'REBDS', 'REBIL', 'RECALL', 'REFIL', 'REFIL0', 'REFIL1', 'REFILL', 'REFMI', 'REFNA', 'RESID', 'RESTR', 'REWRK', 'RIDER', 'ROUTN', 'RSC', 'RSPAPR', 'RXAC', 'RXAGE', 'RXBC', 'RXCUT', 'RXDSG', 'RXEDT', 'RXGEN', 'RXLMT', 'RXNC', 'RXORG', 'RXRCT', 'SCNC', 'SCNOC', 'SEAL', 'SETTL', 'SETTLD', 'SLFINS', 'SNFNC', 'SPEECH', 'SUBST', 'SVLR', 'TERMGRP', 'TOSN', 'TOSNE', 'TRVLNC', 'UNID', 'UNKPT', 'VISION', 'WASTE', 'WCOMP', 'WCOMP', 'WWEST', 'YMAXBN' )
        THEN 'FINAL'
        ELSE NULL
        END AS "DENIAL TYPE",
        
        B.CARRIER_CD AS "HMS CARRIER CODE",
        B.CARRIER_NM AS HMS_CARRIER_NAME,
        A.CAR_CLM_REF_NUM AS "CARRIER DENIAL CLAIM NUMBER",
        A.CARRIER_ACK_DTM AS "CARRIER ACTION DATE",
        TO_CHAR(A.BILL_DT,'MM/DD/YY') AS "ORIGINAL BILL DATE",
        TO_CHAR(A.LAST_BILL_DT, 'MM/DD/YY') AS "REBILL DATE",

    CASE 
    WHEN A.TRANSACT_STATUS_CD = 'D' THEN 'DENIED'
    WHEN A.TRANSACT_STATUS_CD = 'V' THEN 'VOID'
    WHEN A.TRANSACT_STATUS_CD = '8' THEN 'REVERSED'
        ELSE 'OTHER' 
        END AS STATUS,

    CASE WHEN A.ORIG_SRCE_ELIG_CD = 'RS' THEN 'RSC'
    ELSE 'TPL'
    END AS "SOURCE CODE"

    FROM	EDW_AR_FL.ARTBASE A

    LEFT JOIN EDW_AR_FL.ARTCARM B
    ON A.CARRIER_CD = B.CARRIER_CD

    RIGHT OUTER JOIN EDW_AR_FL.DENIALS C
    ON A.CONTRACT_NUM = C.CONTRACT_NUM AND A.AR_SEQ_NUM = C.AR_SEQ_NUM

    WHERE	A.CONTRACT_NUM='478' AND
    /* FOR FFS CLAIMS */  (SUBSTR(A.ICN_NUM,1,2) NOT IN('60', '70'))
        AND CAST(A.REMIT_DT AS DATE) BETWEEN '{first_day_prev_month}' AND '{last_day_prev_month}'  AND B.DEFAULT_IND = 'Y'
        AND A.ACTION_CD NOT IN('AMBORG',  'INPRC', 'INSADD', 'PCFDNY', 'BADEL', 'BDSRCE', 'MATCH', 'NOCOD','A2~20','A1~20','A2~19~P','A1~19~P','A2~20~S','A1~16','A1~19', 'GENREJ', 'ACK', '277ACC', 'LOWDOL', 'REWRK', 'BADEL', '24OLD', 'AGED', 'AGED1', 'ARDUP', 'ARSRC', 'CTOHA', 'DB2, DISAL', 'ELGIS', 'EMPNA', 'FWPXX', 'ICNPD', 'MATCH', 'MFDNY', 'NOBIL', 'NOCOL', 'NOELG', 'NPICN', 'PCFVOID', 'POSTIN', 'PREVOD', 'REBDS', 'REBIL', 'RECALL', 'RESENT', 'RSC', 'RSCINFO', 'SETTL', 'SETTLD', 'ZPDPR', 'ZPDRS') 
        AND A.TRANSACT_STATUS_CD IN ('D','8','V')
        GROUP BY 1,2,3,4,5,6,7,8,9,10,11,12,14,15,16,17,18,19,20,21,22,23,24,25, 26, 27, 28
    """

    cursor.execute(teradata_query1)
    columns3 = [column[0] for column in cursor.description]
    results3 = cursor.fetchall()
    results3 = [list(row) for row in results3] 
    print("Results from FFS Denial query:")
    for row in results3:
        print(row)

    # Export FFS Denied results to Excel and TXT
    df3 = pd.DataFrame(results3, columns=columns3)

    # Save to Excel with borders and without bold column names
    with pd.ExcelWriter(ffs_denied_excel_filename, engine='openpyxl') as writer:
        df3.to_excel(writer, index=False)
        workbook  = writer.book
        worksheet = writer.sheets['Sheet1']
        add_borders(worksheet)
        
        # Align column names to the left and remove bold formatting
        for cell in worksheet["1:1"]:
            cell.alignment = Alignment(horizontal="left")
            cell.font = Font(bold=False, name="Aptos Narrow")
        
        # Format columns as dates or currency and apply font to all cells
        format_columns_and_apply_font(worksheet)

    # Save to TXT with commas and quotes around each row
    with open(ffs_denied_txt_filename, 'w') as f:
        f.write('"' + '","'.join(columns3) + '"\n')
        for row in results3:
            f.write('"' + '","'.join(map(str, row)) + '"\n')

    print("FFS Denied Files Successfully Exported")

    #CCO DENIED CLAIMS
    teradata_query2 = f"""
    SELECT DISTINCT
            A.AR_SEQ_NUM AS "ARSEQ",
            A.LAST_NM AS "CLIENT LAST NAME",
            A.FIRST_NM AS "CLIENT FIRST NAME",
            A.MA_NUM AS "RECIPIENT ID", 
            SUBSTR(A.ICN_NUM,1,LENGTH(ICN_NUM)-2) AS ICN,
            RIGHT(A.ICN_NUM,2) AS "ICN DETAIL LINE",  
            TO_CHAR(A.FROM_DOS_DT, 'MM/DD/YY') AS "FROM DOS",
            TO_CHAR(A.THRU_DOS_DT, 'MM/DD/YY') AS "THRU DOS",
            A.BILL_AMT AS "BILLED AMOUNT",
            A.MA_PAID_AMT AS "LINE MEDICAID PAID",
            A.MED_HDRPD_AMT AS "MEDICAID PAID",
            TO_CHAR(A.MA_PAID_DT, 'MM/DD/YY') AS "MEDICAID PAID DATE",
            SUM(A.REMIT_AMT) AS "ACTUAL AMOUNT RECOVERED",
            C.ACTCD_RF1 AS "HMS ACTION CD",
            CASE WHEN A.ACTION_CD = 'ADDR' THEN 'Submit Claim to a Different Address'
            WHEN A.ACTION_CD = 'ALCARE' THEN 'Alternate Services were Available and Should Have Been Utilized'
            WHEN A.ACTION_CD = 'BCLCL   ' THEN 'File Claim with Local BCBS Plan'
            WHEN A.ACTION_CD = 'BPNPI' THEN 'Billing provider NPI'
            WHEN A.ACTION_CD = 'BPTAX' THEN 'Invalid billing provider Tax ID'
            WHEN A.ACTION_CD = 'BPTXN' THEN 'Missing/invalid billing provider taxonomy ID'
            WHEN A.ACTION_CD = 'CCLM' THEN 'Corrected Claim Needed'
            WHEN A.ACTION_CD = 'CLAMAJ' THEN 'Care may not be Covered by Another Payer per Coordination of Benefits'
            WHEN A.ACTION_CD = 'CLMFWD' THEN 'Claim Forwarded to Payer by NEIC'
            WHEN A.ACTION_CD = 'COBPD' THEN 'Paid in Accordance with COB'
            WHEN A.ACTION_CD = 'CPT' THEN 'CPT Code Missing'
            WHEN A.ACTION_CD = 'DUP' THEN 'Duplicate Claim Submission'
            WHEN A.ACTION_CD = 'DUPHMS' THEN 'Duplicate to Medicaid Claim'
            WHEN A.ACTION_CD = 'ELEC' THEN 'Submit Claims Electronically'
            WHEN A.ACTION_CD = 'ERRDES' THEN 'No description/definition of the denial code found on the EOB in DocDNA'
            WHEN A.ACTION_CD = 'ERRDNA' THEN 'Either the claim or the denial code could not be found in DocDNA'
            WHEN A.ACTION_CD = 'ERREOB' THEN 'There is no denial code present on the EOB in DocDNA'
            WHEN A.ACTION_CD = 'EXPER' THEN 'Procedure Considered Experimental'
            WHEN A.ACTION_CD = 'FAULT' THEN 'Applied to No Fault Benefit'
            WHEN A.ACTION_CD = 'HCFA' THEN 'Claim Needs to be Submitted on CMS/HCFA 1500 Form'
            WHEN A.ACTION_CD = 'HMSI' THEN 'Additional Info Needed (crosswalk to INFO)'
            WHEN A.ACTION_CD = 'IACC' THEN 'Accident Report Required'
            WHEN A.ACTION_CD = 'IANPI' THEN 'Missing/Invalid Attending NPI Number'
            WHEN A.ACTION_CD = 'IBNPI' THEN 'Missing/Invalid Billing NPI Number'
            WHEN A.ACTION_CD = 'ICD10' THEN 'Invalid ICD-10 code.'
            WHEN A.ACTION_CD = 'ICLF' THEN 'Claim Form Requested'
            WHEN A.ACTION_CD = 'ICOB' THEN 'COB Information Needed'
            WHEN A.ACTION_CD = 'ICPD' THEN 'CPT4 Code Description Needed'
            WHEN A.ACTION_CD = 'ICPT' THEN 'CPT4 Code Required'
            WHEN A.ACTION_CD = 'IDAYS' THEN 'Missing/Invalid Days Supply'
            WHEN A.ACTION_CD = 'IDEA' THEN 'Missing/Invalid DEA Number'
            WHEN A.ACTION_CD = 'IDIG' THEN 'DX Code Required'
            WHEN A.ACTION_CD = 'IDOB' THEN 'Invalid/Missing DOB'
            WHEN A.ACTION_CD = 'IDOS' THEN 'Date(s) of Service Needed'
            WHEN A.ACTION_CD = 'IEOB' THEN 'Primary EOB Needed'
            WHEN A.ACTION_CD = 'ILAC' THEN 'Nature of Illness or Accident Required'
            WHEN A.ACTION_CD = 'ILLEG' THEN 'Illegible Code on EOB'
            WHEN A.ACTION_CD = 'IMCV' THEN 'Medicare EOB Required'
            WHEN A.ACTION_CD = 'INAME' THEN 'Invalid Character in Name or Illegible Name'
            WHEN A.ACTION_CD = 'INFO' THEN 'Additional Information Needed'
            WHEN A.ACTION_CD = 'INOPID' THEN 'Missing/Invalid Operating Physician ID '
            WHEN A.ACTION_CD = 'INPI' THEN 'Missing/Invalid NPI Number'
            WHEN A.ACTION_CD = 'INPRC' THEN 'Invalid Procedure Code'
            WHEN A.ACTION_CD = 'INVADD' THEN 'Invalid/Incomplete Address'
            WHEN A.ACTION_CD = 'INVAM' THEN 'Invalid Amount'
            WHEN A.ACTION_CD = 'INVBN' THEN 'Missing/Invalid BIN Number'
            WHEN A.ACTION_CD = 'INVBT' THEN 'Invalid Bill Type'
            WHEN A.ACTION_CD = 'INVCD' THEN 'Invalid Code'
            WHEN A.ACTION_CD = 'INVDOS' THEN 'Invalid Date of Service'
            WHEN A.ACTION_CD = 'INVDT' THEN 'Invalid Admit Date'
            WHEN A.ACTION_CD = 'INVDX' THEN 'DX Not Valid for Procedure'
            WHEN A.ACTION_CD = 'INVES' THEN 'Claim Being Investigated - Includes Possible Mismatch Patients'
            WHEN A.ACTION_CD = 'INVGR' THEN 'Invalid Group Number (use INVGRP)'
            WHEN A.ACTION_CD = 'INVGRP  ' THEN 'Invalid Group Number'
            WHEN A.ACTION_CD = 'INVID   ' THEN 'Invalid ID Number'
            WHEN A.ACTION_CD = 'INVMOD' THEN 'Missing/Invalid Modifier'
            WHEN A.ACTION_CD = 'INVNAM' THEN 'Invalid Name'
            WHEN A.ACTION_CD = 'INVND' THEN 'Non-Matched Service Provider Id'
            WHEN A.ACTION_CD = 'INVPCN' THEN 'Missing/Invalid PCN Number'
            WHEN A.ACTION_CD = 'INVPOS' THEN 'Incorrect Place of Service'
            WHEN A.ACTION_CD = 'IPOA' THEN 'Missing/Invalid POA Indicator'
            WHEN A.ACTION_CD = 'IPRO' THEN 'Provider Information Required'
            WHEN A.ACTION_CD = 'IREC    ' THEN 'Medical Records Needed'
            WHEN A.ACTION_CD = 'IREL' THEN 'Patient Relation to Insured Required'
            WHEN A.ACTION_CD = 'IRNPI' THEN 'Missing/Invalid Referring NPI'
            WHEN A.ACTION_CD = 'ITAX' THEN 'Invalid billing provider Tax ID'
            WHEN A.ACTION_CD = 'ITMZ' THEN 'Itemized Bill Required'
            WHEN A.ACTION_CD = 'KEYED' THEN 'Denial Code Keyed Incorrectly'
            WHEN A.ACTION_CD = 'LEGAL   ' THEN 'HMS Legal Working on Claims Population'
            WHEN A.ACTION_CD = 'LIABL' THEN 'Injury/Illness is Covered by the Liability Carrier'
            WHEN A.ACTION_CD = 'MCV' THEN 'Medicare Voucher Requested (Use IMCV)'
            WHEN A.ACTION_CD = 'MDCLM' THEN 'Carrier Acknowledges as a Medicaid Reclamation Claim'
            WHEN A.ACTION_CD = 'NCRTP' THEN 'No Code Carrier Refused to Process Claims'
            WHEN A.ACTION_CD = 'NEWMOM' THEN 'Newborn Claims Must be Submitted Under the Mothers Policy'
            WHEN A.ACTION_CD = 'NOAUTH' THEN 'No Authorization Obtained'
            WHEN A.ACTION_CD = 'NOCOD   ' THEN 'No Denial Code on Remittance'
            WHEN A.ACTION_CD = 'NOINT' THEN 'Interim Bills Not Processed, Submit for Entire Admission'
            WHEN A.ACTION_CD = 'NOTES' THEN 'Physicians Orders or Nursing Notes Needed'
            WHEN A.ACTION_CD = 'NOXWK' THEN 'No-Crosswalk'
            WHEN A.ACTION_CD = 'NSXWK' THEN 'Non Standard Crosswalk'
            WHEN A.ACTION_CD = 'NWAIT   ' THEN 'Waiting Period for Benefit Not Satisfied'
            WHEN A.ACTION_CD = 'OCCUR' THEN 'NUBC Occurrence Code(s)'
            WHEN A.ACTION_CD = 'OPN' THEN 'Open Claim Not Yet Adjudicated by Carrier'
            WHEN A.ACTION_CD = 'OUTNW' THEN 'Out of Network'
            WHEN A.ACTION_CD = 'PAPER   ' THEN 'Resubmit on Paper'
            WHEN A.ACTION_CD = 'PBM' THEN 'Send Claim to PBM'
            WHEN A.ACTION_CD = 'PCPRE' THEN 'Not Authorized by PCP (Use NOAUTH)'
            WHEN A.ACTION_CD = 'PDIA' THEN 'Invalid Principal/Admit Diagnosis'
            WHEN A.ACTION_CD = 'PEND    ' THEN 'Claim Pended'
            WHEN A.ACTION_CD = 'PREAUT' THEN 'Pre Auth Needed (use NOAUTH)'
            WHEN A.ACTION_CD = 'PTINFO' THEN 'Need Information From Patient'
            WHEN A.ACTION_CD = 'QUAL' THEN 'Qualifying Procedure Not Received'
            WHEN A.ACTION_CD = 'RENPRO' THEN 'Pending Information from Rendering Provider'
            WHEN A.ACTION_CD = 'RESENT  ' THEN 'Resent Claims'
            WHEN A.ACTION_CD = 'RESUB   ' THEN 'Submit Claim to Another Entity'
            WHEN A.ACTION_CD = 'RNID' THEN 'Missing/invalid rendering provider ID number'
            WHEN A.ACTION_CD = 'SPLIT' THEN 'Split into Multiple Claims at Carrier'
            WHEN A.ACTION_CD = 'SPPBM' THEN 'Must Fill through Specialty Pharmacy'
            WHEN A.ACTION_CD = 'STDNT' THEN 'Coverage for Student Terminated due to Reaching Maximum Age'
            WHEN A.ACTION_CD = 'SUBMH' THEN 'Submit to Mental Health Carrier'
            WHEN A.ACTION_CD = 'TIMEL' THEN 'Claim Past Timely Filing Limit'
            WHEN A.ACTION_CD = 'TIMELY' THEN 'Time Limit for filing has expired'
            WHEN A.ACTION_CD = 'TIMRX' THEN 'Prescription Too Old'
            WHEN A.ACTION_CD = 'TOOTH' THEN 'Missing/Invalid Tooth Number'
            WHEN A.ACTION_CD = 'TPA' THEN 'Send Claim to TPA'
            WHEN A.ACTION_CD = 'TPLNC' THEN 'Group does not Allow Third Party Claims'
            WHEN A.ACTION_CD = 'UB92' THEN 'Resubmit Claim on a UB Form'
            WHEN A.ACTION_CD = 'UMGRP' THEN 'Group Number does not Match Carriers System (use INVGRP)'
            WHEN A.ACTION_CD = 'UMID    ' THEN 'ID Number does not Match Carriers System (use INVID)'
            WHEN A.ACTION_CD = 'UNITS' THEN 'Units Field Invalid for Number of Days'
            WHEN A.ACTION_CD = 'UNPRO' THEN 'Unprocessed Claim'
            WHEN A.ACTION_CD = 'ADJUD' THEN 'Adjudicated per Plan Contract/Allowable'
            WHEN A.ACTION_CD = 'AMBORG' THEN 'Facility point of origin and destination - ambulance.'
            WHEN A.ACTION_CD = 'APDTF   ' THEN 'Timely Filing Appeal Denied'
            WHEN A.ACTION_CD = 'ARDUP   ' THEN 'Claim is a Duplicate of Previously Billed Claim'
            WHEN A.ACTION_CD = 'AUTHDN' THEN 'Pre-Authorization Denied Prior to Service'
            WHEN A.ACTION_CD = 'BANK' THEN 'Carrier Has Filed for Bankruptcy'
            WHEN A.ACTION_CD = 'BILLER  ' THEN 'Billing Error'
            WHEN A.ACTION_CD = 'CANCEL' THEN 'Policy Canceled'
            WHEN A.ACTION_CD = 'CAPIT' THEN 'Capitated Service'
            WHEN A.ACTION_CD = 'CLMAJ   ' THEN 'Claim Adjusted'
            WHEN A.ACTION_CD = 'CLMFRQ' THEN 'Frequency of service.'
            WHEN A.ACTION_CD = 'CMPMI' THEN 'Missing/Invalid Compound Code'
            WHEN A.ACTION_CD = 'CMPNC' THEN 'Compounds Not Covered'
            WHEN A.ACTION_CD = 'COINS' THEN 'MA Paid Less Than Co-Insurance'
            WHEN A.ACTION_CD = 'CONTR   ' THEN 'Contraception Not Covered'
            WHEN A.ACTION_CD = 'COPAY' THEN 'MA Paid Less Than Copay'
            WHEN A.ACTION_CD = 'DAYSUP  ' THEN 'Days Supply Exceeds Plan Limits'
            WHEN A.ACTION_CD = 'DEDUC' THEN 'Payment Applied to Patient Deductible'
            WHEN A.ACTION_CD = 'DEFER' THEN 'Portion of Payment Deferred'
            WHEN A.ACTION_CD = 'DEFUNCT' THEN 'Carrier No longer Exist'
            WHEN A.ACTION_CD = 'DEPNE   ' THEN 'Dependent Not Eligible'
            WHEN A.ACTION_CD = 'DEPPG' THEN 'Dependent Pregnancy Not Covered'
            WHEN A.ACTION_CD = 'DNYPD' THEN 'Uncollectable Claim that was Paid to Provider, Patient or State'
            WHEN A.ACTION_CD = 'DOSNE' THEN 'Coverage not in effect at time of service '
            WHEN A.ACTION_CD = 'DPWMI' THEN 'Date Prescription Written Missing/Invalid'
            WHEN A.ACTION_CD = 'DRGNC' THEN 'Drug Not Covered'
            WHEN A.ACTION_CD = 'DSPMI' THEN 'Missing/Invalid Dispense as Written Code'
            WHEN A.ACTION_CD = 'DUPPRO' THEN 'Duplicate to Provider Claim'
            WHEN A.ACTION_CD = 'DURCF' THEN 'Insert Fail DUR-Conflict'
            WHEN A.ACTION_CD = 'DURNC' THEN 'Durable Medical Equipment Not Covered'
            WHEN A.ACTION_CD = 'DXNC' THEN 'Diagnosis Not Covered'
            WHEN A.ACTION_CD = 'EMPNE' THEN 'Employee not Eligible'
            WHEN A.ACTION_CD = 'EXFREQ' THEN 'Service Exceeds Approved Frequency'
            WHEN A.ACTION_CD = 'EXRCR' THEN 'Exceeds Reasonable and Customary Rate'
            WHEN A.ACTION_CD = 'FLEX    ' THEN 'Flexible Spending Account Payments'
            WHEN A.ACTION_CD = 'FRMLRY' THEN 'Product Not On Formulary'
            WHEN A.ACTION_CD = 'FRONT   ' THEN 'Front End Edit Reject'
            WHEN A.ACTION_CD = 'GENSB' THEN 'Generic Substitution Required'
            WHEN A.ACTION_CD = 'GRPNE' THEN 'Group Not Eligible'
            WHEN A.ACTION_CD = 'HOSPIC' THEN 'Patient not covered for Hospice care'
            WHEN A.ACTION_CD = 'ICNPD   ' THEN 'DO NOT USE - Duplicate Claim Closed Out - ICN billed to and paid by different carrier (Auto Deny Job)'
            WHEN A.ACTION_CD = 'IDOL' THEN 'Mapaid > Billed'
            WHEN A.ACTION_CD = 'INCL' THEN 'Carrier will not Pay Separately for this Service'
            WHEN A.ACTION_CD = 'INDC' THEN 'Missing or Invalid NDC Number'
            WHEN A.ACTION_CD = 'INGMX' THEN 'Ingredient Cost Reduced to Maximum'
            WHEN A.ACTION_CD = 'INVGN' THEN 'Invalid Gender'
            WHEN A.ACTION_CD = 'INVPL' THEN 'Invalid or Incomplete Protocol Requirements'
            WHEN A.ACTION_CD = 'INVQT' THEN 'Invalid Quantity Entered for Medication Package '
            WHEN A.ACTION_CD = 'IRX     ' THEN 'Missing/Invalid RX Number'
            WHEN A.ACTION_CD = 'LMAX' THEN 'Lifetime Benefit Maximum Met'
            WHEN A.ACTION_CD = 'LOSC' THEN 'Invalid Location of Service'
            WHEN A.ACTION_CD = 'MAXBEN' THEN 'Benefit Maximum has been reached'
            WHEN A.ACTION_CD = 'MAXBN' THEN 'Maximum Benefits Reached (Use MAXBEN)'
            WHEN A.ACTION_CD = 'MEDAJ' THEN 'Medicaid Claim Adjudicated'
            WHEN A.ACTION_CD = 'MEDNCO' THEN 'Not a Medicare Covered Service'
            WHEN A.ACTION_CD = 'MEDNEC' THEN 'Claim Not Medically Necessary'
            WHEN A.ACTION_CD = 'MHNC    ' THEN 'Mental Health Not Covered'
            WHEN A.ACTION_CD = 'NABP    ' THEN 'Carrier Needs NABP Number'
            WHEN A.ACTION_CD = 'NBPT' THEN 'Non-Billable Provider Type'
            WHEN A.ACTION_CD = 'NDCNC' THEN 'NDC code Not Covered'
            WHEN A.ACTION_CD = 'NEICPD' THEN 'Payment Per Negotiated Rate'
            WHEN A.ACTION_CD = 'NENTPT' THEN 'Entity Not Found: Patient'
            WHEN A.ACTION_CD = 'NOBIL   ' THEN 'Claim Should Not Have Been Billed per Client - Do Not Work/Rebill Claims'
            WHEN A.ACTION_CD = 'NOCOV' THEN 'No Coverage'
            WHEN A.ACTION_CD = 'NODENT' THEN 'No Dental Coverage'
            WHEN A.ACTION_CD = 'NOMAT' THEN 'Maternity Charges Not Covered'
            WHEN A.ACTION_CD = 'NOMED   ' THEN 'Medical Services Not Covered'
            WHEN A.ACTION_CD = 'NONEW' THEN 'Insured has No Coverage for Newborns'
            WHEN A.ACTION_CD = 'NONURS  ' THEN 'Home Nursing Services Not Covered'
            WHEN A.ACTION_CD = 'NOOBES' THEN 'Obesity Services Not Covered'
            WHEN A.ACTION_CD = 'NOPAY' THEN 'No Payment to be Issued'
            WHEN A.ACTION_CD = 'NOVIS' THEN 'Vision Services Not Covered'
            WHEN A.ACTION_CD = 'NPHARM' THEN 'Non-Matched Pharmacy Number'
            WHEN A.ACTION_CD = 'OTCNC   ' THEN 'Over the Counter Drugs Not Covered'
            WHEN A.ACTION_CD = 'PARTB' THEN 'Claim not Processed - Medicare Part B policy'
            WHEN A.ACTION_CD = 'PARTD' THEN 'Claim not Processed - Medicare Part D policy'
            WHEN A.ACTION_CD = 'PDCR    ' THEN 'Paid at Customary and Reasonable Rate'
            WHEN A.ACTION_CD = 'PDPVPB' THEN 'Paid to Provider Prior to HMS Billing'
            WHEN A.ACTION_CD = 'PDTOPT  ' THEN 'Claim Paid to Patient'
            WHEN A.ACTION_CD = 'PDTOPV' THEN 'Claim Paid to Provider of Service'
            WHEN A.ACTION_CD = 'PDTOST  ' THEN 'Paid to State Medicaid Agency'
            WHEN A.ACTION_CD = 'PDTOWC' THEN 'Paid to Incorrect HMS Client'
            WHEN A.ACTION_CD = 'PHARNE  ' THEN 'Pharmacy Not Eligible'
            WHEN A.ACTION_CD = 'PHYNC' THEN 'Physician Not Covered '
            WHEN A.ACTION_CD = 'PRCNC' THEN 'Procedure Code Not Covered'
            WHEN A.ACTION_CD = 'PREEX' THEN 'Pre-Existing Condition'
            WHEN A.ACTION_CD = 'PROBL' THEN 'Servicing Provider Must Bill Carrier Directly'
            WHEN A.ACTION_CD = 'PROVN' THEN 'Provider Not Covered (use PRVNC)'
            WHEN A.ACTION_CD = 'PRSC' THEN 'Expected a Prescriber Field'
            WHEN A.ACTION_CD = 'PRSNC' THEN 'Drug Not Covered for Prescriber'
            WHEN A.ACTION_CD = 'PRVNC' THEN 'Provider Not Covered'
            WHEN A.ACTION_CD = 'PTAGE' THEN 'Patient Outside Age Limit for this Type of Benefit'
            WHEN A.ACTION_CD = 'PTRES' THEN 'Patient Responsibility'
            WHEN A.ACTION_CD = 'QUANT   ' THEN 'Quantity Not Covered'
            WHEN A.ACTION_CD = 'REBIL   ' THEN 'Claim Closed to Pass through New Cycle'
            WHEN A.ACTION_CD = 'REFIL' THEN 'Refill Too Soon'
            WHEN A.ACTION_CD = 'REFMI' THEN 'Refill Number Missing/Invalid'
            WHEN A.ACTION_CD = 'RIDER' THEN 'Considered Under Rider Coverage'
            WHEN A.ACTION_CD = 'ROUTN' THEN 'Routine Services Not Covered'
            WHEN A.ACTION_CD = 'RXAGE' THEN 'Participants Age Restricts Medication Coverage'
            WHEN A.ACTION_CD = 'RXDSG' THEN 'Fail Rx Dosage Rule Table'
            WHEN A.ACTION_CD = 'RXGEN' THEN 'Participants Gender Restricts Medication Coverage'
            WHEN A.ACTION_CD = 'RXLMT   ' THEN 'Medication Exceeds Plan Limits'
            WHEN A.ACTION_CD = 'RXNC' THEN 'Prescription Drugs Not Covered'
            WHEN A.ACTION_CD = 'TERMGRP' THEN 'The group termed with this carrier and the run out period has ended'
            WHEN A.ACTION_CD = 'TOSNE' THEN 'Type of service not covered'
            WHEN A.ACTION_CD = 'TRVLNC' THEN 'Travel/Transportation Not Covered'
            WHEN A.ACTION_CD = 'UNID' THEN 'Unable to Identify Member'
            WHEN A.ACTION_CD = 'WASTE' THEN 'Procedure code that are not covered by the carrier and should never go back to the same carrier'
            WHEN A.ACTION_CD = 'WCOMP' THEN 'Claim should be Processed by Workers Comp Carrier'
            WHEN A.ACTION_CD = 'WWEST'THEN 'Outsourced to Washington & West (Massachusetts Medicaid)'
            WHEN A.ACTION_CD = 'YMAXBN'THEN 'Annual Maximum Benefits Reached (use MAXBEN)'
            WHEN A.ACTION_CD = 'ZPDPR'THEN 'Zero Pay Carrier - Billed from a PAR/Imputed Carrier Feed'
            WHEN A.ACTION_CD = 'ZPDRS'THEN 'Zero Pay Carrier - Billed from the State Resource Feed'
            WHEN A.ACTION_CD = 'PO1'THEN 'Full payment received'
            WHEN A.ACTION_CD = 'PO2'THEN 'Partial Payment received'
            WHEN A.ACTION_CD = 'P03'THEN 'Voluntary/Excess payment received'
            WHEN A.ACTION_CD = 'PEO1'THEN 'Provider grant extension.'
            WHEN A.ACTION_CD = 'PEO2'THEN 'Credit for Claim due to HMS (do not recoup)'
            WHEN A.ACTION_CD = 'RO1'THEN 'Provider agrees with recoupment'
            WHEN A.ACTION_CD = 'RO2'THEN 'Recoupment scheduled for future dates'
            WHEN A.ACTION_CD = 'RO3'THEN 'Recoupment for future date (final balance)'
            WHEN A.ACTION_CD = 'RECUP10'THEN 'Recovery has been submitted for invoicing'
            WHEN A.ACTION_CD = 'V01'THEN 'Not Eligible for Medicare on DOS'
            WHEN A.ACTION_CD = 'V02'THEN 'Medicare Benefits Exhausted'
            WHEN A.ACTION_CD = 'V03'THEN 'Invalid MC HIC # '
            WHEN A.ACTION_CD = 'V05'THEN 'Non-covered Medicare Service(Provider Relations denial)'
            WHEN A.ACTION_CD = 'V06'THEN 'Prior MA Recoupment'
            WHEN A.ACTION_CD = 'V10'THEN 'Untimely'
            WHEN A.ACTION_CD = 'V11'THEN 'Administrative days or PRO denial (bed hold days)'
            WHEN A.ACTION_CD = 'V12'THEN 'Correctly Billed by Medicare and Medicaid'
            WHEN A.ACTION_CD = 'V13'THEN 'Not a Medicare Assigned Provider'
            WHEN A.ACTION_CD = 'V14'THEN 'Mail returned'
            WHEN A.ACTION_CD = 'V15'THEN 'HH:Patient not home bound'
            WHEN A.ACTION_CD = 'V16'THEN 'HH:PT/OT/ST not under treatment plan'
            WHEN A.ACTION_CD = 'V17'THEN 'HH:Nursing care not under treatment plan'
            WHEN A.ACTION_CD = 'V18'THEN 'HH: Nursing care not skilled.'
            WHEN A.ACTION_CD = 'V19'THEN 'HH: Nursing no intermittent'
            WHEN A.ACTION_CD = 'V20'THEN 'HH: Patient condition chronic'
            WHEN A.ACTION_CD = 'V21 'THEN 'Disallow amount is for DED/Coins'
            WHEN A.ACTION_CD = 'V29'THEN 'Patient not affiliated with provider. (Patient not found by provider.)'
            WHEN A.ACTION_CD = 'VO4'THEN 'Bankrupt provider'
            WHEN A.ACTION_CD = 'VO7'THEN 'Prior providers refund'
            WHEN A.ACTION_CD = 'VO8'THEN 'SNF/Non-skilled level of care'
            WHEN A.ACTION_CD = 'VO9'THEN 'Legal providers (MAT)'
            WHEN A.ACTION_CD = 'V30'THEN 'Case Disallowed '
            WHEN A.ACTION_CD = 'V31'THEN 'Direct Payment'
            WHEN A.ACTION_CD = 'V32'THEN 'Duplicate - MRM'
            WHEN A.ACTION_CD = 'V33'THEN 'Duplicate - vendor'
            WHEN A.ACTION_CD = 'V34'THEN 'Failure to pursue'
            WHEN A.ACTION_CD = 'V35'THEN 'Member Paid'
            WHEN A.ACTION_CD = 'V36'THEN 'Negotiated Rate'
            WHEN A.ACTION_CD = 'V37'THEN 'No Auth for Service'
            WHEN A.ACTION_CD = 'V38'THEN 'No Causal Relationship'
            WHEN A.ACTION_CD = 'V39'THEN 'Provider Takeback'
            WHEN A.ACTION_CD = 'V40'THEN 'Refund MRM Fee'
            WHEN A.ACTION_CD = 'V41'THEN 'Settlement Approved'
            WHEN A.ACTION_CD = 'V42'THEN 'Aetna Unrecoverable Amount'
            WHEN A.ACTION_CD = 'V43'THEN 'Provider Denied'
            WHEN A.ACTION_CD = 'OPN01 'THEN 'Open Claim This Claim Will Be Recouped'
            WHEN A.ACTION_CD = 'OPN02 'THEN 'No Response Extension'
            WHEN A.ACTION_CD = 'OPN03 'THEN 'Claim Recovered/Missing Documentation'
            WHEN A.ACTION_CD = 'OPN04 'THEN 'No Medical Records Received'
            WHEN A.ACTION_CD = 'OPN05 'THEN 'No Response During 2nd Extension'
            WHEN A.ACTION_CD = 'OPN06 'THEN 'Final Notice Of Recovery Sent'
            WHEN A.ACTION_CD = 'OPN07 'THEN 'No Response During 3rd Extension'
            WHEN A.ACTION_CD = 'P01 'THEN 'Full Payment Received'
            WHEN A.ACTION_CD = 'P02 'THEN 'Partial Payment Received'
            WHEN A.ACTION_CD = 'P03 'THEN 'Partial Payment P03'
            WHEN A.ACTION_CD = 'P04 'THEN 'MD Checks Received Via Provider'
            WHEN A.ACTION_CD = 'P05 'THEN 'Extension Payment Received'
            WHEN A.ACTION_CD = 'P06 'THEN 'Payment From EMOMED Forms'
            WHEN A.ACTION_CD = 'P50 'THEN 'Paid By Check - Private Health Insurance Paid'
            WHEN A.ACTION_CD = 'P51 'THEN 'Paid By Check - Medicaid Paid Twice (Duplicate)'
            WHEN A.ACTION_CD = 'P52 'THEN 'Paid By Check - Medicare Paid'
            WHEN A.ACTION_CD = 'P53 'THEN 'Paid By Check - Patient Paid.'
            WHEN A.ACTION_CD = 'P54 'THEN 'Paid By Check - Medicare/Medicaid Cross-Over Pymt'
            WHEN A.ACTION_CD = 'P55 'THEN 'Paid By Check - Other Noted By Provider'
            WHEN A.ACTION_CD = 'P56 'THEN 'Paid By Check - Casualty/Estate Recovery'
            WHEN A.ACTION_CD = 'P57 'THEN 'Paid By Check - Billing Error'
            WHEN A.ACTION_CD = 'P58 'THEN 'Paid By Check - Medicare SNF Stay'
            WHEN A.ACTION_CD = 'PAMDP 'THEN 'Pending Additional Medical Records To Permedion'
            WHEN A.ACTION_CD = 'PE01 'THEN 'Provider Granted Extension'
            WHEN A.ACTION_CD = 'PE02 'THEN 'Provider Granted 2nd 60day Extension (CATPL Only)'
            WHEN A.ACTION_CD = 'PE03 'THEN 'Provider Granted 3rd 60day Extension (CATPL Only)'
            WHEN A.ACTION_CD = 'PE04 'THEN 'Provider Is Sending Check Into Lockbox'
            WHEN A.ACTION_CD = 'PE05 'THEN 'Claims Are Pending/ In Process Of Being Worked'
            WHEN A.ACTION_CD = 'PE06 'THEN 'Medical Records Rec`D In Process Of Being Worked'
            WHEN A.ACTION_CD = 'PE10 'THEN 'CATPL ONLY Extension Per CA PD 120 Days'
            WHEN A.ACTION_CD = 'PE100 'THEN 'Provider Requesting Extension'
            WHEN A.ACTION_CD = 'PE30 'THEN '1st 30 Day Extension Request'
            WHEN A.ACTION_CD = 'PE500 'THEN '(PIRA)Providers Wants To Appeal, Pending Records'
            WHEN A.ACTION_CD = 'PE505 'THEN 'Rec`D Appeal Docs In Process Of Working Docs'
            WHEN A.ACTION_CD = 'PE510 'THEN '(PIRA) Pending Appeal To State'
            WHEN A.ACTION_CD = 'PE60 'THEN '2nd 30 Day Extension Request'
            WHEN A.ACTION_CD = 'PE600 'THEN 'CA-Provider Has Billed Or Attempted To Bill'
            WHEN A.ACTION_CD = 'PE90 'THEN '3rd 30 Day Extension Request'
            WHEN A.ACTION_CD = 'R01 'THEN 'Provider Agrees With Recoupment'
            WHEN A.ACTION_CD = 'R02 'THEN 'Recoupment Scheduled For Future Date'
            WHEN A.ACTION_CD = 'R03 'THEN 'Result Of Appeal Process Thru Permedion'
            WHEN A.ACTION_CD = 'R04 'THEN 'Provider Verbally Agreed To Recoupment'
            WHEN A.ACTION_CD = 'R05 'THEN 'Patient Liability Not Applicable Partial Recoup'
            WHEN A.ACTION_CD = 'R06 'THEN 'Recoup During Extension'
            WHEN A.ACTION_CD = 'R07 'THEN 'Recoupment Confirmed'
            WHEN A.ACTION_CD = 'R08 'THEN 'Adjustment Form Received-Not Processed By Client'
            WHEN A.ACTION_CD = 'R09 'THEN 'Provider Verbally Agree To Recoup During Extension'
            WHEN A.ACTION_CD = 'R10 'THEN 'Recoup/Untimely Denied/Appeal'
            WHEN A.ACTION_CD = 'R11 'THEN 'Provider Agrees To Recoupment - NC Recoup File'
            WHEN A.ACTION_CD = 'R12 'THEN 'Recoup During 2nd - 30 Day Extension'
            WHEN A.ACTION_CD = 'R13 'THEN 'Recoup During 3rd- 30 Day Extension'
            WHEN A.ACTION_CD = 'R14 'THEN 'Provider Verbally Agrees To Recoup - NC File'
            WHEN A.ACTION_CD = 'R15 'THEN 'Partial Recoupment'
            WHEN A.ACTION_CD = 'R16 'THEN 'Recoupment File Processed'
            WHEN A.ACTION_CD = 'R17 'THEN 'Follow-Up Completed.'
            WHEN A.ACTION_CD = 'R18 'THEN 'Possible Double Recoupment'
            WHEN A.ACTION_CD = 'R19 'THEN 'Unacceptable Documentation Received'
            WHEN A.ACTION_CD = 'R20 'THEN 'Claim Recouped Mid Cycle'
            WHEN A.ACTION_CD = 'R21 'THEN 'Incomp Documentation Rec`D 1st Ext (Days 90-120)'
            WHEN A.ACTION_CD = 'R22 'THEN 'Incomp Documentation Rec`D 2nd Ext (Days 120-150)'
            WHEN A.ACTION_CD = 'R23 'THEN 'CX Provider Agrees To Refund'
            WHEN A.ACTION_CD = 'R24 'THEN 'Responded But No Documentation Provided'
            WHEN A.ACTION_CD = 'R25 'THEN 'VA-Provider Initated Adj/Refund Due To HMS Audit'
            WHEN A.ACTION_CD = 'R26 'THEN 'Extension Request Denied By Account Team'
            WHEN A.ACTION_CD = 'R27 'THEN 'CX Cleveland Clinic No Response From Provider'
            WHEN A.ACTION_CD = 'R30 'THEN 'Recoup At 30 Days'
            WHEN A.ACTION_CD = 'R50 'THEN 'Recoupment - Private Health Insurance Paid'
            WHEN A.ACTION_CD = 'R51 'THEN 'Recoupment - Medicaid Paid Twice. (Duplicate)'
            WHEN A.ACTION_CD = 'R52 'THEN 'Recoupment - Medicare Paid'
            WHEN A.ACTION_CD = 'R53 'THEN 'Recoupment - Patient Paid'
            WHEN A.ACTION_CD = 'R54 'THEN 'Recoupment - Medicare/Medcaid Crossover Payment'
            WHEN A.ACTION_CD = 'R55 'THEN 'Recoupment - Other Noted By Provider'
            WHEN A.ACTION_CD = 'R56 'THEN 'Recoupment - Casualty/Estate Recovery'
            WHEN A.ACTION_CD = 'R57 'THEN 'Recoupment - Billing Error'
            WHEN A.ACTION_CD = 'R58 'THEN 'Recoupment - Medicare SNF Stay'
            WHEN A.ACTION_CD = 'R59 'THEN 'Claim Voided Via Mmis By Provider'
            WHEN A.ACTION_CD = 'R60 'THEN 'Recoup At 60 Days'
            WHEN A.ACTION_CD = 'R70 'THEN 'Reconsideration- Upheld'
            WHEN A.ACTION_CD = 'R72 'THEN 'Docs Does Not Support Scenario'
            WHEN A.ACTION_CD = 'R80 'THEN 'Unacceptable Docs Received During Extension'
            WHEN A.ACTION_CD = 'R81 'THEN 'Claim Denied For Untimely Filing During Extension'
            WHEN A.ACTION_CD = 'R82 'THEN 'No Documentation Received'
            WHEN A.ACTION_CD = 'R90 'THEN 'Recoup At 90 Days'
            WHEN A.ACTION_CD = 'R91 'THEN 'Preliminary Findings'
            WHEN A.ACTION_CD = 'R92 'THEN 'Upheld/ Result Of Appeal'
            WHEN A.ACTION_CD = 'R93 'THEN 'Upheld/ Appeal To State'
            WHEN A.ACTION_CD = 'R94 'THEN 'Technical Denial'
            WHEN A.ACTION_CD = 'R95 'THEN 'Final Recovery Letter Sent'
            WHEN A.ACTION_CD = 'RECON 'THEN 'Reconsideration Request Received'
            WHEN A.ACTION_CD = 'RITA 'THEN 'Untimely Rebuttal Received - Uphold'
            WHEN A.ACTION_CD = 'RLTNO 'THEN 'Tentative Notice Of Overpayment Sent'
            WHEN A.ACTION_CD = 'RPCOM 'THEN 'Complex Review Payment Received'
            WHEN A.ACTION_CD = 'RUP 'THEN 'Upheld Letter Generated'
            WHEN A.ACTION_CD = 'SDR 'THEN 'Supporting Documentation Rec`Vd-In Review Status'
            WHEN A.ACTION_CD = 'TNOOS 'THEN 'Tentative Notic Of Overpayment Sent'
            WHEN A.ACTION_CD = 'U01 'THEN 'Rac Identified Underpayments'
            WHEN A.ACTION_CD = 'U02 'THEN 'RAC Identified Dollar Value Of Underpayments'
            WHEN A.ACTION_CD = 'V01 'THEN 'No Coverage For DOS (Dates Of Service)'
            WHEN A.ACTION_CD = 'V02 'THEN 'Medicare Benefit Days Exhausted'
            WHEN A.ACTION_CD = 'V03 'THEN 'Medicaid Recipient Mismatched'
            WHEN A.ACTION_CD = 'V04 'THEN 'Bankrupt Provider'
            WHEN A.ACTION_CD = 'V05 'THEN 'Non-Covered Service'
            WHEN A.ACTION_CD = 'V06 'THEN 'Prior Medicaid Recoupment'
            WHEN A.ACTION_CD = 'V07 'THEN 'Prior Provider Refund'
            WHEN A.ACTION_CD = 'V08 'THEN 'Medicaid Recouped Before Cycle'
            WHEN A.ACTION_CD = 'V09 'THEN 'MATPL ONLY Legal Providers'
            WHEN A.ACTION_CD = 'V10 'THEN 'Time Limit For Filing Has Expired'
            WHEN A.ACTION_CD = 'V100 'THEN 'Provider Does Not Agree With Recoupment'
            WHEN A.ACTION_CD = 'V11 'THEN 'Death Certificate Rec'
            WHEN A.ACTION_CD = 'V12 'THEN 'Correctly Billed By Blue Cross, Medicare, Medicaid'
            WHEN A.ACTION_CD = 'V13 'THEN 'Non-Participating Provider'
            WHEN A.ACTION_CD = 'V14 'THEN 'Provider Electronic Adjustment'
            WHEN A.ACTION_CD = 'V15 'THEN 'Missing Modifier-No Impact On Reimbursement'
            WHEN A.ACTION_CD = 'V16 'THEN 'HH: PT/OT/ST Not Under Treatment Plan'
            WHEN A.ACTION_CD = 'V17 'THEN 'HH:Nursing Care Not Under Treatment Plan'
            WHEN A.ACTION_CD = 'V18 'THEN 'HH: Nursing Care Not Skilled'
            WHEN A.ACTION_CD = 'V19 'THEN 'HH: Nursing No Intermittent'
            WHEN A.ACTION_CD = 'V20 'THEN 'HH: Patient Condition Chronic'
            WHEN A.ACTION_CD = 'V21 'THEN 'Disallow Amount Is For Ded/Co (MA Liability)'
            WHEN A.ACTION_CD = 'V22 'THEN 'Dupe Surgeons (Billed Correctly)'
            WHEN A.ACTION_CD = 'V23 'THEN 'Non-Dupes- Twin Birth'
            WHEN A.ACTION_CD = 'V24 'THEN 'Billed Correctly(Same Proc Left/Right Forced Clm)'
            WHEN A.ACTION_CD = 'V25 'THEN 'Wrong Claim Selected/Supporting Claim'
            WHEN A.ACTION_CD = 'V26 'THEN 'Bill Correctly(Diffrent Level Of Care)'
            WHEN A.ACTION_CD = 'V27 'THEN 'Paid Correctly - Crossover Claim'
            WHEN A.ACTION_CD = 'V28 'THEN 'Provider Refuses To Bill MC (MATPL)'
            WHEN A.ACTION_CD = 'V29 'THEN 'Patient Not Affiliated With Provider'
            WHEN A.ACTION_CD = 'V30 'THEN 'Void Per Client/PD'
            WHEN A.ACTION_CD = 'V31 'THEN 'SNF - Nursing Care Not Skilled'
            WHEN A.ACTION_CD = 'V32 'THEN 'SNF Beneficiary Not In A Medicare-Certified Bed'
            WHEN A.ACTION_CD = 'V33 'THEN 'SNF Benefit Days Exhausted (Please Verify W/CWF)'
            WHEN A.ACTION_CD = 'V34 'THEN 'SNF Hosp Stay Unrelated To Subseq SNF Skilled Care'
            WHEN A.ACTION_CD = 'V35 'THEN 'Nv - Soi, Noridian Admin. Svc.'
            WHEN A.ACTION_CD = 'V36 'THEN 'NV - SOI, Mutual Of Omaha'
            WHEN A.ACTION_CD = 'V37 'THEN 'MO - Prev. Adj, Credit Or Void'
            WHEN A.ACTION_CD = 'V38 'THEN 'Pre-Cert/Pre-Auth Denied'
            WHEN A.ACTION_CD = 'V39 'THEN 'Void Per Client/PD MA Paid Amount To Be Recouped'
            WHEN A.ACTION_CD = 'V40 'THEN 'Entered As Self Audit When Actually MCA'
            WHEN A.ACTION_CD = 'V41 'THEN 'Change Of Ownership- Provider Not Held Responsible'
            WHEN A.ACTION_CD = 'V42 'THEN 'BC Benefit Days Exhausted'
            WHEN A.ACTION_CD = 'V43 'THEN 'Date Of Death Incorrect'
            WHEN A.ACTION_CD = 'V44 'THEN 'Provider Closed'
            WHEN A.ACTION_CD = 'V45 'THEN 'Under Carrier Review'
            WHEN A.ACTION_CD = 'V46 'THEN 'Recoupment Confirmed (Non Credit)'
            WHEN A.ACTION_CD = 'V47 'THEN 'Verbally Stated Time Limit For Filing Has Expired'
            WHEN A.ACTION_CD = 'V48 'THEN 'Return Mail'
            WHEN A.ACTION_CD = 'V49 'THEN 'Other Insurance Is Primary'
            WHEN A.ACTION_CD = 'V50 'THEN 'Provider Refunded Another Entity'
            WHEN A.ACTION_CD = 'V51 'THEN 'Claims For Timely Filing Appealed Correctly'
            WHEN A.ACTION_CD = 'V52 'THEN 'Provider Initiated Online Adj Due To HMS Audit'
            WHEN A.ACTION_CD = 'V53 'THEN 'No Findings'
            WHEN A.ACTION_CD = 'V54 'THEN 'Reconsideration-Overturned'
            WHEN A.ACTION_CD = 'V55 'THEN 'Overturn/Result Of Appeal'
            WHEN A.ACTION_CD = 'V56 'THEN 'Overturn/Appeal To State'
            WHEN A.ACTION_CD = 'V57 'THEN 'Suspected Fraud'
            WHEN A.ACTION_CD = 'V58 'THEN 'Incorrect Policy Number'
            WHEN A.ACTION_CD = 'V59 'THEN 'Policy Not In The States MMIS'
            WHEN A.ACTION_CD = 'V60 'THEN 'Hardship/Disaster Relief'
            WHEN A.ACTION_CD = 'V61 'THEN 'Appeal Overturned'
            WHEN A.ACTION_CD = 'V62 'THEN 'No Records Received, Closing Per Client'
            WHEN A.ACTION_CD = 'V63 'THEN 'Nj Mmis Void'
            WHEN A.ACTION_CD = 'V64 'THEN 'Void Claim Prior To Mailing'
            WHEN A.ACTION_CD = 'V65 'THEN 'Void Audit Overlap Claim'
            WHEN A.ACTION_CD = 'V66 'THEN 'Claim Closed, Recovery Initiated'
            WHEN A.ACTION_CD = 'V67 'THEN 'Corrected Claim Submitted By Provider'
            WHEN A.ACTION_CD = 'V68 'THEN 'Benefits Exhausted -Ci Cycle La'
            WHEN A.ACTION_CD = 'V72'THEN 'Re-Disallow Claims'
            WHEN A.ACTION_CD = 'V99 'THEN 'Void Duplicate Claim'
            WHEN A.ACTION_CD = 'v99 'THEN 'Duplicate Claims Disallowed'
            WHEN A.ACTION_CD = 'BDRFPV'THEN 'Bad referring provider info'
            WHEN A.ACTION_CD = 'BLANKPIN'THEN 'the rendering provider number is missing'
            WHEN A.ACTION_CD = 'DODB4C'THEN 'Date of death prior to DOS'
            WHEN A.ACTION_CD = 'DOSNE'THEN 'Coverage not in effect at time of service '
            WHEN A.ACTION_CD = 'DUPE'THEN 'Duplicate Claim (use DUP)'
            WHEN A.ACTION_CD = 'DXINC'THEN 'DX code incomplete'
            WHEN A.ACTION_CD = 'DXNOTCOV'THEN 'DX is not covered'
            WHEN A.ACTION_CD = 'FSSINFO'THEN 'missing facility name'
            WHEN A.ACTION_CD = 'HOSPIC'THEN 'Patient not covered for Hospice care'
            WHEN A.ACTION_CD = 'INAPPSER'THEN 'Treatment rendered inappropriate'
            WHEN A.ACTION_CD = 'INVLDDX/INVLDX'THEN 'Claim denied based on diagnosis'
            WHEN A.ACTION_CD = 'INVLDPIN'THEN 'invalid rendering provider number'
            WHEN A.ACTION_CD = 'INVRNFN'THEN 'Invalid rendering provider first name'
            WHEN A.ACTION_CD = 'INVRNLN'THEN 'Invalid rendering provider last name'
            WHEN A.ACTION_CD = 'MAXBEN'THEN 'Benefit Maximum has been reached'
            WHEN A.ACTION_CD = 'MAXCHG'THEN 'Maximum charges exceeded'
            WHEN A.ACTION_CD = 'MISSINFO/MSINFO'THEN 'Claim lacks info needed for adjudication'
            WHEN A.ACTION_CD = 'NOHICNUM'THEN 'Missing HIC number'
            WHEN A.ACTION_CD = 'NOTPDSEP'THEN 'Procedure is not paid separately'
            WHEN A.ACTION_CD = 'NOUPIN'THEN 'the UPIN is missing on the claim'
            WHEN A.ACTION_CD = 'PRNOCERT/PRNOCE'THEN 'Provider was not certified for this procedure on this date of service'
            WHEN A.ACTION_CD = 'PRVNOB'THEN 'Provider not allowed to bill for service'
            WHEN A.ACTION_CD = 'Routine'THEN 'Non covered charges, routine exam'
            WHEN A.ACTION_CD = 'SEPINP'THEN 'Not paid separately when patient in Hosp.'
            WHEN A.ACTION_CD = 'Timely'THEN 'Time Limit for filing has expired (USE Timel)'
            WHEN A.ACTION_CD = 'TOSNE'THEN 'Type of service not covered'
            WHEN A.ACTION_CD = 'V01'THEN 'Coverage termination prior to DOS'
            WHEN A.ACTION_CD = 'V02'THEN 'DOS prior to coverage effective date'
            WHEN A.ACTION_CD = 'V03'THEN 'Dependent not eligible for coverage'
            WHEN A.ACTION_CD = 'V04'THEN 'Patient not covered under this policy'
            WHEN A.ACTION_CD = 'V05'THEN 'Carrier issued duplicate payment'
            WHEN A.ACTION_CD = 'V06'THEN 'Type of service not eligible for coverage'
            WHEN A.ACTION_CD = 'V07'THEN 'Carrier benefits incorrectly calculated'
            WHEN A.ACTION_CD = 'V08'THEN 'Patients maximum has been exceeded'
            WHEN A.ACTION_CD = 'V09'THEN 'Carrier filing limit exceeded'
            WHEN A.ACTION_CD = 'V10'THEN 'Pre-existing conditions not covered under this policy'
            WHEN A.ACTION_CD = 'V11'THEN 'Services not covered without prior authorization'
            WHEN A.ACTION_CD = 'V12'THEN 'Coverage was not coordinated correctly'
            WHEN A.ACTION_CD = 'V13'THEN 'Services rendered by this provider not covered'
            WHEN A.ACTION_CD = 'V14'THEN 'Services not rendered by provider'
            WHEN A.ACTION_CD = 'V15'THEN 'Carrier paid in excess of Medicaid paid amount'
            WHEN A.ACTION_CD = 'V16'THEN 'Member was retro-terminated'
    ELSE NULL
    END AS "HMS DESCRIPTION",

    C.CARRIER_ACTION_CD1 AS "CARRIER ACTION CD1",
    CASE 
    WHEN A.CARRIER_ACTION_CD = '277ACC' THEN '277 ACCEPTANCE/ACKNOWLEDGMENT'
    WHEN A.CARRIER_ACTION_CD = '-999' THEN 'Reason Code Unavailable'
    WHEN A.CARRIER_ACTION_CD = 'ACCDT' THEN 'Reason Code Unavailable'
    WHEN A.CARRIER_ACTION_CD = 'ADJUD' THEN 'Adjudicated per Plan Contract/Allowable'
    WHEN A.CARRIER_ACTION_CD = 'APID' THEN 'Reason Code Unavailable'
    WHEN A.CARRIER_ACTION_CD = 'ARDUP' THEN 'Claim is a Duplicate of Previously Billed Claim'
    WHEN A.CARRIER_ACTION_CD = 'B2BREJ' THEN 'Reason Code Unavailable'
    WHEN A.CARRIER_ACTION_CD = 'BADCLMS' THEN 'Reason Code Unavailable'
    WHEN A.CARRIER_ACTION_CD = 'BADEL' THEN 'Reason Code Unavailable'
    WHEN A.CARRIER_ACTION_CD = 'BILLER' THEN 'Billing Error'
    WHEN A.CARRIER_ACTION_CD = 'BPNAP' THEN 'Reason Code Unavailable'
    WHEN A.CARRIER_ACTION_CD = 'BPNPI' THEN 'Reason Code Unavailable'
    WHEN A.CARRIER_ACTION_CD = 'BPTAX' THEN 'Reason Code Unavailable'
    WHEN A.CARRIER_ACTION_CD = 'BPTXN' THEN 'Reason Code Unavailable'
    WHEN A.CARRIER_ACTION_CD = 'CAID' THEN 'Reason Code Unavailable'
    WHEN A.CARRIER_ACTION_CD = 'CANCEL' THEN 'Reason Code Unavailable'
    WHEN A.CARRIER_ACTION_CD = 'CAPIT' THEN 'Capitated Service'
    WHEN A.CARRIER_ACTION_CD = 'CCLM' THEN 'Corrected Claim Needed'
    WHEN A.CARRIER_ACTION_CD = 'CLAMAJ' THEN 'Care may not be Covered by Another Payer per Coordination of Benefits'
    WHEN A.CARRIER_ACTION_CD = 'CLMAJ' THEN 'Claim Adjusted'
    WHEN A.CARRIER_ACTION_CD = 'CLMFRQ' THEN 'Reason Code Unavailable'
    WHEN A.CARRIER_ACTION_CD = 'CLMFWD' THEN 'Claim Forwarded to Payer by NEIC'
    WHEN A.CARRIER_ACTION_CD = 'CMPMI' THEN 'Missing/Invalid Compound Code'
    WHEN A.CARRIER_ACTION_CD = 'CNBC' THEN 'Reason Code Unavailable'
    WHEN A.CARRIER_ACTION_CD = 'CO45' THEN 'Charge exceeds fee schedule/maximum allowable or contracted/legislated fee arrangement. Usage: This adjustment amount cannot equal the total service or claim charge amount; and must not duplicate provider adjustment amounts (payments and contractual reductions) that have resulted from prior payer(s) adjudication. (Use only with Group Codes PR or CO depending upon liability)'
    WHEN A.CARRIER_ACTION_CD = 'COBPD' THEN 'Paid in Accordance with COB'
    WHEN A.CARRIER_ACTION_CD = 'COINS' THEN 'MA Paid Less Than Co-Insurance'
    WHEN A.CARRIER_ACTION_CD = 'COPAY' THEN 'MA Paid Less Than Copay'
    WHEN A.CARRIER_ACTION_CD = 'DAVITA' THEN 'Reason Code Unavailable'
    WHEN A.CARRIER_ACTION_CD = 'DAYSUP' THEN 'Days Supply Exceeds Plan Limits'
    WHEN A.CARRIER_ACTION_CD = 'DEDUC' THEN 'Payment Applied to Patient Deductible'
    WHEN A.CARRIER_ACTION_CD = 'DEFER' THEN 'Portion of Payment Deferred'
    WHEN A.CARRIER_ACTION_CD = 'DENTL' THEN 'Group Has Dental Coverage Only'
    WHEN A.CARRIER_ACTION_CD = 'DEPNE' THEN 'Dependent Not Eligible'
    WHEN A.CARRIER_ACTION_CD = 'DEPPG' THEN 'Dependent Pregnancy Not Covered'
    WHEN A.CARRIER_ACTION_CD = 'DISDT' THEN 'Reason Code Unavailable'
    WHEN A.CARRIER_ACTION_CD = 'DISSTT' THEN 'Reason Code Unavailable'
    WHEN A.CARRIER_ACTION_CD = 'DOSNE' THEN 'Patient Not Eligible on Date of Service'
    WHEN A.CARRIER_ACTION_CD = 'DPWMI' THEN 'Date Prescription Written Missing/Invalid'
    WHEN A.CARRIER_ACTION_CD = 'DRGNC' THEN 'Drug Not Covered'
    WHEN A.CARRIER_ACTION_CD = 'DSPMI' THEN 'Missing/Invalid Dispense as Written Code'
    WHEN A.CARRIER_ACTION_CD = 'DUP' THEN 'Duplicate Claim Submission'
    WHEN A.CARRIER_ACTION_CD = 'DURCF' THEN 'Insert Fail DUR-Conflict'
    WHEN A.CARRIER_ACTION_CD = 'DXNC' THEN 'Diagnosis Not Covered'
    WHEN A.CARRIER_ACTION_CD = 'EDITGOV' THEN 'Reason Code Unavailable'
    WHEN A.CARRIER_ACTION_CD = 'EMPNE' THEN 'Employee not Eligible'
    WHEN A.CARRIER_ACTION_CD = 'ENDUP' THEN 'Reason Code Unavailable'
    WHEN A.CARRIER_ACTION_CD = 'ERRDNA' THEN 'Reason Code Unavailable'
    WHEN A.CARRIER_ACTION_CD = 'ERREOB' THEN 'Reason Code Unavailable'
    WHEN A.CARRIER_ACTION_CD = 'EXFREQ' THEN 'Service Exceeds Approved Frequency'
    WHEN A.CARRIER_ACTION_CD = 'EXPER' THEN 'Procedure Considered Experimental'
    WHEN A.CARRIER_ACTION_CD = 'EXRCR' THEN 'Exceeds Reasonable and Customary Rate'
    WHEN A.CARRIER_ACTION_CD = 'FRMLRY' THEN 'Product Not On Formulary'
    WHEN A.CARRIER_ACTION_CD = 'GENREJ' THEN 'Reason Code Unavailable'
    WHEN A.CARRIER_ACTION_CD = 'GRPNE' THEN 'Group Not Eligible'
    WHEN A.CARRIER_ACTION_CD = 'GRPSTL' THEN 'Reason Code Unavailable'
    WHEN A.CARRIER_ACTION_CD = 'HCPCS' THEN 'Reason Code Unavailable'
    WHEN A.CARRIER_ACTION_CD = 'HOSPIC' THEN 'Hospice Care Not Covered'
    WHEN A.CARRIER_ACTION_CD = 'ICD10' THEN 'Reason Code Unavailable'
    WHEN A.CARRIER_ACTION_CD = 'ICNPD' THEN 'DO NOT USE - Duplicate Claim Closed Out - ICN billed to and paid by different carrier (Auto Deny Job'
    WHEN A.CARRIER_ACTION_CD = 'ICOB' THEN 'COB Information Needed'
    WHEN A.CARRIER_ACTION_CD = 'IDAYS' THEN 'Missing/Invalid Days Supply'
    WHEN A.CARRIER_ACTION_CD = 'IDEA' THEN 'Missing/Invalid DEA Number'
    WHEN A.CARRIER_ACTION_CD = 'IDIG' THEN 'DX Code Required'
    WHEN A.CARRIER_ACTION_CD = 'IDOB' THEN 'Invalid/Missing DOB'
    WHEN A.CARRIER_ACTION_CD = 'IDOS' THEN 'Date(s) of Service Needed'
    WHEN A.CARRIER_ACTION_CD = 'IEOB' THEN 'Primary EOB Needed'
    WHEN A.CARRIER_ACTION_CD = 'ILAC' THEN 'Nature of Illness or Accident Required'
    WHEN A.CARRIER_ACTION_CD = 'IMCV' THEN 'Medicare EOB Required'
    WHEN A.CARRIER_ACTION_CD = 'INAME' THEN 'Invalid Character in Name or Illegible Name'
    WHEN A.CARRIER_ACTION_CD = 'INCL' THEN 'Carrier will not Pay Separately for this Service'
    WHEN A.CARRIER_ACTION_CD = 'INDC' THEN 'Missing or Invalid NDC Number'
    WHEN A.CARRIER_ACTION_CD = 'INFO' THEN 'Additional Information Needed'
    WHEN A.CARRIER_ACTION_CD = 'INGEX' THEN 'Reject - Ingredient Cost Exceeds Plan Maximum'
    WHEN A.CARRIER_ACTION_CD = 'INGMX' THEN 'Ingredient Cost Reduced to Maximum'
    WHEN A.CARRIER_ACTION_CD = 'INOPID' THEN 'Reason Code Unavailable'
    WHEN A.CARRIER_ACTION_CD = 'INPI' THEN 'Missing/Invalid NPI Number'
    WHEN A.CARRIER_ACTION_CD = 'INPRC' THEN 'Invalid Procedure Code'
    WHEN A.CARRIER_ACTION_CD = 'INSADD' THEN 'Reason Code Unavailable'
    WHEN A.CARRIER_ACTION_CD = 'INVADD' THEN 'Invalid/Incomplete Address'
    WHEN A.CARRIER_ACTION_CD = 'INVBN' THEN 'Missing/Invalid BIN Number'
    WHEN A.CARRIER_ACTION_CD = 'INVBT' THEN 'Invalid Bill Type'
    WHEN A.CARRIER_ACTION_CD = 'INVCD' THEN 'Invalid Code'
    WHEN A.CARRIER_ACTION_CD = 'INVDOS' THEN 'Invalid Date of Service'
    WHEN A.CARRIER_ACTION_CD = 'INVDT' THEN 'Invalid Admit Date'
    WHEN A.CARRIER_ACTION_CD = 'INVDX' THEN 'DX Not Valid for Procedure'
    WHEN A.CARRIER_ACTION_CD = 'INVGN' THEN 'Invalid Gender'
    WHEN A.CARRIER_ACTION_CD = 'INVGR' THEN 'UNKNOWN'
    WHEN A.CARRIER_ACTION_CD = 'INVGRP' THEN 'Invalid Group Number'
    WHEN A.CARRIER_ACTION_CD = 'INVID' THEN 'Invalid ID Number'
    WHEN A.CARRIER_ACTION_CD = 'INVMOD' THEN 'Missing/Invalid Modifier'
    WHEN A.CARRIER_ACTION_CD = 'INVNAM' THEN 'Invalid Name'
    WHEN A.CARRIER_ACTION_CD = 'INVND' THEN 'Non-Matched Service Provider Id'
    WHEN A.CARRIER_ACTION_CD = 'INVPCN' THEN 'Missing/Invalid PCN Number'
    WHEN A.CARRIER_ACTION_CD = 'INVPL' THEN 'Invalid or Incomplete Protocol Requirements'
    WHEN A.CARRIER_ACTION_CD = 'INVPOS' THEN 'Incorrect Place of Service'
    WHEN A.CARRIER_ACTION_CD = 'INVPY' THEN 'Invalid Payee Code for Medicaid Agency'
    WHEN A.CARRIER_ACTION_CD = 'INVQT' THEN 'Invalid Quantity Entered for Medication Package '
    WHEN A.CARRIER_ACTION_CD = 'INVUC' THEN 'Invalid Usual and Customary Amount'
    WHEN A.CARRIER_ACTION_CD = 'IPRO' THEN 'Provider Information Required'
    WHEN A.CARRIER_ACTION_CD = 'IREC' THEN 'Medical Records Needed'
    WHEN A.CARRIER_ACTION_CD = 'IREL' THEN 'Patient Relation to Insured Required'
    WHEN A.CARRIER_ACTION_CD = 'IRNPI' THEN 'Missing/Invalid Referring NPI'
    WHEN A.CARRIER_ACTION_CD = 'ITAX' THEN 'Need Tax ID Information'
    WHEN A.CARRIER_ACTION_CD = 'ITMZ' THEN 'Itemized Bill Required'
    WHEN A.CARRIER_ACTION_CD = 'KEYED' THEN 'Denial Code Keyed Incorrectly'
    WHEN A.CARRIER_ACTION_CD = 'LIABL' THEN 'Injury/Illness is Covered by the Liability Carrier'
    WHEN A.CARRIER_ACTION_CD = 'LINE' THEN 'Reason Code Unavailable'
    WHEN A.CARRIER_ACTION_CD = 'LOWDOL' THEN 'Reason Code Unavailable'
    WHEN A.CARRIER_ACTION_CD = 'M86' THEN 'Service denied because payment already made for same/similar procedure within set time frame.'
    WHEN A.CARRIER_ACTION_CD = 'MA04' THEN 'Secondary payment cannot be considered without the identity of or payment information from the primary payer. The information was either not reported or was illegible.'
    WHEN A.CARRIER_ACTION_CD = 'MA130' THEN 'Your claim contains incomplete and/or invalid information, and no appeal rights are afforded because the claim is unprocessable. Please submit a new claim with the complete/correct information.'
    WHEN A.CARRIER_ACTION_CD = 'MA66' THEN 'Missing/incomplete/invalid principal procedure code.'
    WHEN A.CARRIER_ACTION_CD = 'MAXBEN' THEN 'Maximum Benefits Reached'
    WHEN A.CARRIER_ACTION_CD = 'MAXBN' THEN 'UNKNOWN'
    WHEN A.CARRIER_ACTION_CD = 'MCAID' THEN 'Reason Code Unavailable'
    WHEN A.CARRIER_ACTION_CD = 'MEDAJ' THEN 'Medicaid Claim Adjudicated'
    WHEN A.CARRIER_ACTION_CD = 'MEDNCO' THEN 'Not a Medicare Covered Service'
    WHEN A.CARRIER_ACTION_CD = 'MEDNCOV' THEN 'UNKNOWN'
    WHEN A.CARRIER_ACTION_CD = 'MEDNEC' THEN 'Claim Not Medically Necessary'
    WHEN A.CARRIER_ACTION_CD = 'N122' THEN 'Add-on code cannot be billed by itself.'
    WHEN A.CARRIER_ACTION_CD = 'N130' THEN 'Consult plan benefit documents/guidelines for information about restrictions for this service.'
    WHEN A.CARRIER_ACTION_CD = 'N179' THEN 'Additional information has been requested from the member. The charges will be reconsidered upon receipt of that information.'
    WHEN A.CARRIER_ACTION_CD = 'N30' THEN 'Patient ineligible for this service.'
    WHEN A.CARRIER_ACTION_CD = 'N465' THEN 'Missing Physical Therapy Notes/Report.'
    WHEN A.CARRIER_ACTION_CD = 'N522' THEN 'Duplicate of a claim processed, or to be processed, as a crossover claim.'
    WHEN A.CARRIER_ACTION_CD = 'N525' THEN 'These services are not covered when performed within the global period of another service.'
    WHEN A.CARRIER_ACTION_CD = 'N578' THEN 'Coverages do not apply to this loss.'
    WHEN A.CARRIER_ACTION_CD = 'N674' THEN 'Not covered unless a pre-requisite procedure/service has been provided.'
    WHEN A.CARRIER_ACTION_CD = 'NABP' THEN 'Carrier Needs NABP Number'
    WHEN A.CARRIER_ACTION_CD = 'NBPT' THEN 'Non-Billable Provider Type'
    WHEN A.CARRIER_ACTION_CD = 'NDCNC' THEN 'NDC code Not Covered'
    WHEN A.CARRIER_ACTION_CD = 'NEICPD' THEN 'Payment Per Negotiated Rate'
    WHEN A.CARRIER_ACTION_CD = 'NENTBP' THEN 'Reason Code Unavailable'
    WHEN A.CARRIER_ACTION_CD = 'NENTIN' THEN 'Reason Code Unavailable'
    WHEN A.CARRIER_ACTION_CD = 'NENTPT' THEN 'Reason Code Unavailable'
    WHEN A.CARRIER_ACTION_CD = 'NENTRF' THEN 'Reason Code Unavailable'
    WHEN A.CARRIER_ACTION_CD = 'NOAUTH' THEN 'No Authorization Obtained'
    WHEN A.CARRIER_ACTION_CD = 'NOBIL' THEN 'Claim Should Not Have Been Billed per Client - Do Not Work/Rebill Claims'
    WHEN A.CARRIER_ACTION_CD = 'NOCOD' THEN 'No Denial Code on Remittance'
    WHEN A.CARRIER_ACTION_CD = 'NOCOV' THEN 'No Coverage'
    WHEN A.CARRIER_ACTION_CD = 'NOELG' THEN 'Reason Code Unavailable'
    WHEN A.CARRIER_ACTION_CD = 'NONURS' THEN 'Home Nursing Services Not Covered'
    WHEN A.CARRIER_ACTION_CD = 'NOPAY' THEN 'No Payment to be Issued'
    WHEN A.CARRIER_ACTION_CD = 'NOTES' THEN 'Physicians Orders or Nursing Notes Needed'
    WHEN A.CARRIER_ACTION_CD = 'NOXWK' THEN 'No-Crosswalk'
    WHEN A.CARRIER_ACTION_CD = 'NPHARM' THEN 'Non-Matched Pharmacy Number'
    WHEN A.CARRIER_ACTION_CD = 'NUBCCD' THEN 'Reason Code Unavailable'
    WHEN A.CARRIER_ACTION_CD = 'NUBCVL' THEN 'Reason Code Unavailable'
    WHEN A.CARRIER_ACTION_CD = 'OA133' THEN 'The disposition of this service line is pending further review. (Use only with Group Code OA). Usage: Use of this code requires a reversal and correction when the service line is finalized (use only in Loop 2110 CAS segment of the 835 or Loop 2430 of the 837).'
    WHEN A.CARRIER_ACTION_CD = 'OA18' THEN 'Exact duplicate claim/service (Use only with Group Code OA except where state workers compensation regulations requires CO)'
    WHEN A.CARRIER_ACTION_CD = 'OA23' THEN 'The impact of prior payer(s) adjudication including payments and/or adjustments. (Use only with Group Code OA)'
    WHEN A.CARRIER_ACTION_CD = 'OAB13' THEN 'Previously paid. Payment for this claim/service may have been provided in a previous payment.'
    WHEN A.CARRIER_ACTION_CD = 'OUTNW' THEN 'Out of Network'
    WHEN A.CARRIER_ACTION_CD = 'P' THEN 'Paid'
    WHEN A.CARRIER_ACTION_CD = 'PARTB' THEN 'Claim not Processed - Medicare Part B policy'
    WHEN A.CARRIER_ACTION_CD = 'PCFDNY' THEN 'Reason Code Unavailable'
    WHEN A.CARRIER_ACTION_CD = 'PCFVOID' THEN 'Reason Code Unavailable'
    WHEN A.CARRIER_ACTION_CD = 'PCPRE' THEN 'UNKNOWN'
    WHEN A.CARRIER_ACTION_CD = 'PDIA' THEN 'Invalid Principal/Admit Diagnosis'
    WHEN A.CARRIER_ACTION_CD = 'PDPVPB' THEN 'Paid to Provider Prior to HMS Billing'
    WHEN A.CARRIER_ACTION_CD = 'PDTOPV' THEN 'Claim Paid to Provider of Service'
    WHEN A.CARRIER_ACTION_CD = 'PEND' THEN 'Claim Pended'
    WHEN A.CARRIER_ACTION_CD = 'PHARNE' THEN 'Pharmacy Not Eligible'
    WHEN A.CARRIER_ACTION_CD = 'PHYNC' THEN 'Physician Not Covered '
    WHEN A.CARRIER_ACTION_CD = 'PR1' THEN 'Deductible Amount'
    WHEN A.CARRIER_ACTION_CD = 'PR20' THEN 'This injury/illness is covered by the liability carrier.'
    WHEN A.CARRIER_ACTION_CD = 'PR204' THEN 'This service/equipment/drug is not covered under the patients current benefit plan'
    WHEN A.CARRIER_ACTION_CD = 'PR27' THEN 'Expenses incurred after coverage terminated.'
    WHEN A.CARRIER_ACTION_CD = 'PR33' THEN 'Insured has no dependent coverage.'
    WHEN A.CARRIER_ACTION_CD = 'PR49' THEN 'This is a non-covered service because it is a routine/preventive exam or a diagnostic/screening procedure done in conjunction with a routine/preventive exam. Usage: Refer to the 835 Healthcare Policy Identification Segment (loop 2110 Service Payment Information REF), if present.'
    WHEN A.CARRIER_ACTION_CD = 'PRB7' THEN 'This provider was not certified/eligible to be paid for this procedure/service on this date of service. Usage: Refer to the 835 Healthcare Policy Identification Segment (loop 2110 Service Payment Information REF), if present.'
    WHEN A.CARRIER_ACTION_CD = 'PRCNC' THEN 'Procedure Code Not Covered'
    WHEN A.CARRIER_ACTION_CD = 'PRPRC' THEN 'Reason Code Unavailable'
    WHEN A.CARRIER_ACTION_CD = 'PRPRD' THEN 'Reason Code Unavailable'
    WHEN A.CARRIER_ACTION_CD = 'PRSC' THEN 'Expected a Prescriber Field'
    WHEN A.CARRIER_ACTION_CD = 'PRVNC' THEN 'Provider Not Covered'
    WHEN A.CARRIER_ACTION_CD = 'PTAGE' THEN 'Patient Outside Age Limit for this Type of Benefit'
    WHEN A.CARRIER_ACTION_CD = 'PTINFO' THEN 'Need Information From Patient'
    WHEN A.CARRIER_ACTION_CD = 'PTRES' THEN 'Patient Responsibility'
    WHEN A.CARRIER_ACTION_CD = 'QUAL' THEN 'Qualifying Procedure Not Received'
    WHEN A.CARRIER_ACTION_CD = 'QUANT' THEN 'Quantity Not Covered'
    WHEN A.CARRIER_ACTION_CD = 'REBIL' THEN 'Claim Closed to Pass through New Cycle'
    WHEN A.CARRIER_ACTION_CD = 'RECAL' THEN 'Reason Code Unavailable'
    WHEN A.CARRIER_ACTION_CD = 'RECALL' THEN 'Claims Closed - Completed Client Run Out Phase'
    WHEN A.CARRIER_ACTION_CD = 'REFIL' THEN 'Refill Too Soon'
    WHEN A.CARRIER_ACTION_CD = 'REFMI' THEN 'Refill Number Missing/Invalid'
    WHEN A.CARRIER_ACTION_CD = 'RENPRO' THEN 'Pending Information from Rendering Provider'
    WHEN A.CARRIER_ACTION_CD = 'RESENT' THEN 'Resent Claims'
    WHEN A.CARRIER_ACTION_CD = 'RESUB' THEN 'Submit Claim to Another Entity'
    WHEN A.CARRIER_ACTION_CD = 'REVCOD' THEN 'Reason Code Unavailable'
    WHEN A.CARRIER_ACTION_CD = 'RNID' THEN 'Reason Code Unavailable'
    WHEN A.CARRIER_ACTION_CD = 'RNNAP' THEN 'Reason Code Unavailable'
    WHEN A.CARRIER_ACTION_CD = 'ROMRAT' THEN 'Reason Code Unavailable'
    WHEN A.CARRIER_ACTION_CD = 'ROUTN' THEN 'Routine Services Not Covered'
    WHEN A.CARRIER_ACTION_CD = 'RXAGE' THEN 'Participant Age Restricts Medication Coverage'
    WHEN A.CARRIER_ACTION_CD = 'RXLMT' THEN 'Medication Exceeds Plan Limits'
    WHEN A.CARRIER_ACTION_CD = 'RXNC' THEN 'Prescription Drugs Not Covered'
    WHEN A.CARRIER_ACTION_CD = 'SETTL' THEN 'Claim Covered in Carrier Settlement'
    WHEN A.CARRIER_ACTION_CD = 'SPPBM' THEN 'Must Fill through Specialty Pharmacy'
    WHEN A.CARRIER_ACTION_CD = 'SUBID' THEN 'Reason Code Unavailable'
    WHEN A.CARRIER_ACTION_CD = 'TERMGRP' THEN 'Reason Code Unavailable'
    WHEN A.CARRIER_ACTION_CD = 'TIMELY' THEN 'Claim Past Timely Filing Limit'
    WHEN A.CARRIER_ACTION_CD = 'TOSNE' THEN 'Type of Service Not Eligible'
    WHEN A.CARRIER_ACTION_CD = 'TPA' THEN 'Send Claim to TPA'
    WHEN A.CARRIER_ACTION_CD = 'TPLNC' THEN 'Group Does Not Allow Third Party Claims'
    WHEN A.CARRIER_ACTION_CD = 'UMID' THEN 'UNKNOWN'
    WHEN A.CARRIER_ACTION_CD = 'UNID' THEN 'Unable to Identify Member'
    WHEN A.CARRIER_ACTION_CD = 'UNITS' THEN 'Units Field Invalid for Number of Days'
    WHEN A.CARRIER_ACTION_CD = 'UNPRO' THEN 'Unprocessed Claim'
    WHEN A.CARRIER_ACTION_CD = 'WASTE' THEN 'Reason Code Unavailable'
    WHEN A.CARRIER_ACTION_CD = '' THEN 'No Data Provided'
    ELSE 'NOCOD'
            END  AS   "CARRIER DESCRIPTION1",
            
    C.CARRIER_ACTION_CD2 AS "CARRIER ACTION CD2",
    CASE 
    WHEN C.CARRIER_ACTION_CD2 = '277ACC' THEN '277 ACCEPTANCE/ACKNOWLEDGMENT'
    WHEN C.CARRIER_ACTION_CD2 = '-999' THEN 'Reason Code Unavailable'
    WHEN C.CARRIER_ACTION_CD2 = 'ACCDT' THEN 'Reason Code Unavailable'
    WHEN C.CARRIER_ACTION_CD2 = 'ADJUD' THEN 'Adjudicated per Plan Contract/Allowable'
    WHEN C.CARRIER_ACTION_CD2 = 'APID' THEN 'Reason Code Unavailable'
    WHEN C.CARRIER_ACTION_CD2 = 'ARDUP' THEN 'Claim is a Duplicate of Previously Billed Claim'
    WHEN C.CARRIER_ACTION_CD2 = 'B2BREJ' THEN 'Reason Code Unavailable'
    WHEN C.CARRIER_ACTION_CD2 = 'BADCLMS' THEN 'Reason Code Unavailable'
    WHEN C.CARRIER_ACTION_CD2 = 'BADEL' THEN 'Reason Code Unavailable'
    WHEN C.CARRIER_ACTION_CD2 = 'BILLER' THEN 'Billing Error'
    WHEN C.CARRIER_ACTION_CD2 = 'BPNAP' THEN 'Reason Code Unavailable'
    WHEN C.CARRIER_ACTION_CD2 = 'BPNPI' THEN 'Reason Code Unavailable'
    WHEN C.CARRIER_ACTION_CD2 = 'BPTAX' THEN 'Reason Code Unavailable'
    WHEN C.CARRIER_ACTION_CD2 = 'BPTXN' THEN 'Reason Code Unavailable'
    WHEN C.CARRIER_ACTION_CD2 = 'CAID' THEN 'Reason Code Unavailable'
    WHEN C.CARRIER_ACTION_CD2 = 'CANCEL' THEN 'Reason Code Unavailable'
    WHEN C.CARRIER_ACTION_CD2 = 'CAPIT' THEN 'Capitated Service'
    WHEN C.CARRIER_ACTION_CD2 = 'CCLM' THEN 'Corrected Claim Needed'
    WHEN C.CARRIER_ACTION_CD2 = 'CLAMAJ' THEN 'Care may not be Covered by Another Payer per Coordination of Benefits'
    WHEN C.CARRIER_ACTION_CD2 = 'CLMAJ' THEN 'Claim Adjusted'
    WHEN C.CARRIER_ACTION_CD2 = 'CLMFRQ' THEN 'Reason Code Unavailable'
    WHEN C.CARRIER_ACTION_CD2 = 'CLMFWD' THEN 'Claim Forwarded to Payer by NEIC'
    WHEN C.CARRIER_ACTION_CD2 = 'CMPMI' THEN 'Missing/Invalid Compound Code'
    WHEN C.CARRIER_ACTION_CD2 = 'CNBC' THEN 'Reason Code Unavailable'
    WHEN C.CARRIER_ACTION_CD2 = 'CO45' THEN 'Charge exceeds fee schedule/maximum allowable or contracted/legislated fee arrangement. Usage: This adjustment amount cannot equal the total service or claim charge amount; and must not duplicate provider adjustment amounts (payments and contractual reductions) that have resulted from prior payer(s) adjudication. (Use only with Group Codes PR or CO depending upon liability)'
    WHEN C.CARRIER_ACTION_CD2 = 'COBPD' THEN 'Paid in Accordance with COB'
    WHEN C.CARRIER_ACTION_CD2 = 'COINS' THEN 'MA Paid Less Than Co-Insurance'
    WHEN C.CARRIER_ACTION_CD2 = 'COPAY' THEN 'MA Paid Less Than Copay'
    WHEN C.CARRIER_ACTION_CD2 = 'DAVITA' THEN 'Reason Code Unavailable'
    WHEN C.CARRIER_ACTION_CD2 = 'DAYSUP' THEN 'Days Supply Exceeds Plan Limits'
    WHEN C.CARRIER_ACTION_CD2 = 'DEDUC' THEN 'Payment Applied to Patient Deductible'
    WHEN C.CARRIER_ACTION_CD2 = 'DEFER' THEN 'Portion of Payment Deferred'
    WHEN C.CARRIER_ACTION_CD2 = 'DENTL' THEN 'Group Has Dental Coverage Only'
    WHEN C.CARRIER_ACTION_CD2 = 'DEPNE' THEN 'Dependent Not Eligible'
    WHEN C.CARRIER_ACTION_CD2 = 'DEPPG' THEN 'Dependent Pregnancy Not Covered'
    WHEN C.CARRIER_ACTION_CD2 = 'DISDT' THEN 'Reason Code Unavailable'
    WHEN C.CARRIER_ACTION_CD2 = 'DISSTT' THEN 'Reason Code Unavailable'
    WHEN C.CARRIER_ACTION_CD2 = 'DOSNE' THEN 'Patient Not Eligible on Date of Service'
    WHEN C.CARRIER_ACTION_CD2 = 'DPWMI' THEN 'Date Prescription Written Missing/Invalid'
    WHEN C.CARRIER_ACTION_CD2 = 'DRGNC' THEN 'Drug Not Covered'
    WHEN C.CARRIER_ACTION_CD2 = 'DSPMI' THEN 'Missing/Invalid Dispense as Written Code'
    WHEN C.CARRIER_ACTION_CD2 = 'DUP' THEN 'Duplicate Claim Submission'
    WHEN C.CARRIER_ACTION_CD2 = 'DURCF' THEN 'Insert Fail DUR-Conflict'
    WHEN C.CARRIER_ACTION_CD2 = 'DXNC' THEN 'Diagnosis Not Covered'
    WHEN C.CARRIER_ACTION_CD2 = 'EDITGOV' THEN 'Reason Code Unavailable'
    WHEN C.CARRIER_ACTION_CD2 = 'EMPNE' THEN 'Employee not Eligible'
    WHEN C.CARRIER_ACTION_CD2 = 'ENDUP' THEN 'Reason Code Unavailable'
    WHEN C.CARRIER_ACTION_CD2 = 'ERRDNA' THEN 'Reason Code Unavailable'
    WHEN C.CARRIER_ACTION_CD2 = 'ERREOB' THEN 'Reason Code Unavailable'
    WHEN C.CARRIER_ACTION_CD2 = 'EXFREQ' THEN 'Service Exceeds Approved Frequency'
    WHEN C.CARRIER_ACTION_CD2 = 'EXPER' THEN 'Procedure Considered Experimental'
    WHEN C.CARRIER_ACTION_CD2 = 'EXRCR' THEN 'Exceeds Reasonable and Customary Rate'
    WHEN C.CARRIER_ACTION_CD2 = 'FRMLRY' THEN 'Product Not On Formulary'
    WHEN C.CARRIER_ACTION_CD2 = 'GENREJ' THEN 'Reason Code Unavailable'
    WHEN C.CARRIER_ACTION_CD2 = 'GRPNE' THEN 'Group Not Eligible'
    WHEN C.CARRIER_ACTION_CD2 = 'GRPSTL' THEN 'Reason Code Unavailable'
    WHEN C.CARRIER_ACTION_CD2 = 'HCPCS' THEN 'Reason Code Unavailable'
    WHEN C.CARRIER_ACTION_CD2 = 'HOSPIC' THEN 'Hospice Care Not Covered'
    WHEN C.CARRIER_ACTION_CD2 = 'ICD10' THEN 'Reason Code Unavailable'
    WHEN C.CARRIER_ACTION_CD2 = 'ICNPD' THEN 'DO NOT USE - Duplicate Claim Closed Out - ICN billed to and paid by different carrier (Auto Deny Job'
    WHEN C.CARRIER_ACTION_CD2 = 'ICOB' THEN 'COB Information Needed'
    WHEN C.CARRIER_ACTION_CD2 = 'IDAYS' THEN 'Missing/Invalid Days Supply'
    WHEN C.CARRIER_ACTION_CD2 = 'IDEA' THEN 'Missing/Invalid DEA Number'
    WHEN C.CARRIER_ACTION_CD2 = 'IDIG' THEN 'DX Code Required'
    WHEN C.CARRIER_ACTION_CD2 = 'IDOB' THEN 'Invalid/Missing DOB'
    WHEN C.CARRIER_ACTION_CD2 = 'IDOS' THEN 'Date(s) of Service Needed'
    WHEN C.CARRIER_ACTION_CD2 = 'IEOB' THEN 'Primary EOB Needed'
    WHEN C.CARRIER_ACTION_CD2 = 'ILAC' THEN 'Nature of Illness or Accident Required'
    WHEN C.CARRIER_ACTION_CD2 = 'IMCV' THEN 'Medicare EOB Required'
    WHEN C.CARRIER_ACTION_CD2 = 'INAME' THEN 'Invalid Character in Name or Illegible Name'
    WHEN C.CARRIER_ACTION_CD2 = 'INCL' THEN 'Carrier will not Pay Separately for this Service'
    WHEN C.CARRIER_ACTION_CD2 = 'INDC' THEN 'Missing or Invalid NDC Number'
    WHEN C.CARRIER_ACTION_CD2 = 'INFO' THEN 'Additional Information Needed'
    WHEN C.CARRIER_ACTION_CD2 = 'INGEX' THEN 'Reject - Ingredient Cost Exceeds Plan Maximum'
    WHEN C.CARRIER_ACTION_CD2 = 'INGMX' THEN 'Ingredient Cost Reduced to Maximum'
    WHEN C.CARRIER_ACTION_CD2 = 'INOPID' THEN 'Reason Code Unavailable'
    WHEN C.CARRIER_ACTION_CD2 = 'INPI' THEN 'Missing/Invalid NPI Number'
    WHEN C.CARRIER_ACTION_CD2 = 'INPRC' THEN 'Invalid Procedure Code'
    WHEN C.CARRIER_ACTION_CD2 = 'INSADD' THEN 'Reason Code Unavailable'
    WHEN C.CARRIER_ACTION_CD2 = 'INVADD' THEN 'Invalid/Incomplete Address'
    WHEN C.CARRIER_ACTION_CD2 = 'INVBN' THEN 'Missing/Invalid BIN Number'
    WHEN C.CARRIER_ACTION_CD2 = 'INVBT' THEN 'Invalid Bill Type'
    WHEN C.CARRIER_ACTION_CD2 = 'INVCD' THEN 'Invalid Code'
    WHEN C.CARRIER_ACTION_CD2 = 'INVDOS' THEN 'Invalid Date of Service'
    WHEN C.CARRIER_ACTION_CD2 = 'INVDT' THEN 'Invalid Admit Date'
    WHEN C.CARRIER_ACTION_CD2 = 'INVDX' THEN 'DX Not Valid for Procedure'
    WHEN C.CARRIER_ACTION_CD2 = 'INVGN' THEN 'Invalid Gender'
    WHEN C.CARRIER_ACTION_CD2 = 'INVGR' THEN 'UNKNOWN'
    WHEN C.CARRIER_ACTION_CD2 = 'INVGRP' THEN 'Invalid Group Number'
    WHEN C.CARRIER_ACTION_CD2 = 'INVID' THEN 'Invalid ID Number'
    WHEN C.CARRIER_ACTION_CD2 = 'INVMOD' THEN 'Missing/Invalid Modifier'
    WHEN C.CARRIER_ACTION_CD2 = 'INVNAM' THEN 'Invalid Name'
    WHEN C.CARRIER_ACTION_CD2 = 'INVND' THEN 'Non-Matched Service Provider Id'
    WHEN C.CARRIER_ACTION_CD2 = 'INVPCN' THEN 'Missing/Invalid PCN Number'
    WHEN C.CARRIER_ACTION_CD2 = 'INVPL' THEN 'Invalid or Incomplete Protocol Requirements'
    WHEN C.CARRIER_ACTION_CD2 = 'INVPOS' THEN 'Incorrect Place of Service'
    WHEN C.CARRIER_ACTION_CD2 = 'INVPY' THEN 'Invalid Payee Code for Medicaid Agency'
    WHEN C.CARRIER_ACTION_CD2 = 'INVQT' THEN 'Invalid Quantity Entered for Medication Package '
    WHEN C.CARRIER_ACTION_CD2 = 'INVUC' THEN 'Invalid Usual and Customary Amount'
    WHEN C.CARRIER_ACTION_CD2 = 'IPRO' THEN 'Provider Information Required'
    WHEN C.CARRIER_ACTION_CD2 = 'IREC' THEN 'Medical Records Needed'
    WHEN C.CARRIER_ACTION_CD2 = 'IREL' THEN 'Patient Relation to Insured Required'
    WHEN C.CARRIER_ACTION_CD2 = 'IRNPI' THEN 'Missing/Invalid Referring NPI'
    WHEN C.CARRIER_ACTION_CD2 = 'ITAX' THEN 'Need Tax ID Information'
    WHEN C.CARRIER_ACTION_CD2 = 'ITMZ' THEN 'Itemized Bill Required'
    WHEN C.CARRIER_ACTION_CD2 = 'KEYED' THEN 'Denial Code Keyed Incorrectly'
    WHEN C.CARRIER_ACTION_CD2 = 'LIABL' THEN 'Injury/Illness is Covered by the Liability Carrier'
    WHEN C.CARRIER_ACTION_CD2 = 'LINE' THEN 'Reason Code Unavailable'
    WHEN C.CARRIER_ACTION_CD2 = 'LOWDOL' THEN 'Reason Code Unavailable'
    WHEN C.CARRIER_ACTION_CD2 = 'M86' THEN 'Service denied because payment already made for same/similar procedure within set time frame.'
    WHEN C.CARRIER_ACTION_CD2 = 'MA04' THEN 'Secondary payment cannot be considered without the identity of or payment information from the primary payer. The information was either not reported or was illegible.'
    WHEN C.CARRIER_ACTION_CD2 = 'MA130' THEN 'Your claim contains incomplete and/or invalid information, and no appeal rights are afforded because the claim is unprocessable. Please submit a new claim with the complete/correct information.'
    WHEN C.CARRIER_ACTION_CD2 = 'MA66' THEN 'Missing/incomplete/invalid principal procedure code.'
    WHEN C.CARRIER_ACTION_CD2 = 'MAXBEN' THEN 'Maximum Benefits Reached'
    WHEN C.CARRIER_ACTION_CD2 = 'MAXBN' THEN 'UNKNOWN'
    WHEN C.CARRIER_ACTION_CD2 = 'MCAID' THEN 'Reason Code Unavailable'
    WHEN C.CARRIER_ACTION_CD2 = 'MEDAJ' THEN 'Medicaid Claim Adjudicated'
    WHEN C.CARRIER_ACTION_CD2 = 'MEDNCO' THEN 'Not a Medicare Covered Service'
    WHEN C.CARRIER_ACTION_CD2 = 'MEDNCOV' THEN 'UNKNOWN'
    WHEN C.CARRIER_ACTION_CD2 = 'MEDNEC' THEN 'Claim Not Medically Necessary'
    WHEN C.CARRIER_ACTION_CD2 = 'N122' THEN 'Add-on code cannot be billed by itself.'
    WHEN C.CARRIER_ACTION_CD2 = 'N130' THEN 'Consult plan benefit documents/guidelines for information about restrictions for this service.'
    WHEN C.CARRIER_ACTION_CD2 = 'N179' THEN 'Additional information has been requested from the member. The charges will be reconsidered upon receipt of that information.'
    WHEN C.CARRIER_ACTION_CD2 = 'N30' THEN 'Patient ineligible for this service.'
    WHEN C.CARRIER_ACTION_CD2 = 'N465' THEN 'Missing Physical Therapy Notes/Report.'
    WHEN C.CARRIER_ACTION_CD2 = 'N522' THEN 'Duplicate of a claim processed, or to be processed, as a crossover claim.'
    WHEN C.CARRIER_ACTION_CD2 = 'N525' THEN 'These services are not covered when performed within the global period of another service.'
    WHEN C.CARRIER_ACTION_CD2 = 'N578' THEN 'Coverages do not apply to this loss.'
    WHEN C.CARRIER_ACTION_CD2 = 'N674' THEN 'Not covered unless a pre-requisite procedure/service has been provided.'
    WHEN C.CARRIER_ACTION_CD2 = 'NABP' THEN 'Carrier Needs NABP Number'
    WHEN C.CARRIER_ACTION_CD2 = 'NBPT' THEN 'Non-Billable Provider Type'
    WHEN C.CARRIER_ACTION_CD2 = 'NDCNC' THEN 'NDC code Not Covered'
    WHEN C.CARRIER_ACTION_CD2 = 'NEICPD' THEN 'Payment Per Negotiated Rate'
    WHEN C.CARRIER_ACTION_CD2 = 'NENTBP' THEN 'Reason Code Unavailable'
    WHEN C.CARRIER_ACTION_CD2 = 'NENTIN' THEN 'Reason Code Unavailable'
    WHEN C.CARRIER_ACTION_CD2 = 'NENTPT' THEN 'Reason Code Unavailable'
    WHEN C.CARRIER_ACTION_CD2 = 'NENTRF' THEN 'Reason Code Unavailable'
    WHEN C.CARRIER_ACTION_CD2 = 'NOAUTH' THEN 'No Authorization Obtained'
    WHEN C.CARRIER_ACTION_CD2 = 'NOBIL' THEN 'Claim Should Not Have Been Billed per Client - Do Not Work/Rebill Claims'
    WHEN C.CARRIER_ACTION_CD2 = 'NOCOD' THEN 'No Denial Code on Remittance'
    WHEN C.CARRIER_ACTION_CD2 = 'NOCOV' THEN 'No Coverage'
    WHEN C.CARRIER_ACTION_CD2 = 'NOELG' THEN 'Reason Code Unavailable'
    WHEN C.CARRIER_ACTION_CD2 = 'NONURS' THEN 'Home Nursing Services Not Covered'
    WHEN C.CARRIER_ACTION_CD2 = 'NOPAY' THEN 'No Payment to be Issued'
    WHEN C.CARRIER_ACTION_CD2 = 'NOTES' THEN 'Physicians Orders or Nursing Notes Needed'
    WHEN C.CARRIER_ACTION_CD2 = 'NOXWK' THEN 'No-Crosswalk'
    WHEN C.CARRIER_ACTION_CD2 = 'NPHARM' THEN 'Non-Matched Pharmacy Number'
    WHEN C.CARRIER_ACTION_CD2 = 'NUBCCD' THEN 'Reason Code Unavailable'
    WHEN C.CARRIER_ACTION_CD2 = 'NUBCVL' THEN 'Reason Code Unavailable'
    WHEN C.CARRIER_ACTION_CD2 = 'OA133' THEN 'The disposition of this service line is pending further review. (Use only with Group Code OA). Usage: Use of this code requires a reversal and correction when the service line is finalized (use only in Loop 2110 CAS segment of the 835 or Loop 2430 of the 837).'
    WHEN C.CARRIER_ACTION_CD2 = 'OA18' THEN 'Exact duplicate claim/service (Use only with Group Code OA except where state workers compensation regulations requires CO)'
    WHEN C.CARRIER_ACTION_CD2 = 'OA23' THEN 'The impact of prior payer(s) adjudication including payments and/or adjustments. (Use only with Group Code OA)'
    WHEN C.CARRIER_ACTION_CD2 = 'OAB13' THEN 'Previously paid. Payment for this claim/service may have been provided in a previous payment.'
    WHEN C.CARRIER_ACTION_CD2 = 'OUTNW' THEN 'Out of Network'
    WHEN C.CARRIER_ACTION_CD2 = 'P' THEN 'Paid'
    WHEN C.CARRIER_ACTION_CD2 = 'PARTB' THEN 'Claim not Processed - Medicare Part B policy'
    WHEN C.CARRIER_ACTION_CD2 = 'PCFDNY' THEN 'Reason Code Unavailable'
    WHEN C.CARRIER_ACTION_CD2 = 'PCFVOID' THEN 'Reason Code Unavailable'
    WHEN C.CARRIER_ACTION_CD2 = 'PCPRE' THEN 'UNKNOWN'
    WHEN C.CARRIER_ACTION_CD2 = 'PDIA' THEN 'Invalid Principal/Admit Diagnosis'
    WHEN C.CARRIER_ACTION_CD2 = 'PDPVPB' THEN 'Paid to Provider Prior to HMS Billing'
    WHEN C.CARRIER_ACTION_CD2 = 'PDTOPV' THEN 'Claim Paid to Provider of Service'
    WHEN C.CARRIER_ACTION_CD2 = 'PEND' THEN 'Claim Pended'
    WHEN C.CARRIER_ACTION_CD2 = 'PHARNE' THEN 'Pharmacy Not Eligible'
    WHEN C.CARRIER_ACTION_CD2 = 'PHYNC' THEN 'Physician Not Covered '
    WHEN C.CARRIER_ACTION_CD2 = 'PR1' THEN 'Deductible Amount'
    WHEN C.CARRIER_ACTION_CD2 = 'PR20' THEN 'This injury/illness is covered by the liability carrier.'
    WHEN C.CARRIER_ACTION_CD2 = 'PR204' THEN 'This service/equipment/drug is not covered under the patients current benefit plan'
    WHEN C.CARRIER_ACTION_CD2 = 'PR27' THEN 'Expenses incurred after coverage terminated.'
    WHEN C.CARRIER_ACTION_CD2 = 'PR33' THEN 'Insured has no dependent coverage.'
    WHEN C.CARRIER_ACTION_CD2 = 'PR49' THEN 'This is a non-covered service because it is a routine/preventive exam or a diagnostic/screening procedure done in conjunction with a routine/preventive exam. Usage: Refer to the 835 Healthcare Policy Identification Segment (loop 2110 Service Payment Information REF), if present.'
    WHEN C.CARRIER_ACTION_CD2 = 'PRB7' THEN 'This provider was not certified/eligible to be paid for this procedure/service on this date of service. Usage: Refer to the 835 Healthcare Policy Identification Segment (loop 2110 Service Payment Information REF), if present.'
    WHEN C.CARRIER_ACTION_CD2 = 'PRCNC' THEN 'Procedure Code Not Covered'
    WHEN C.CARRIER_ACTION_CD2 = 'PRPRC' THEN 'Reason Code Unavailable'
    WHEN C.CARRIER_ACTION_CD2 = 'PRPRD' THEN 'Reason Code Unavailable'
    WHEN C.CARRIER_ACTION_CD2 = 'PRSC' THEN 'Expected a Prescriber Field'
    WHEN C.CARRIER_ACTION_CD2 = 'PRVNC' THEN 'Provider Not Covered'
    WHEN C.CARRIER_ACTION_CD2 = 'PTAGE' THEN 'Patient Outside Age Limit for this Type of Benefit'
    WHEN C.CARRIER_ACTION_CD2 = 'PTINFO' THEN 'Need Information From Patient'
    WHEN C.CARRIER_ACTION_CD2 = 'PTRES' THEN 'Patient Responsibility'
    WHEN C.CARRIER_ACTION_CD2 = 'QUAL' THEN 'Qualifying Procedure Not Received'
    WHEN C.CARRIER_ACTION_CD2 = 'QUANT' THEN 'Quantity Not Covered'
    WHEN C.CARRIER_ACTION_CD2 = 'REBIL' THEN 'Claim Closed to Pass through New Cycle'
    WHEN C.CARRIER_ACTION_CD2 = 'RECAL' THEN 'Reason Code Unavailable'
    WHEN C.CARRIER_ACTION_CD2 = 'RECALL' THEN 'Claims Closed - Completed Client Run Out Phase'
    WHEN C.CARRIER_ACTION_CD2 = 'REFIL' THEN 'Refill Too Soon'
    WHEN C.CARRIER_ACTION_CD2 = 'REFMI' THEN 'Refill Number Missing/Invalid'
    WHEN C.CARRIER_ACTION_CD2 = 'RENPRO' THEN 'Pending Information from Rendering Provider'
    WHEN C.CARRIER_ACTION_CD2 = 'RESENT' THEN 'Resent Claims'
    WHEN C.CARRIER_ACTION_CD2 = 'RESUB' THEN 'Submit Claim to Another Entity'
    WHEN C.CARRIER_ACTION_CD2 = 'REVCOD' THEN 'Reason Code Unavailable'
    WHEN C.CARRIER_ACTION_CD2 = 'RNID' THEN 'Reason Code Unavailable'
    WHEN C.CARRIER_ACTION_CD2 = 'RNNAP' THEN 'Reason Code Unavailable'
    WHEN C.CARRIER_ACTION_CD2 = 'ROMRAT' THEN 'Reason Code Unavailable'
    WHEN C.CARRIER_ACTION_CD2 = 'ROUTN' THEN 'Routine Services Not Covered'
    WHEN C.CARRIER_ACTION_CD2 = 'RXAGE' THEN 'Participant Age Restricts Medication Coverage'
    WHEN C.CARRIER_ACTION_CD2 = 'RXLMT' THEN 'Medication Exceeds Plan Limits'
    WHEN C.CARRIER_ACTION_CD2 = 'RXNC' THEN 'Prescription Drugs Not Covered'
    WHEN C.CARRIER_ACTION_CD2 = 'SETTL' THEN 'Claim Covered in Carrier Settlement'
    WHEN C.CARRIER_ACTION_CD2 = 'SPPBM' THEN 'Must Fill through Specialty Pharmacy'
    WHEN C.CARRIER_ACTION_CD2 = 'SUBID' THEN 'Reason Code Unavailable'
    WHEN C.CARRIER_ACTION_CD2 = 'TERMGRP' THEN 'Reason Code Unavailable'
    WHEN C.CARRIER_ACTION_CD2 = 'TIMELY' THEN 'Claim Past Timely Filing Limit'
    WHEN C.CARRIER_ACTION_CD2 = 'TOSNE' THEN 'Type of Service Not Eligible'
    WHEN C.CARRIER_ACTION_CD2 = 'TPA' THEN 'Send Claim to TPA'
    WHEN C.CARRIER_ACTION_CD2 = 'TPLNC' THEN 'Group Does Not Allow Third Party Claims'
    WHEN C.CARRIER_ACTION_CD2 = 'UMID' THEN 'UNKNOWN'
    WHEN C.CARRIER_ACTION_CD2 = 'UNID' THEN 'Unable to Identify Member'
    WHEN C.CARRIER_ACTION_CD2 = 'UNITS' THEN 'Units Field Invalid for Number of Days'
    WHEN C.CARRIER_ACTION_CD2 = 'UNPRO' THEN 'Unprocessed Claim'
    WHEN C.CARRIER_ACTION_CD2 = 'WASTE' THEN 'Reason Code Unavailable'
    WHEN C.CARRIER_ACTION_CD2 = '' THEN 'Reason Code Unavailable'
    ELSE 'NOCOD'
            END  AS   "CARRIER DESCRIPTION2",
        
        CASE WHEN A.ACTION_CD IN ('1100', '1109', '147C', '2100', '210B', '24OLD', '277ACC', '277REJ', '7U00', 'ABATE', 'ACCDX', 'ADDR', 'ADMIN', 'AETNA', 'AGED', 'AGED1', 'ALCARE', 'ARSRC', 'B2BREJ', 'BABFL', 'BABY', 'BCLCL', 'BILLE', 'BNFAD', 'BNFAS', 'BNFPB', 'CCLM', 'CERT', 'CGCIG', 'CLAMAJ', 'CLEAN', 'CLMFM', 'CLMFW', 'CLMFWD', 'CNTID', 'COBPD', 'CPT', 'D', 'DRSTM', 'DUA', 'DUP', 'DUPHMS', 'EFT', 'ELEC', 'EMPCER', 'EOMB', 'ERRDES', 'ERRDNA', 'ERREOB', 'EXPER', 'FAULT', 'FUTFL', 'GENREJ', 'HCFA', 'HMSA', 'HMSCLF', 'HMSCPT', 'HMSDIG', 'HMSDOS', 'HMSEOB', 'HMSI', 'HMSIC', 'HMSMCV', 'HMSN', 'HMSNC', 'HMSNPI', 'HMSO', 'HMSP', 'HMSPE', 'HMSPND', 'HMSPRO', 'HMSPTP', 'HMSPV', 'HMSR', 'HMSREC', 'HMSRJ', 'HMSSTU', 'HMSTMZ', 'HPSC', 'HRA', 'HSACC', 'IACC', 'IANPI', 'IBNPI', 'ICER', 'ICERT', 'ICLF', 'ICOB', 'ICOU', 'ICPD', 'ICPT', 'IDAYS', 'IDEA', 'IDIG', 'IDOB', 'IDOS', 'IEOB', 'IGRP', 'ILAC', 'ILLEG', 'IMAP', 'IMCV', 'INADD', 'INAME', 'INFO', 'INOPID', 'INPI', 'INPRC', 'INSADD', 'INVAD', 'INVADD', 'INVAM', 'INVBN', 'INVBT', 'INVCD', 'INVDOB', 'INVDOS', 'INVDS', 'INVDT', 'INVDX', 'INVES', 'INVGR', 'INVGRP', 'INVID', 'INVLDADD', 'INVLDCAR', 'INVLDFMT', 'INVLDOF', 'INVMOD', 'INVNAM', 'INVND', 'INVOF', 'INVPCN', 'INVPIN', 'INVPOS', 'IPAD', 'IPCS', 'IPOA', 'IPRO', 'IPRS', 'IPTA', 'IREC', 'IREL', 'IRXST', 'ISNPI', 'ISTU', 'ITAX', 'ITMZ', 'ITPL', 'KEYED', 'LEGAL', 'LIABL', 'LOST', 'LWR21', 'MAGEL', 'MCAID', 'MCPRID', 'MCV', 'MDCLM', 'MEDREC', 'MEDSUP', 'MODIF', 'NAS', 'NCRPT', 'NCRTP', 'NENTAP', 'NENTBP', 'NENTRF', 'NENTRP', 'NENTSF', 'NOAUT', 'NOAUTH', 'NOCOB', 'NOCOD', 'NOINFO', 'NOINT', 'NOPAP', 'NOTES', 'NOXWK', 'NSXWK', 'NWAIT', 'OPN', 'OTNWK', 'OUTNW', 'PAPER', 'PAYCL', 'PBM', 'PBMPD', 'PCPRE', 'PDIA', 'PEND', 'PND', 'PRCINF', 'PREAU', 'PREAUT', 'PRIMID', 'PTADD', 'PTINFO', 'PURCHASED', 'QUAL', 'REDWT', 'RENNPI', 'RENPRO', 'REPRI', 'RESENT', 'RESUB', 'REVCD', 'RFINF', 'RSCINFO', 'SBAPS', 'SPLIT', 'SPPBM', 'STDNT', 'SUBMH', 'SUBNAP', 'SURDT', 'TARS', 'TFLCL', 'TIMEL', 'TIMELY', 'TIMRX', 'TOOTH', 'TPA', 'TPLNC', 'TPLNE', 'UB92', 'UMGRP', 'UMID', 'UNITS', 'UNPRO', 'VALOP', 'ZPDPR', 'ZPDRS' )
        --C.CARRIER_ACTION_CD2 IN ('277ACC', '999', 'ACCDT', 'APID', 'B2BREJ', 'BADCLMS', 'BADEL', 'BPNAP', 'BPNPI', 'BPTAX', 'BPTXN', 'CAID', 'CANCEL', 'CCLM', 'CLAMAJ', 'CLMFRQ', 'CLMFWD', 'CNBC', 'COBPD', 'DAVITA', 'DISDT', 'DISSTT', 'DUP', 'EDITGOV', 'ENDUP', 'ERRDNA', 'ERREOB', 'EXPER', 'GENREJ', 'GRPSTL', 'HCPCS', 'ICD10', 'ICOB', 'IDAYS', 'IDEA', 'IDIG', 'IDOB', 'IDOS', 'IEOB', 'ILAC', 'IMCV', 'INAME', 'INFO', 'INOPID', 'INPI', 'INPRC', 'INSADD', 'INVADD', 'INVBN', 'INVBT', 'INVCD', 'INVDOS', 'INVDT', 'INVDX', 'INVGR', 'INVGRP', 'INVID', 'INVMOD', 'INVNAM', 'INVND', 'INVPCN', 'INVPOS', 'IPRO', 'IREC', 'IREL', 'ITAX', 'ITMZ', 'KEYED', 'LIABL', 'LINE', 'LOWDOL', 'MAXBN', 'MCAID', 'MEDNCOV', 'NENTBP', 'NENTIN', 'NENTPT', 'NENTRF', 'NOAUTH', 'NOCOD', 'NOELG', 'NOTES', 'NOXWK', 'NUBCCD', 'NUBCVL', 'OUTNW', 'PCFDNY', 'PCFVOID', 'PCPRE', 'PDIA', 'PEND', 'PRPRC', 'PRPRD', 'PTINFO', 'QUAL', 'RECAL', 'RENPRO', 'RESENT', 'RESUB', 'REVCOD', 'RNID', 'RNNAP', 'ROMRAT', 'SPPBM', 'SUBID', 'TERMGRP', 'TIMELY', 'TPA', 'TPLNC', 'UMID', 'UNITS', 'UNPRO', 'WASTE' )
        --OR C.CARRIER_ACTION_CD2 IN ('277ACC', '999', 'ACCDT', 'APID', 'B2BREJ', 'BADCLMS', 'BADEL', 'BPNAP', 'BPNPI', 'BPTAX', 'BPTXN', 'CAID', 'CANCEL', 'CCLM', 'CLAMAJ', 'CLMFRQ', 'CLMFWD', 'CNBC', 'COBPD', 'DAVITA', 'DISDT', 'DISSTT', 'DUP', 'EDITGOV', 'ENDUP', 'ERRDNA', 'ERREOB', 'EXPER', 'GENREJ', 'GRPSTL', 'HCPCS', 'ICD10', 'ICOB', 'IDAYS', 'IDEA', 'IDIG', 'IDOB', 'IDOS', 'IEOB', 'ILAC', 'IMCV', 'INAME', 'INFO', 'INOPID', 'INPI', 'INPRC', 'INSADD', 'INVADD', 'INVBN', 'INVBT', 'INVCD', 'INVDOS', 'INVDT', 'INVDX', 'INVGR', 'INVGRP', 'INVID', 'INVMOD', 'INVNAM', 'INVND', 'INVPCN', 'INVPOS', 'IPRO', 'IREC', 'IREL', 'ITAX', 'ITMZ', 'KEYED', 'LIABL', 'LINE', 'LOWDOL', 'MAXBN', 'MCAID', 'MEDNCOV', 'NENTBP', 'NENTIN', 'NENTPT', 'NENTRF', 'NOAUTH', 'NOCOD', 'NOELG', 'NOTES', 'NOXWK', 'NUBCCD', 'NUBCVL', 'OUTNW', 'PCFDNY', 'PCFVOID', 'PCPRE', 'PDIA', 'PEND', 'PRPRC', 'PRPRD', 'PTINFO', 'QUAL', 'RECAL', 'RENPRO', 'RESENT', 'RESUB', 'REVCOD', 'RNID', 'RNNAP', 'ROMRAT', 'SPPBM', 'SUBID', 'TERMGRP', 'TIMELY', 'TPA', 'TPLNC', 'UMID', 'UNITS', 'UNPRO', 'WASTE' )
        THEN 'ACTIONABLE'
        WHEN A.ACTION_CD IN('1', '2', '3', '4', '5', 'ADJ', 'ADJUD', 'AMBORG', 'AMBUL', 'APDNA', 'APDTF', 'APMRB', 'APPDN', 'ARDUP', 'ASSIGN', 'AUTHDN', 'AUTO', 'BADCLMS', 'BANK', 'BENEFITCHG', 'BILLER', 'BILPR', 'BKRPT', 'CANCEL', 'CAPAJ', 'CAPIT', 'CARD', 'CARE', 'CCSOL', 'CCU', 'CDHNE', 'CFLDP', 'CLMAJ', 'CLMFRQ', 'CLMVD', 'CMPMI', 'CMPNC', 'CMPPR', 'CMPPT', 'CNBC', 'CNPHA', 'COINS', 'CONTR', 'COPAY', 'COSMT', 'CREDIT', 'CSFAL', 'CSNID', 'CTOHA', 'CUSTD', 'DAVITA', 'DAWPT', 'DAYLI', 'DAYSU', 'DAYSUP', 'DB2', 'DEDUC', 'DEFER', 'DEFUNCT', 'DENTL', 'DEPNE', 'DEPPG', 'DISAL', 'DISPOL', 'DNYPD', 'DODIN', 'DOSNE', 'DPWMI', 'DRGEX', 'DRGNA', 'DRGNC', 'DRGPG', 'DRGTB', 'DSPMI', 'DSPMX', 'DUPPRO', 'DURCF', 'DURNC', 'DXNC', 'ELGIS', 'EMPN', 'EMPNA', 'EMPNE', 'ENDUP', 'EXDRG', 'EXFREQ', 'EXRCR', 'FACNC', 'FAIL', 'FLEX', 'FRMLRY', 'FRONT', 'FWPXX', 'GENSB', 'GRAND', 'GROSS', 'GRPNE', 'HDNA', 'HFAMILY', 'HMSADJ', 'HMSC', 'HMSD', 'HMSE', 'HMSG', 'HMSH', 'HMSM', 'HMSS', 'HMST', 'HMSU', 'HMSX', 'HMSZ', 'HOSPIC', 'HOSPOL', 'ICNPD', 'ID12M', 'ID24M', 'IDEXP', 'IDOL', 'IFILL', 'INCL', 'INDC', 'INDEM', 'INDN', 'INGEX', 'INGMI', 'INGMX', 'INPNC', 'INVDO', 'INVGN', 'INVNDC', 'INVPL', 'INVPY', 'INVQT', 'INVUC', 'IPA', 'IRGC', 'IRX', 'LIFEPO', 'LMAX', 'LOSC', 'LOWDOL', 'LRDEX', 'LTCNC', 'LTD', 'MAB', 'MAIL', 'MAMCO', 'MAPAD', 'MAREF', 'MATCH', 'MAXBE', 'MAXBEN', 'MAXBN', 'MCCOB', 'MEDAJ', 'MEDNCO', 'MEDNEC', 'MFDNY', 'MGDCR', 'MHNC', 'MMNS', 'MXBEN', 'NABP', 'NBPAT', 'NBPT', 'NDCGP', 'NDCNC', 'NEIC', 'NEICDX', 'NEICPD', 'NENTIN', 'NENTPT', 'NENTSB', 'NLP', 'NOBIL', 'NOBILL', 'NOBRK', 'NOCOL', 'NOCOV', 'NOCVGE', 'NODENT', 'NOELG', 'NOMAT', 'NOMC', 'NOMCA', 'NOMCB', 'NOMED', 'NOMHC', 'NONEW', 'NONURS', 'NOOBES', 'NOPAR', 'NOPAY', 'NOPAYADJ', 'NOPRO', 'NOSUB', 'NOVIS', 'NPHARM', 'NPICN', 'OFFADJ', 'OOP', 'OOPET', 'OTCNC', 'OTHPRO', 'OVDRG', 'OVERP', 'P', 'PARPY', 'PARTB', 'PARTD', 'PAYEE', 'PAYME', 'PAYPR', 'PCFDNY', 'PCFVOID', 'PCUHC', 'PDCR', 'PDPVPB', 'PDSTCLM', 'PDTOP', 'PDTOPT', 'PDTOPV', 'PDTOST', 'PDTOWC', 'PENL', 'PHARN', 'PHARNE', 'PHYNC', 'PIPDY', 'POSTIN', 'POUHC', 'PRCNC', 'PREEX', 'PREVOD', 'PREVOI', 'PREVPD', 'PRICE', 'PRIME', 'PROBL', 'PROD', 'PROVN', 'PROVNC', 'PRSC', 'PRSCAD', 'PRSNC', 'PRVADD', 'PRVNC', 'PTAGE', 'PTINF', 'PTRES', 'PTSTAT', 'PUHC', 'QUANT', 'RDCBN', 'REBDS', 'REBIL', 'RECALL', 'REFIL', 'REFIL0', 'REFIL1', 'REFILL', 'REFMI', 'REFNA', 'RESID', 'RESTR', 'REWRK', 'RIDER', 'ROUTN', 'RSC', 'RSPAPR', 'RXAC', 'RXAGE', 'RXBC', 'RXCUT', 'RXDSG', 'RXEDT', 'RXGEN', 'RXLMT', 'RXNC', 'RXORG', 'RXRCT', 'SCNC', 'SCNOC', 'SEAL', 'SETTL', 'SETTLD', 'SLFINS', 'SNFNC', 'SPEECH', 'SUBST', 'SVLR', 'TERMGRP', 'TOSN', 'TOSNE', 'TRVLNC', 'UNID', 'UNKPT', 'VISION', 'WASTE', 'WCOMP', 'WCOMP', 'WWEST', 'YMAXBN' )
        THEN 'FINAL'
        ELSE NULL
        END AS "DENIAL TYPE",
        
        B.CARRIER_CD AS "HMS CARRIER CODE",
        B.CARRIER_NM AS HMS_CARRIER_NAME,
        A.CAR_CLM_REF_NUM AS "CARRIER DENIAL CLAIM NUMBER",
        A.CARRIER_ACK_DTM AS "CARRIER ACTION DATE",
        TO_CHAR(A.BILL_DT,'MM/DD/YY') AS "ORIGINAL BILL DATE",
        TO_CHAR(A.LAST_BILL_DT, 'MM/DD/YY') AS "REBILL DATE",

    CASE 
    WHEN A.TRANSACT_STATUS_CD = 'D' THEN 'DENIED'
    WHEN A.TRANSACT_STATUS_CD = 'V' THEN 'VOID'
    WHEN A.TRANSACT_STATUS_CD = '8' THEN 'REVERSED'
        ELSE 'OTHER' 
        END AS STATUS,

    CASE WHEN A.ORIG_SRCE_ELIG_CD = 'RS' THEN 'RSC'
    ELSE 'TPL'
    END AS "SOURCE CODE"

    FROM	EDW_AR_FL.ARTBASE A

    LEFT JOIN EDW_AR_FL.ARTCARM B
    ON A.CARRIER_CD = B.CARRIER_CD

    RIGHT OUTER JOIN EDW_AR_FL.DENIALS C
    ON A.CONTRACT_NUM = C.CONTRACT_NUM AND A.AR_SEQ_NUM = C.AR_SEQ_NUM

    WHERE	A.CONTRACT_NUM='478' AND
    /* FOR CCO CLAIMS */  (SUBSTR(A.ICN_NUM,1,2) IN('60', '70'))
        AND CAST(A.REMIT_DT AS DATE) BETWEEN '{first_day_prev_month}' AND '{last_day_prev_month}'  AND B.DEFAULT_IND = 'Y'
        AND A.ACTION_CD NOT IN('AMBORG',  'INPRC', 'INSADD', 'PCFDNY', 'BADEL', 'BDSRCE', 'MATCH', 'NOCOD','A2~20','A1~20','A2~19~P','A1~19~P','A2~20~S','A1~16','A1~19', 'GENREJ', 'ACK', '277ACC', 'LOWDOL', 'REWRK', 'BADEL', '24OLD', 'AGED', 'AGED1', 'ARDUP', 'ARSRC', 'CTOHA', 'DB2, DISAL', 'ELGIS', 'EMPNA', 'FWPXX', 'ICNPD', 'MATCH', 'MFDNY', 'NOBIL', 'NOCOL', 'NOELG', 'NPICN', 'PCFVOID', 'POSTIN', 'PREVOD', 'REBDS', 'REBIL', 'RECALL', 'RESENT', 'RSC', 'RSCINFO', 'SETTL', 'SETTLD', 'ZPDPR', 'ZPDRS') 
        AND A.TRANSACT_STATUS_CD IN ('D','8','V')
        GROUP BY 1,2,3,4,5,6,7,8,9,10,11,12,14,15,16,17,18,19,20,21,22,23,24,25, 26, 27, 28
    """

    cursor.execute(teradata_query2)
    columns4 = [column[0] for column in cursor.description]
    results4 = cursor.fetchall()
    results4 = [list(row) for row in results4] 
    print("Results from CCO Denial query:")
    for row in results4:
        print(row)

    # Export FFS Denied results to Excel and TXT
    df4 = pd.DataFrame(results4, columns=columns4)

    # Save to Excel with borders and without bold column names
    with pd.ExcelWriter(cco_denied_excel_filename, engine='openpyxl') as writer:
        df4.to_excel(writer, index=False)
        workbook  = writer.book
        worksheet = writer.sheets['Sheet1']
        add_borders(worksheet)
        
        # Align column names to the left and remove bold formatting
        for cell in worksheet["1:1"]:
            cell.alignment = Alignment(horizontal="left")
            cell.font = Font(bold=False, name="Aptos Narrow")
        
        # Format columns as dates or currency and apply font to all cells
        format_columns_and_apply_font(worksheet)

    # Save to TXT with commas and quotes around each row
    with open(cco_denied_txt_filename, 'w') as f:
        f.write('"' + '","'.join(columns4) + '"\n')
        for row in results4:
            f.write('"' + '","'.join(map(str, row)) + '"\n')

    print("CCO Denied Files Successfully Exported")

except pyodbc.DatabaseError as e:
    print("Database connection error:", e)
finally:
    if cursor:
        cursor.close()
    if connection:
        connection.close()
    print("Connection closed.")

# List of files to attach
attachments = [
    ffs_excel_filename,
    cco_excel_filename,
    ffs_txt_filename,
    cco_txt_filename,
    ffs_denied_excel_filename,
    cco_denied_excel_filename,
    ffs_denied_txt_filename,
    cco_denied_txt_filename
]

# Create an Outlook Email
outlook = win32.Dispatch("Outlook.Application")
mail = outlook.CreateItem(0)

# Set Email Details
mail.To = "momentum@hms.com"
mail.CC = "paige.wall@hms.com"
mail.Subject = "sendORMTHLY"
mail.BodyFormat = 1
mail.Body = ("send")

# Attach Files
for file in attachments:
    mail.Attachments.Add(file)

# Send Email
mail.Send()

print("Email Sent Successfully")